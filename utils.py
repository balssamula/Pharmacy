import streamlit as st
import pandas as pd
import io
import os
import requests
import json
import logging
import re
from datetime import datetime
from typing import Optional, List, Dict, Any
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SALLA_API_URL = "https://api.salla.dev/admin/v2/specialoffers"

OFFER_TYPES_MAP = {
    "buy_x_get_y": "اذا اشترى العميل X يحصل على Y",
    "fixed_amount": "مبلغ ثابت من قيمة مشتريات العميل",
    "percentage": "نسبة من قيمة مشتريات العميل",
    "discounts_table": "جدول الخصومات",
    "special_price": "سعر ثابت",
    "tiered_offer": "عرض الفئات"
}
REV_OFFER_TYPES_MAP = {v: k for k, v in OFFER_TYPES_MAP.items()}

CHANNELS_MAP = {
    "browser": "متصفح المتجر",
    "app": "تطبيق المتجر",
    "browser_and_application": "متصفح وتطبيق المتجر",
    "pos": "سلة بوينت"
}
REV_CHANNELS_MAP = {v: k for k, v in CHANNELS_MAP.items()}

APPLIED_TO_MAP = {
    "all": "جميع المنتجات",
    "product": "منتجات مختارة",
    "category": "تصنيفات مختارة",
    "paymentMethod": "طرق دفع مختارة",
    "brand": "علامات تجارية مختارة",
    "tag": "وسوم مختارة"
}
REV_APPLIED_TO_MAP = {v: k for k, v in APPLIED_TO_MAP.items()}

# ==========================================
# 🛠️ دوال الأساسيات والربط
# ==========================================

def get_headers():
    token = st.session_state.get('access_token', '')
    if not token:
        st.warning("⚠️ الرجاء إدخال مفتاح الربط (Access Token)")
        return None
    return {"Authorization": f"Bearer {token}"}

def safe_float(val: Any, default: float = 0.0) -> float:
    if val is None: return default
    try: return float(val)
    except (ValueError, TypeError): return default

def safe_parse_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str: return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%a %b %d %Y %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
        try:
            clean_str = re.sub(r' GMT.*$', '', str(date_str)).replace('T', ' ')
            return datetime.strptime(clean_str[:19], fmt.replace('T', ' '))
        except (ValueError, TypeError): pass
    return None

def parse_products_cleanly(offer_section: Dict) -> str:
    if not offer_section or not isinstance(offer_section, dict):
        return "جميع الأصناف المشمولة"
    clean_elements = []
    products = offer_section.get('products', [])
    if products and isinstance(products, list):
        for p in products:
            if isinstance(p, dict):
                clean_elements.append(f"• صنف: {p.get('name', 'غير معرف')} [ID: {p.get('id', 'N/A')}] [SKU: {p.get('sku', 'N/A')}]")
            else:
                clean_elements.append(f"• معرف منتج رقم: {p}")
    categories = offer_section.get('categories', [])
    if categories and isinstance(categories, list):
        for c in categories:
            if isinstance(c, dict):
                clean_elements.append(f"• تصنيف: {c.get('name', 'غير معرف')} [ID: {c.get('id', 'N/A')}]")
            else:
                clean_elements.append(f"• معرف تصنيف رقم: {c}")
    return "\n".join(clean_elements) if clean_elements else "جميع الأصناف المشمولة"

def get_flat_price(price_field: Any) -> float:
    if not price_field: return 0.0
    if isinstance(price_field, dict):
        return safe_float(price_field.get('amount', 0.0))
    return safe_float(price_field)

def safe_api_request(method: str, url: str, headers: Dict, **kwargs) -> Optional[Dict]:
    try:
        if 'json' in kwargs:
            headers = headers.copy()
            headers['Content-Type'] = 'application/json; charset=utf-8'
            
        response = requests.request(method, url, headers=headers, timeout=30, **kwargs)
        if response.status_code >= 400:
            if response.status_code != 404:
                try: error_detail = json.dumps(response.json(), ensure_ascii=False)
                except: error_detail = response.text[:500]
                st.error(f"⚠️ خطأ {response.status_code}: {error_detail}")
            return None
        return response.json()
    except Exception as e:
        st.error(f"⚠️ خطأ في الاتصال: {str(e)}")
        return None

def style_excel_file(ws, is_template=True, header_color="0F1C2E"):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), 
                         top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 22

    if ws.title in ["Salla Offers Template", "Salla Offers", "مسودة العروض"]:
        validations = {
            "A": '"إنشاء,تحديث,حذف"', "D": f'"{",".join(OFFER_TYPES_MAP.values())}"',
            "E": f'"{",".join(CHANNELS_MAP.values())}"', "F": f'"{",".join(APPLIED_TO_MAP.values())}"',
            "I": '"نعم,لا"', "N": '"منتج,تصنيف,ماركة"', "Q": '"منتج,تصنيف,ماركة"',
            "T": '"منتج مجاني,خصم بنسبة,مبلغ ثابت"', "W": '"نشط,غير نشط"'
        }
        for col_letter, formula1 in validations.items():
            if col_letter <= get_column_letter(ws.max_column):
                try:
                    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}1000")
                except: pass

def prepare_import_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    import_df = df.copy()
    column_mapping = {
        'معرف المنتج': 'id', 'SKU': 'sku', 'اسم المنتج': 'name', 'النوع': 'type',
        'نوع المنتج': 'product_type', 'حالة المنتج': 'status', 'السعر (SAR)': 'price',
        'السعر المخفض (SAR)': 'sale_price', 'بداية التخفيض': 'sale_start',
        'نهاية التخفيض': 'sale_end', 'كمية غير محدودة': 'unlimited_quantity',
        'خاضع للضريبة': 'with_tax', 'سبب عدم الخضوع': 'tax_reason_code',
        'العنوان الترويجي': 'promotion_title', 'العنوان الفرعي': 'promotion_subtitle'
    }
    import_df = import_df.rename(columns=column_mapping)
    
    product_type_mapping = {'منتج جاهز': 'product', 'مجموعة منتجات': 'group_products', 'بطاقة رقمية': 'codes', 'منتج رقمي': 'digital', 'أكل': 'food', 'خدمة حسب الطلب': 'service', 'منتج حجز': 'booking'}
    if 'product_type' in import_df.columns: import_df['product_type'] = import_df['product_type'].map(product_type_mapping).fillna('product')
    
    status_mapping = {'معروض': 'sale', 'مخفي': 'hidden'}
    if 'status' in import_df.columns: import_df['status'] = import_df['status'].map(status_mapping).fillna('sale')
    
    tax_mapping = {'نعم': 'true', 'لا': 'false'}
    if 'with_tax' in import_df.columns: import_df['with_tax'] = import_df['with_tax'].map(tax_mapping).fillna('true')
    
    unlimited_mapping = {'نعم': 'true', 'لا': 'false'}
    if 'unlimited_quantity' in import_df.columns: import_df['unlimited_quantity'] = import_df['unlimited_quantity'].map(unlimited_mapping).fillna('false')
    
    import_df = import_df.fillna('')
    return import_df

# ==========================================
# 🎁 دوال العروض الخاصة (محسنة لخطأ 422)
# ==========================================

def generate_salla_excel_template() -> bytes:
    from openpyxl import Workbook
    headers = [
        "الإجراء", "معرف العرض", "اسم العرض", "نوع العرض", "المنصة", "تطبيق على", "تاريخ البدء", "تاريخ الانتهاء", 
        "تطبيق مع كوبون", "الحد الأقصى للخصم", "الحد الأدنى للشراء", "الحد الأدنى للكمية", 
        "مجموعات العملاء", "نوع شراء X", "كمية شراء X", "عناصر شراء X (IDs)", 
        "نوع عرض Y", "كمية عرض Y", "عناصر عرض Y (IDs)", "نوع الخصم", "قيمة الخصم", 
        "رسالة العرض", "حالة العرض"
    ]
    example_row = [
        "إنشاء", "", "عرض الشتاء المميز", "اذا اشترى العميل X يحصل على Y", "متصفح وتطبيق المتجر", 
        "منتجات مختارة", "2026-08-01 00:00:00", "2026-08-30 23:59:59", "لا", 100, 50, 1, 
        "", "منتج", 1, "1234,5678", "منتج", 1, "91011", "خصم بنسبة", 50, "تسوق الآن واستمتع!", "نشط"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Offers Template"
    ws.append(headers)
    ws.append(example_row)
    style_excel_file(ws, is_template=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def export_offers_to_excel(offers_list: List[Dict]) -> bytes:
    try:
        from openpyxl import Workbook
        headers = [
            "الإجراء", "معرف العرض", "اسم العرض", "نوع العرض", "المنصة", "تطبيق على", "تاريخ البدء", "تاريخ الانتهاء", 
            "تطبيق مع كوبون", "الحد الأقصى للخصم", "الحد الأدنى للشراء", "الحد الأدنى للكمية", 
            "مجموعات العملاء", "نوع شراء X", "كمية شراء X", "عناصر شراء X (IDs)", 
            "نوع عرض Y", "كمية عرض Y", "عناصر عرض Y (IDs)", "نوع الخصم", "قيمة الخصم", 
            "رسالة العرض", "حالة العرض"
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Salla Offers"
        ws.append(headers)

        for o in offers_list:
            buy = o.get('buy', {})
            get = o.get('get', {})
            
            b_type = buy.get('type', 'product')
            if isinstance(b_type, dict): b_type = b_type.get('id', 'product')
            buy_type_ar = {"product": "منتج", "category": "تصنيف", "brand": "ماركة"}.get(b_type, "منتج")
            
            buy_elems = []
            if b_type == 'product': buy_elems = [str(p.get('id', p) if isinstance(p, dict) else p) for p in buy.get('products', [])]
            elif b_type == 'category': buy_elems = [str(c.get('id', c) if isinstance(c, dict) else c) for c in buy.get('categories', [])]
            elif b_type == 'brand': buy_elems = [str(b.get('id', b) if isinstance(b, dict) else b) for b in buy.get('brands', [])]
            
            g_type = get.get('type', 'product')
            if isinstance(g_type, dict): g_type = g_type.get('id', 'product')
            get_type_ar = {"product": "منتج", "category": "تصنيف", "brand": "ماركة"}.get(g_type, "منتج")
            
            get_elems = []
            if g_type == 'product': get_elems = [str(p.get('id', p) if isinstance(p, dict) else p) for p in get.get('products', [])]
            elif g_type == 'category': get_elems = [str(c.get('id', c) if isinstance(c, dict) else c) for c in get.get('categories', [])]
            elif g_type == 'brand': get_elems = [str(b.get('id', b) if isinstance(b, dict) else b) for b in get.get('brands', [])]
            
            disc_type_id = get.get('discount_type', 'percentage')
            if isinstance(disc_type_id, dict): disc_type_id = disc_type_id.get('id', 'percentage')
            disc_type_ar = "منتج مجاني" if disc_type_id == "free-product" else ("خصم بنسبة" if disc_type_id == "percentage" else "مبلغ ثابت")
            
            cust_groups = [str(g.get('id', g) if isinstance(g, dict) else g) for g in o.get('customer_groups', [])]
            o_type_id = o.get('offer_type', '')
            chan_id = o.get('applied_channel', '')
            app_to_id = o.get('applied_to', '')
            
            row = [
                "تحديث", o.get('id', ''), o.get('name', ''), OFFER_TYPES_MAP.get(o_type_id, o_type_id),
                CHANNELS_MAP.get(chan_id, chan_id), APPLIED_TO_MAP.get(app_to_id, app_to_id),
                o.get('start_date', ''), o.get('expiry_date', ''), "نعم" if o.get('applied_with_coupon') else "لا",
                o.get('max_discount_amount', 0), o.get('min_purchase_amount', 0), o.get('min_items_count', 0),
                ",".join(cust_groups), buy_type_ar, buy.get('quantity', 1), ",".join(buy_elems),
                get_type_ar, get.get('quantity', 1), ",".join(get_elems), disc_type_ar,
                get.get('discount_amount', 0), o.get('message', ''), "نشط" if o.get('status') == 'active' else "غير نشط"
            ]
            ws.append(row)
            
        style_excel_file(ws, is_template=False)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except Exception as e:
        return b""

def process_excel_import(df) -> dict:
    """معالجة ورفع العروض وإصلاح متطلبات سلة الصارمة (422)"""
    results = {"success": [], "errors": []}
    headers = get_headers()
    if not headers:
        results["errors"].append("فشل في المصادقة.")
        return results

    rev_offer_types = {v: k for k, v in OFFER_TYPES_MAP.items()}
    rev_channels = {v: k for k, v in CHANNELS_MAP.items()}
    rev_applied_to = {v: k for k, v in APPLIED_TO_MAP.items()}
    
    for idx, row in df.iterrows():
        try:
            action = str(row.get('الإجراء', '')).strip()
            offer_id = str(row.get('معرف العرض', '')).strip() if pd.notna(row.get('معرف العرض')) else ''
            
            if action == 'حذف':
                if offer_id and offer_id != 'nan':
                    res = safe_api_request("DELETE", f"{SALLA_API_URL}/{offer_id}", headers)
                    if res is not None: results["success"].append(f"تم حذف العرض بنجاح.")
                    else: results["errors"].append(f"فشل حذف العرض رقم {offer_id}")
                else:
                    results["errors"].append(f"السطر {idx+1}: لا يمكن حذف العرض بدون معرف.")
                continue
                
            o_type = rev_offer_types.get(str(row.get('نوع العرض', '')).strip(), 'buy_x_get_y')
            chan = rev_channels.get(str(row.get('المنصة', '')).strip(), 'browser')
            app_to = rev_applied_to.get(str(row.get('تطبيق على', '')).strip(), 'all')
            with_coupon = str(row.get('تطبيق مع كوبون', '')).strip() == 'نعم'
            
            b_type = {"منتج": "product", "تصنيف": "category", "ماركة": "brand"}.get(str(row.get('نوع شراء X', '')).strip(), "product")
            b_elems_raw = str(row.get('عناصر شراء X (IDs)', '')).strip()
            b_elems = [int(i.strip()) for i in b_elems_raw.split(',')] if b_elems_raw and b_elems_raw != 'nan' else []
            
            g_type = {"منتج": "product", "تصنيف": "category", "ماركة": "brand"}.get(str(row.get('نوع عرض Y', '')).strip(), "product")
            g_elems_raw = str(row.get('عناصر عرض Y (IDs)', '')).strip()
            g_elems = [int(i.strip()) for i in g_elems_raw.split(',')] if g_elems_raw and g_elems_raw != 'nan' else []
            
            disc_type_ar = str(row.get('نوع الخصم', '')).strip()
            disc_type = "free-product" if disc_type_ar == "منتج مجاني" else ("percentage" if disc_type_ar == "خصم بنسبة" else "fixed_amount")
            
            # ✅ إصلاح خطأ 422: منتج مجاني يجب أن يرسل له الخصم 100 إجبارياً
            if disc_type == "free-product":
                discount_amount = 100.0
            else:
                discount_amount = float(row.get('قيمة الخصم', 0) or 0)
            
            cg_raw = str(row.get('مجموعات العملاء', '')).strip()
            c_groups = [int(g.strip()) for g in cg_raw.split(',')] if cg_raw and cg_raw != 'nan' else []
            status = "active" if str(row.get('حالة العرض', '')).strip() == "نشط" else "inactive"
            
            payload = {
                "name": str(row.get('اسم العرض', 'عرض')), 
                "offer_type": o_type, 
                "applied_channel": chan,
                "applied_to": app_to, 
                "start_date": str(row.get('تاريخ البدء', '')).strip(), 
                "expiry_date": str(row.get('تاريخ الانتهاء', '')).strip(),
                "status": status, 
                "applied_with_coupon": with_coupon, 
                "max_discount_amount": float(row.get('الحد الأقصى للخصم', 0) or 0),
                "min_purchase_amount": float(row.get('الحد الأدنى للشراء', 0) or 0), 
                "min_items_count": int(row.get('الحد الأدنى للكمية', 0) or 0),
                "message": str(row.get('رسالة العرض', '')).strip(),
                "buy": {"type": b_type, "quantity": int(row.get('كمية شراء X', 1) or 1)},
                "get": {
                    "type": g_type, 
                    "quantity": int(row.get('كمية عرض Y', 1) or 1), 
                    "discount_type": disc_type, 
                    "discount_amount": discount_amount
                }
            }
            if c_groups: payload["customer_groups"] = c_groups
            if b_type == 'product' and b_elems: payload["buy"]["products"] = b_elems
            elif b_type == 'category' and b_elems: payload["buy"]["categories"] = b_elems
            elif b_type == 'brand' and b_elems: payload["buy"]["brands"] = b_elems
            
            if g_type == 'product' and g_elems: payload["get"]["products"] = g_elems
            elif g_type == 'category' and g_elems: payload["get"]["categories"] = g_elems
            elif g_type == 'brand' and g_elems: payload["get"]["brands"] = g_elems
            
            if pd.isna(payload['message']) or payload['message'] == 'nan': payload['message'] = ''
            
            if action == 'تحديث' and offer_id and offer_id != 'nan':
                res = safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json=payload)
                if res: results["success"].append(f"تم تحديث العرض: {payload['name']}")
                else: results["errors"].append(f"فشل تحديث العرض: {payload['name']}")
            elif action == 'إنشاء':
                res = safe_api_request("POST", SALLA_API_URL, headers, json=payload)
                if res: results["success"].append(f"تم إنشاء العرض: {payload['name']}")
                else: results["errors"].append(f"فشل إنشاء العرض: {payload['name']}")
        except Exception as e:
            results["errors"].append(f"خطأ في الصف {idx+1}: {str(e)}")
    return results

# ==========================================
# 📦 دوال المنتجات والصور والضرائب
# ==========================================

def update_product_status(product_id: int, status: str) -> bool:
    headers = get_headers()
    url = f"https://api.salla.dev/admin/v2/products/{product_id}/status"
    res = safe_api_request("POST", url, headers, json={"status": status})
    return res is not None

def export_products_to_excel(products: List[Dict]) -> bytes:
    try:
        data = []
        for p in products:
            price = get_flat_price(p.get('price', 0))
            sale_price = get_flat_price(p.get('sale_price', 0))
            regular_price = get_flat_price(p.get('regular_price', 0))
            base_price = regular_price if regular_price > 0 else price
            promo = p.get('promotion', {})
            promo_title = p.get('promotion_title') or (promo.get('title') if isinstance(promo, dict) else '') or "لا يوجد"
            promo_sub = (promo.get('sub_title') if isinstance(promo, dict) else '') or "لا يوجد"
            sale_start = p.get('sale_start') or (p.get('sale_price', {}).get('start_at') if isinstance(p.get('sale_price'), dict) else None) or "غير محدد"
            sale_end = p.get('sale_end') or (p.get('sale_price', {}).get('expired_at') if isinstance(p.get('sale_price'), dict) else None) or "غير محدد"
            data.append({
                'المعرف': p.get('id', ''), 'الاسم': p.get('name', ''), 'SKU': p.get('sku', ''),
                'السعر الأساسي الأصل': base_price, 'السعر المخفض الحالي': sale_price if sale_price > 0 else 'لا يوجد',
                'خاضع للضريبة': 'نعم' if p.get('with_tax', True) else 'لا',
                'العنوان الترويجي': promo_title, 'العنوان الفرعي': promo_sub,
                'تاريخ بداية التخفيض': sale_start, 'تاريخ نهاية التخفيض': sale_end,
                'المخزون': p.get('quantity', 0), 'المبيعات': p.get('sold_quantity', 0),
                'الحالة': 'معروض' if p.get('status') == 'sale' else 'مخفي'
            })
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            style_excel_file(writer.sheets['Sheet1'], is_template=False, header_color="0F1C2E")
        return buffer.getvalue()
    except Exception as e:
        return b""

def fill_salla_template(products: List[Dict], template_path: str = "Salla_Products_Template.xlsx") -> bytes:
    try:
        import os
        import openpyxl
        import io
        if not os.path.exists(template_path): return b""
        wb = openpyxl.load_workbook(template_path)
        ws = wb["Salla Products Template Sheet"]
        if ws.max_row >= 3: ws.delete_rows(3, ws.max_row - 2)
        start_row = 3 
        for i, p in enumerate(products):
            current_row = start_row + i
            price = get_flat_price(p.get('price', 0))
            sale_price = get_flat_price(p.get('sale_price', 0))
            promo_obj = p.get('promotion', {})
            promo_title = p.get('promotion_title') or (promo_obj.get('title') if isinstance(promo_obj, dict) else '')
            sale_start = p.get('sale_start') or (p.get('sale_price', {}).get('start_at') if isinstance(p.get('sale_price'), dict) else None)
            sale_end = p.get('sale_end') or (p.get('sale_price', {}).get('expired_at') if isinstance(p.get('sale_price'), dict) else None)
            if sale_start and isinstance(sale_start, str): sale_start = sale_start[:10]
            if sale_end and isinstance(sale_end, str): sale_end = sale_end[:10]
            p_id = p.get('id')
            ws.cell(row=current_row, column=1).value = p_id if p_id else None
            ws.cell(row=current_row, column=2).value = 'منتج'
            ws.cell(row=current_row, column=3).value = p.get('name') or 'بدون اسم'
            ws.cell(row=current_row, column=7).value = 'منتج جاهز'
            ws.cell(row=current_row, column=8).value = price if price > 0 else 0
            ws.cell(row=current_row, column=10).value = 'نعم'
            p_sku = p.get('sku')
            ws.cell(row=current_row, column=11).value = p_sku if p_sku else None
            ws.cell(row=current_row, column=13).value = sale_price if sale_price > 0 else None
            ws.cell(row=current_row, column=14).value = sale_start if sale_start else None
            ws.cell(row=current_row, column=15).value = sale_end if sale_end else None
            ws.cell(row=current_row, column=19).value = 1
            ws.cell(row=current_row, column=20).value = 'kg'
            ws.cell(row=current_row, column=21).value = 'متاح' if p.get('status', 'sale') == 'sale' else 'مخفي'
            ws.cell(row=current_row, column=23).value = promo_title if promo_title else None
            ws.cell(row=current_row, column=29).value = 'نعم' if p.get('with_tax', True) else 'لا'
            if not p.get('with_tax', True): ws.cell(row=current_row, column=30).value = p.get('tax_exemption_cause') or 'الأدوية والمعدات الطبية'
            else: ws.cell(row=current_row, column=30).value = None
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e: return b""

def attach_product_image_api(product_id: int, image_bytes: bytes=None, filename: str=None, image_url: str=None) -> bool:
    headers = get_headers()
    url = f"https://api.salla.dev/admin/v2/products/{product_id}/images"
    try:
        if image_url:
            headers["Content-Type"] = "application/json"
            response = requests.post(url, headers=headers, json={"original": image_url}, timeout=30)
        elif image_bytes and filename:
            files = {'photo': (filename, image_bytes, 'image/jpeg')}
            response = requests.post(url, headers=headers, files=files, timeout=30)
        else: return False
        return response.status_code < 400
    except: return False

def update_product_promotions_secure(product_id: int, new_promo: str, new_sub: str, headers: dict) -> bool:
    current_res = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current_res or not current_res.get('data'): return False
    p_data = current_res['data']
    price_val = get_flat_price(p_data.get('price', 0))
    sale_val = get_flat_price(p_data.get('sale_price', 0))
    regular_val = get_flat_price(p_data.get('regular_price', 0))
    base_price = regular_val if regular_val > 0 else price_val
    
    payload = {"name": p_data.get('name'), "price": base_price, "status": p_data.get('status', 'sale')}
    if new_promo is not None: payload["promotion_title"] = new_promo
    if new_sub is not None: payload["promotion_subtitle"] = new_sub; payload["subtitle"] = new_sub
    if sale_val > 0: payload['sale_price'] = sale_val
    
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{product_id}", headers, json=payload)
    return res is not None

def update_product_tax_secure(product_id: int, with_tax: bool, tax_cause: str, headers: dict) -> bool:
    current_res = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current_res or not current_res.get('data'): return False
    p_data = current_res['data']
    base_price = get_flat_price(p_data.get('regular_price', 0)) or get_flat_price(p_data.get('price', 0))
    payload = {"name": p_data.get('name'), "price": base_price, "with_tax": with_tax}
    sale_val = get_flat_price(p_data.get('sale_price', 0))
    if sale_val > 0: payload['sale_price'] = sale_val
    if not with_tax and tax_cause: payload["tax_exemption_cause"] = tax_cause
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{product_id}", headers, json=payload)
    return res is not None

def get_branches_list() -> List[Dict]:
    headers = get_headers()
    res = safe_api_request("GET", "https://api.salla.dev/admin/v2/branches", headers)
    return res.get("data", []) if res else []

def generate_quantities_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "تحديث الكميات"
    ws.append(["💡 إرشادات: ادخل رقم الـمنتج، ومعرف الفرع، والكمية، ثم اختر نوع العملية."])
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 24
    ws.append(["Product_SKU", "Branch_ID", "Quantity", "Mode"])
    style_excel_file(ws, is_template=True, header_color="00EBCF")
    dv_mode = DataValidation(type="list", formula1='"increment,decrement,overwrite"', allow_blank=False)
    ws.add_data_validation(dv_mode); dv_mode.add("D3:D1000")
    wb.save(output)
    return output.getvalue()

def process_quantities_import(df: pd.DataFrame) -> Dict:
    results = {"success": [], "errors": []}
    headers = get_headers()
    quantities_payload = []
    for idx, row in df.iterrows():
        if row.isna().all() or str(row.iloc[0]).strip().startswith("💡"): continue
        sku = str(row.get('Product_SKU', '')).strip()
        branch_id = row.get('Branch_ID')
        quantity = int(safe_float(row.get('Quantity', 0)))
        mode = str(row.get('Mode', 'increment')).strip()
        if sku and pd.notna(branch_id):
            quantities_payload.append({"identifer": sku, "identifer_type": "sku", "branch_id": int(float(branch_id)), "quantity": quantity, "mode": mode})
            
    if quantities_payload:
        res = safe_api_request("POST", "https://api.salla.dev/admin/v2/products/quantities/bulk", headers, json={"products": quantities_payload})
        if res: results["success"].append(f"✅ تم تحديث كميات {len(quantities_payload)} سجل بنجاح!")
        else: results["errors"].append("❌ فشل إرسال طلب التحديث الجماعي.")
    return results

def delete_product(product_id: int) -> bool:
    headers = get_headers()
    res = safe_api_request("DELETE", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    return res is not None

def update_product_price(product_id: int, new_price: float) -> bool:
    headers = get_headers()
    current_res = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current_res or not current_res.get('data'): return False
    p_data = current_res['data']
    payload = {"name": p_data.get('name'), "price": new_price, "status": p_data.get('status', 'sale')}
    sale_price = get_flat_price(p_data.get('sale_price', 0))
    if sale_price > 0 and sale_price < new_price: payload['sale_price'] = sale_price
    elif sale_price >= new_price: payload['sale_price'] = None
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{product_id}", headers, json=payload)
    return res is not None

def update_product_sale_price(product_id: int, sale_price: float, sale_start: str = None, sale_end: str = None) -> bool:
    headers = get_headers()
    current_res = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current_res or not current_res.get('data'): return False
    p_data = current_res['data']
    current_price = get_flat_price(p_data.get('price', 0))
    
    if sale_price > 0 and sale_price >= current_price:
        st.error(f"⚠️ السعر المخفض يجب أن يكون أقل من السعر الأصلي")
        return False
    
    payload = {"name": p_data.get('name'), "price": current_price, "status": p_data.get('status', 'sale')}
    if sale_price > 0:
        payload['sale_price'] = sale_price
        if sale_start: payload['sale_start'] = sale_start
        if sale_end: payload['sale_end'] = sale_end
    else:
        payload['sale_price'] = None
        
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{product_id}", headers, json=payload)
    return res is not None

# ==========================================
# 📦 إدارة المجموعات (معالجة خطأ 422 للكميات)
# ==========================================

def get_product_details(product_id: int) -> Optional[Dict]:
    headers = get_headers()
    res = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    return res['data'] if res and res.get('data') else None

def get_group_products(product_id: int) -> List[Dict]:
    headers = get_headers()
    product = get_product_details(product_id)
    if not product or product.get('type') != 'group_products': return []
    
    group_products = []
    if product.get('consisted_products'):
        for item in product['consisted_products']:
            group_products.append({
                'id': item.get('id'), 'name': item.get('name', 'بدون اسم'), 'sku': item.get('sku', 'لا يوجد'),
                'price': get_flat_price(item.get('price', 0)), 'bundle_quantity': item.get('quantity_in_group', 1),
                'stock_quantity': item.get('quantity', 0), 'status': item.get('status', 'sale'),
                'image': item.get('thumbnail') or item.get('main_image'), 'url': item.get('url'),
                'with_tax': item.get('with_tax', True)
            })
    return group_products

def _get_clean_grouped_items(parent_data: dict) -> list:
    """مصفاة ذكية تستخرج العناصر بصيغة grouped_items النظيفة التي تقبلها سلة في طلبات PUT لمنع تخريب المخزون"""
    clean_items = []
    if parent_data.get('consisted_products'):
        for item in parent_data['consisted_products']:
            p_id = item.get('id')
            qty = item.get('quantity_in_group', 1)
            if p_id: clean_items.append({"product_id": p_id, "quantity": qty})
    elif parent_data.get('grouped_items'):
        for item in parent_data['grouped_items']:
            p_id = item.get('product_id') or item.get('product', {}).get('id')
            qty = item.get('quantity', 1)
            if p_id: clean_items.append({"product_id": p_id, "quantity": qty})
    elif parent_data.get('skus'):
        for item in parent_data['skus']:
            p_id = item.get('id')
            qty = item.get('quantity', 1)
            if p_id: clean_items.append({"product_id": p_id, "quantity": qty})
    return clean_items

def update_group_product_quantity(parent_product_id: int, child_product_id: int, new_quantity: int) -> bool:
    """✅ دالة جراحية لمنع خطأ 422 عند تحديث الكميات في مجموعة المنتجات"""
    headers = get_headers()
    parent = get_product_details(parent_product_id)
    if not parent: return False

    clean_items = _get_clean_grouped_items(parent)
    updated = False

    for item in clean_items:
        if str(item['product_id']) == str(child_product_id):
            item['quantity'] = new_quantity
            updated = True
            break

    if not updated: return False

    payload = {
        "name": parent.get('name'),
        "type": "group_products",
        "grouped_items": clean_items
    }

    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{parent_product_id}", headers, json=payload)
    return res is not None

def remove_product_from_group(parent_product_id: int, child_product_id: int) -> bool:
    headers = get_headers()
    parent = get_product_details(parent_product_id)
    if not parent: return False
    
    clean_items = _get_clean_grouped_items(parent)
    new_items = [item for item in clean_items if str(item['product_id']) != str(child_product_id)]
    
    if len(new_items) == len(clean_items): return False
            
    payload = {
        "name": parent.get('name'),
        "type": "group_products",
        "grouped_items": new_items
    }
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{parent_product_id}", headers, json=payload)
    return res is not None

def add_product_to_group(parent_product_id: int, child_product_id: int) -> bool:
    headers = get_headers()
    parent = get_product_details(parent_product_id)
    if not parent: return False
    
    clean_items = _get_clean_grouped_items(parent)
    for item in clean_items:
        if str(item['product_id']) == str(child_product_id): return False
        
    clean_items.append({"product_id": int(child_product_id), "quantity": 1})
    payload = {
        "name": parent.get('name'),
        "type": "group_products",
        "grouped_items": clean_items
    }
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/products/{parent_product_id}", headers, json=payload)
    return res is not None

# ==========================================
# 👥 دوال العملاء والمجموعات (تم استعادتها)
# ==========================================

def get_customers_list(keyword: str = "") -> Optional[Dict]:
    headers = get_headers()
    if not headers: return None
    url = "https://api.salla.dev/admin/v2/customers"
    params = {"keyword": keyword} if keyword else {}
    return safe_api_request("GET", url, headers, params=params)

def create_customer(customer_data: Dict) -> bool:
    headers = get_headers()
    res = safe_api_request("POST", "https://api.salla.dev/admin/v2/customers", headers, json=customer_data)
    return res is not None

def update_customer_api(customer_id: int, customer_data: Dict) -> bool:
    headers = get_headers()
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/customers/{customer_id}", headers, json=customer_data)
    return res is not None

def delete_customer_api(customer_id: int) -> bool:
    headers = get_headers()
    res = safe_api_request("DELETE", f"https://api.salla.dev/admin/v2/customers/{customer_id}", headers)
    return res is not None

def get_customer_groups_list() -> Optional[Dict]:
    headers = get_headers()
    return safe_api_request("GET", "https://api.salla.dev/admin/v2/customers/groups", headers)

def create_customer_group(group_data: Dict) -> bool:
    headers = get_headers()
    res = safe_api_request("POST", "https://api.salla.dev/admin/v2/customers/groups", headers, json=group_data)
    return res is not None

def update_customer_group_api(group_id: int, group_data: Dict) -> bool:
    headers = get_headers()
    res = safe_api_request("PUT", f"https://api.salla.dev/admin/v2/customers/groups/{group_id}", headers, json=group_data)
    return res is not None

def delete_customer_group_api(group_id: int) -> bool:
    headers = get_headers()
    res = safe_api_request("DELETE", f"https://api.salla.dev/admin/v2/customers/groups/{group_id}", headers)
    return res is not None

def export_customers_to_excel(customers: List[Dict]) -> bytes:
    try:
        data = []
        for cust in customers:
            stats = cust.get('stats', {})
            orders_count = stats.get('orders_count', 0) if isinstance(stats, dict) else 0
            orders_amount = safe_float(stats.get('orders_amount', 0.0)) if isinstance(stats, dict) else 0.0
            data.append({
                'معرف العميل': cust.get('id', ''), 'الاسم الأول': cust.get('first_name', ''), 'اسم العائلة': cust.get('last_name', ''),
                'البريد الإلكتروني': cust.get('email', ''), 'الجنس': 'ذكر' if cust.get('gender') == 'male' else 'أنثى',
                'رمز الدولة': cust.get('mobile_code', ''), 'رقم الجوال': cust.get('mobile', ''), 'المدينة': cust.get('city', ''),
                'المنطقة / العنوان': cust.get('location', ''), 'عدد الطلبات': orders_count, 'إجمالي المشتريات (SAR)': orders_amount
            })
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            style_excel_file(writer.sheets['Sheet1'], is_template=False, header_color="0F1C2E")
        return buffer.getvalue()
    except: return b""

def export_customer_groups_to_excel(groups: List[Dict]) -> bytes:
    try:
        data = [{'معرف المجموعة': g.get('id', ''), 'اسم المجموعة': g.get('name', '')} for g in groups]
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            style_excel_file(writer.sheets['Sheet1'], is_template=False, header_color="0F1C2E")
        return buffer.getvalue()
    except: return b""

def generate_salla_new_products_file(products: List[Dict]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Salla Products Template Sheet"

        row1 = ["بيانات المنتج"] + [""] * 18 + ["الخيارات والبيانات الإضافية"] + [""] * 20
        ws.append(row1)

        headers = [
            "النوع ", "أسم المنتج", "تصنيف المنتج", "صورة المنتج", "وصف صورة المنتج",
            "نوع المنتج", "سعر المنتج", "الوصف", "هل يتطلب شحن؟", "رمز المنتج sku",
            "سعر التكلفة", "السعر المخفض", "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
            "اقصي كمية لكل عميل", "إخفاء خيار تحديد الكمية", "اضافة صورة عند الطلب",
            "الوزن", "وحدة الوزن", "الماركة", "العنوان الترويجي", "تثبيت المنتج",
            "الباركود", "السعرات الحرارية", "MPN", "GTIN", "خاضع للضريبة ؟",
            "سبب عدم الخضوع للضريبة"
        ]
        ws.append(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
        ws.merge_cells(start_row=1, start_column=20, end_row=1, end_column=40)
        
        elegant_purple = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        dark_slate = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        font_headers = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), 
                             top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
        
        ws.row_dimensions[1].height = 35
        cell_1 = ws.cell(row=1, column=1); cell_1.fill = elegant_purple; cell_1.font = font_title; cell_1.alignment = center_align
        cell_20 = ws.cell(row=1, column=20); cell_20.fill = elegant_purple; cell_20.font = font_title; cell_20.alignment = center_align

        ws.row_dimensions[2].height = 25
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = dark_slate; cell.font = font_headers; cell.alignment = center_align; cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        for p in products:
            price = get_flat_price(p.get('price', 0))
            is_taxable = p.get('with_tax', True)
            tax_cause = p.get('tax_exemption_cause', 'الأدوية والمعدات الطبية') if not is_taxable else ""
            row_data = [""] * len(headers)
            row_data[0] = 'منتج'; row_data[1] = p.get('name', 'بدون اسم'); row_data[5] = 'منتج جاهز'; row_data[6] = price if price > 0 else 0
            row_data[8] = 'نعم'; row_data[9] = p.get('sku', ""); row_data[17] = 1; row_data[18] = 'kg'; row_data[20] = p.get('promotion_title', "")
            row_data[26] = 'نعم' if is_taxable else 'لا'; row_data[27] = tax_cause
            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1): ws.cell(row=ws.max_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except: return b""

# ==========================================
# 🏷️ دوال التحكم الجماعي في العناوين (ترويجي / فرعي)
# ==========================================

def generate_promotions_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "تحديث العناوين"

    # إضافة صف الإرشادات
    ws.append(["💡 إرشادات: ادخل رقم الـمنتج، العناوين الجديدة (إن وجدت)، واختر الإجراء المطلوب من القائمة المنسدلة."])
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 24

    headers = ["SKU (رقم المنتج)", "العنوان الترويجي", "العنوان الفرعي", "الإجراء المطلوب"]
    ws.append(headers)

    # التنسيق الاحترافي
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # أزرق داكن فخم
    header_font = Font(color="FFFFFF", bold=True, name="Cairo", size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), 
                         top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

    # تطبيق التنسيق على صف العناوين (الصف الثاني)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 25

    # إضافة قائمة منسدلة (Data Validation) لعمود الإجراء المطلوب
    dv = DataValidation(type="list", formula1='"تحديث,مسح الترويجي,مسح الفرعي,مسح الكل"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("D3:D1000")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def process_promotions_bulk(df: pd.DataFrame, products_list: List[Dict], headers: Dict[str, str]) -> Dict:
    results = {"success": [], "errors": []}
    # بناء قاموس لتحويل SKU إلى Product ID بسرعة
    sku_to_id = {str(p.get('sku', '')).strip(): p.get('id') for p in products_list if p.get('sku')}
    
    for idx, row in df.iterrows():
        # تجاوز صف الإرشادات
        if row.isna().all() or str(row.iloc[0]).strip().startswith("💡"): continue
        
        sku = str(row.get('SKU (رقم المنتج)', '')).strip()
        if not sku or sku == 'nan': continue
        
        p_id = sku_to_id.get(sku)
        if not p_id:
            results["errors"].append(f"السطر {idx+2}: لم يتم العثور على منتج بـ SKU ({sku})")
            continue
        
        action = str(row.get('الإجراء المطلوب', 'تحديث')).strip()
        promo_val = str(row.get('العنوان الترويجي', '')).strip()
        sub_val = str(row.get('العنوان الفرعي', '')).strip()
        
        if promo_val == 'nan': promo_val = ""
        if sub_val == 'nan': sub_val = ""
        
        new_promo = None
        new_sub = None
        
        if action == 'تحديث':
            if promo_val: new_promo = promo_val
            if sub_val: new_sub = sub_val
        elif action == 'مسح الترويجي':
            new_promo = ""
            if sub_val: new_sub = sub_val # نحدث الفرعي إذا تم كتابته، وإلا نتركه كما هو
        elif action == 'مسح الفرعي':
            new_sub = ""
            if promo_val: new_promo = promo_val # نحدث الترويجي إذا تم كتابته، وإلا نتركه كما هو
        elif action == 'مسح الكل':
            new_promo = ""
            new_sub = ""
        
        # استدعاء الدالة الآمنة (إذا تم إرسال None، فلن يتم تعديل الحقل)
        if update_product_promotions_secure(p_id, new_promo, new_sub, headers):
            results["success"].append(f"تم تنفيذ ({action}) للمنتج {sku} بنجاح.")
        else:
            results["errors"].append(f"فشل تنفيذ ({action}) للمنتج {sku}.")
            
    return results
