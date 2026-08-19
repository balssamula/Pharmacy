import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import time
import io
import re
import base64
import os
from typing import Dict, List, Any, Optional
from utils import (
    get_headers, safe_api_request, SALLA_API_URL, generate_salla_excel_template,
    process_excel_import, export_offers_to_excel, safe_parse_date,
    OFFER_TYPES_MAP, CHANNELS_MAP, APPLIED_TO_MAP, safe_float,
    update_product_promotions_secure
)

ALERT_SOUND_BASE64 = """
UklGRnoAAABXQVZFZm10IBAAAAABAAEAQB8AAEAfAAABAAgAZGF0YQoAAACBhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqFhYqF......
"""

def get_audio_base64(file_path):
    try:
        with open(file_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception as e:
        return None

# ==========================================
# 🚨 نظام التنبيه لقرب انتهاء العروض (مع أزرار الإجراءات الفورية)
# ==========================================
def render_expiry_alerts(raw_offers, headers=None):
    now = datetime.now()
    expiring_soon_grouped = {}
    total_count = 0
    
    for offer in raw_offers:
        if offer.get('status') != 'active': continue
        
        expiry_date = safe_parse_date(offer.get('expiry_date'))
        if not expiry_date: continue
        
        days_left = (expiry_date - now).days
        
        if days_left <= 2:
            date_str = expiry_date.strftime('%Y-%m-%d')
            if date_str not in expiring_soon_grouped:
                expiring_soon_grouped[date_str] = []
            
            expiring_soon_grouped[date_str].append(offer)
            total_count += 1
    
    if "sound_playing" not in st.session_state: st.session_state["sound_playing"] = True
    if "show_expiry_alerts" not in st.session_state: st.session_state["show_expiry_alerts"] = True
    
    # 🔘 أزرار الإخفاء والتحديث أعلى التنبيهات
    col_t1, col_t2, col_t3 = st.columns([2, 2, 6])
    with col_t1:
        toggle_lbl = "👁️ إخفاء التنبيهات" if st.session_state["show_expiry_alerts"] else "👁️ إظهار التنبيهات"
        if st.button(toggle_lbl, use_container_width=True, key="toggle_alert_btn"):
            st.session_state["show_expiry_alerts"] = not st.session_state["show_expiry_alerts"]
            st.rerun()
    with col_t2:
        if st.button("🔄 مزامنة العروض", use_container_width=True, key="refresh_alert_btn"):
            with st.spinner("جاري سحب ومزامنة كافة العروض..."):
                fresh_offers = fetch_all_pages(SALLA_API_URL, "تحديث العروض", headers)
                if fresh_offers:
                    st.session_state["all_offers"] = fresh_offers
                    
                    # ✅ إعادة بناء خريطة المنتجات المرتبطة بالعروض بذكاء وسرعة
                    po_map = {"ALL_PRODUCTS": []}
                    active_offers = [o for o in fresh_offers if o.get('status') == 'active']
                    for o in active_offers:
                        oid = str(o.get("id"))
                        summary = {"id": oid, "name": o.get("name")}
                        if o.get("applied_to") in ["order", "all"] or o.get("offer_type") in ["cart_offer", "tiered_offer"]:
                            po_map["ALL_PRODUCTS"].append(summary)
                        else:
                            pids = set()
                            for px in (o.get("buy", {}) or {}).get("products", []):
                                pid = str(px.get("id", px) if isinstance(px, dict) else px)
                                if pid.isdigit(): pids.add(pid)
                            for px in (o.get("get", {}) or {}).get("products", []):
                                pid = str(px.get("id", px) if isinstance(px, dict) else px)
                                if pid.isdigit(): pids.add(pid)
                            for px in o.get("products", []):
                                pid = str(px.get("id", px) if isinstance(px, dict) else px)
                                if pid.isdigit(): pids.add(pid)
                            for pid in pids:
                                if pid not in po_map: po_map[pid] = []
                                po_map[pid].append(summary)
                    
                    st.session_state["product_offers_map"] = po_map
                st.rerun()

    # إذا كانت مخفية أو لا توجد عروض، نوقف التنفيذ هنا
    if not st.session_state["show_expiry_alerts"]: return
    
    if total_count > 0:
        st.markdown("""
        <style>
            @keyframes blink-red { 0% { border-color: #ff0000; box-shadow: 0 0 10px rgba(255,0,0,0.2); } 50% { border-color: #cc0000; box-shadow: 0 0 30px rgba(255,0,0,0.6); } 100% { border-color: #ff0000; box-shadow: 0 0 10px rgba(255,0,0,0.2); } }
            .expiry-alert { animation: blink-red 1.5s ease-in-out infinite; padding: 20px; border-radius: 12px; background: linear-gradient(135deg, #1e0508 0%, #2c0b0e 100%); color: white; border: 2px solid #ff6b6b; margin-bottom: 25px; direction: rtl; }
            .date-group { background: rgba(255,255,255,0.05); border-radius: 10px; padding: 12px 15px; margin-bottom: 12px; border-right: 5px solid #ffca28; }
            .date-group.expired { border-right: 5px solid #ff4d4d; background: rgba(255,0,0,0.08); }
        </style>
        """, unsafe_allow_html=True)

        if st.session_state["sound_playing"]:
            audio_url = "https://assets.mixkit.co/active_storage/sfx/2869/2869-preview.mp3"
            st.markdown(f"""
            <div style="background: rgba(0,0,0,0.3); padding: 10px; border-radius: 8px; text-align: center; margin-bottom: 15px;">
                <span style='font-size:13px; color:#98ff54; margin-bottom:5px;'>💡 سياسة المتصفحات تمنع أحياناً تشغيل الصوت تلقائياً. اضغط علامة التشغيل ( ▶ ) لتفعيل جرس الإنذار:</span><br>
                <audio id="alert-sound" controls autoplay loop style="height: 35px; outline: none; margin-top: 8px; border-radius: 20px;">
                    <source src="{audio_url}" type="audio/mp3">
                </audio>
            </div>
            """, unsafe_allow_html=True)    
        
        st.markdown(f"""
        <div class="expiry-alert">
            <div style="text-align: center; margin-bottom: 20px;">
                <h2 style="margin:0; color: #ff6b6b;">🚨 انتباه! إجراء مطلوب</h2>
                <p style="color: #cbd5e1; font-size: 15px; margin-top: 5px;">هناك <b style="background: white; color: red; padding: 2px 10px; border-radius: 12px; font-size: 16px;">{total_count}</b> عرض انتهى بالفعل أو سينتهي قريباً!</p>
            </div>
        """, unsafe_allow_html=True)
        
        for date_str, offers in sorted(expiring_soon_grouped.items()):
            days_left = (safe_parse_date(date_str + " 23:59:59") - now).days
            is_expired = days_left < 0
            group_class = "date-group expired" if is_expired else "date-group"
            status_icon = "🛑" if is_expired else "⏳"
            status_text = "انتهت بالفعل!" if is_expired else ("تنتهي اليوم" if days_left == 0 else "تنتهي قريباً")
            
            st.markdown(f"""
            <div class="{group_class}">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; border-bottom: 1px solid rgba(255,255,255,0.1); padding-bottom: 8px;">
                    <span style="font-size: 16px; font-weight: bold; color: #ffd700;">{status_icon} عروض تنتهي في: {date_str} <span style="font-size: 12px; color: {'#ff4d4d' if is_expired else '#cbd5e1'};">({status_text})</span></span>
                    <span style="background: rgba(255,255,255,0.2); font-size: 13px; font-weight: bold; padding: 4px 12px; border-radius: 20px;">العدد: [{len(offers)} عروض]</span>
                </div>
            """, unsafe_allow_html=True)
            
            for offer in offers:
                o_id = offer['id']
                col_txt, col_btn1, col_btn2, col_btn3 = st.columns([6, 1, 1, 1])
                
                with col_txt:
                    st.markdown(f"<div style='background: rgba(0,0,0,0.4); padding: 8px 12px; border-radius: 8px; font-size: 13px; border: 1px solid rgba(255,255,255,0.15); margin-top:2px;'>🎯 {offer['name']}</div>", unsafe_allow_html=True)
                
                with col_btn1:
                    if st.button("⏹️", key=f"al_stop_{o_id}", help="إيقاف العرض وإلغاء تفعيله"):
                        with st.spinner("⏳"):
                            if safe_api_request("PUT", f"{SALLA_API_URL}/{o_id}/status", headers, json={"status": "inactive"}):
                                # ⚡ إجبار الواجهة على تحديث الذاكرة للاختفاء التلقائي الفوري
                                st.session_state["all_offers"] = [dict(o, status='inactive') if str(o.get('id')) == str(o_id) else o for o in st.session_state["all_offers"]]
                                st.rerun()
                            
                with col_btn2:
                    if st.button("🧹", key=f"al_clr_{o_id}", help="مسح العناوين الترويجية للمنتجات المشمولة"):
                        with st.spinner("⏳"):
                            full_res = safe_api_request("GET", f"{SALLA_API_URL}/{o_id}", headers)
                            if full_res and full_res.get('data'):
                                pids = set()
                                for p in full_res['data'].get('buy', {}).get('products', []):
                                    pid = p.get('id', p) if isinstance(p, dict) else p
                                    if str(pid).isdigit(): pids.add(str(pid))
                                for p in full_res['data'].get('get', {}).get('products', []):
                                    pid = p.get('id', p) if isinstance(p, dict) else p
                                    if str(pid).isdigit(): pids.add(str(pid))
                                for pid in pids: update_product_promotions_secure(int(pid), "", "", headers)
                            st.rerun()
                            
                with col_btn3:
                    if st.button("🗑️", key=f"al_del_{o_id}", help="حذف العرض نهائياً من المتجر"):
                        with st.spinner("⏳"):
                            if safe_api_request("DELETE", f"{SALLA_API_URL}/{o_id}", headers):
                                st.session_state["all_offers"] = [o for o in st.session_state["all_offers"] if str(o.get('id')) != str(o_id)]
                                st.rerun()

            st.markdown("</div>", unsafe_allow_html=True)
        
        st.markdown("<div style='margin-top: 20px; display: flex; gap: 10px; justify-content: center; flex-wrap: wrap;'>", unsafe_allow_html=True)
        
        col_btn_bot1, col_btn_bot2, col_btn_bot3 = st.columns(3)
        with col_btn_bot1:
            if st.button("📅 تمديد العروض", use_container_width=True, type="primary", key="ext_btn_bot"):
                st.session_state["qa_action"] = "end_dates"
                st.rerun()
        with col_btn_bot2:
            if st.button("🧹 إزالة العناوين المجمعة", use_container_width=True, key="clr_btn_bot"):
                st.session_state["qa_action"] = "end_dates"
                st.rerun()
        with col_btn_bot3:
            sound_label = "🔇 إخفاء مشغل الصوت" if st.session_state["sound_playing"] else "🔊 إظهار مشغل الصوت"
            if st.button(sound_label, use_container_width=True, key="snd_btn_bot"):
                st.session_state["sound_playing"] = not st.session_state["sound_playing"]
                st.rerun()
        st.markdown("</div></div>", unsafe_allow_html=True)
    else:
        st.success("✅ جميع العروض النشطة سارية المفعول ولا توجد عروض منتهية أو على وشك الانتهاء.")

@st.cache_data(ttl=300, show_spinner=False)
def fetch_offers_cached(headers, _force_refresh=False):
    return fetch_all_pages(SALLA_API_URL, "جاري سحب العروض من متجرك", headers)

def fetch_all_pages(url_base, loading_text, headers):
    all_data = []
    page = 1
    total_pages = 1
    status_text = st.empty()
    progress_bar = st.progress(0)
    
    while True:
        status_text.info(f"📥 {loading_text} (صفحة {page} من {total_pages if page > 1 else '...'}) | تم تحميل {len(all_data)} عنصر")
        url = f"{url_base}?per_page=60&page={page}" if "?" not in url_base else f"{url_base}&per_page=60&page={page}"
        res = safe_api_request("GET", url, headers)
        if not res or not res.get("data"): break
        if page == 1: total_pages = res.get("pagination", {}).get("totalPages", 1)
        all_data.extend(res["data"])
        progress_bar.progress(min(page / total_pages, 1.0))
        if page >= total_pages: break
        page += 1
    
    progress_bar.empty()
    status_text.empty()
    return all_data

def build_product_offers_map_with_progress(offers, headers):
    po_map = {}
    active_offers = [o for o in offers if o.get("status") == "active"]
    if not active_offers: return po_map
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    total = len(active_offers)
    
    for idx, o in enumerate(active_offers):
        status_text.info(f"🔗 جاري بناء روابط المنتجات بالعروض: {idx + 1} من {total}")
        oid = o.get("id")
        full_o = safe_api_request("GET", f"https://api.salla.dev/admin/v2/specialoffers/{oid}", headers)
        if full_o and full_o.get("data"):
            pids = set()
            for px in full_o["data"].get("buy", {}).get("products", []):
                pid = str(px.get("id", px) if isinstance(px, dict) else px)
                if pid.isdigit(): pids.add(pid)
            for px in full_o["data"].get("get", {}).get("products", []):
                pid = str(px.get("id", px) if isinstance(px, dict) else px)
                if pid.isdigit(): pids.add(pid)
            for pid in pids:
                if pid not in po_map: po_map[pid] = []
                po_map[pid].append({"id": oid, "name": o.get("name")})
                
        time.sleep(0.3)
        progress_bar.progress((idx + 1) / total)
    
    progress_bar.empty()
    status_text.empty()
    return po_map

def ensure_product_offers_mapping(headers):
    if not st.session_state.get("product_offers_map") and st.session_state.get("all_offers"):
        with st.spinner("🔄 جاري بناء روابط المنتجات بالعروض الخاصة..."):
            po_map = build_product_offers_mapping_with_progress(st.session_state["all_offers"], headers)
            st.session_state["product_offers_map"] = po_map
            st.rerun()
            
def render_create_offer_section(headers: Dict[str, str], section_key: str = "main"):
    with st.expander("➕ إنشاء عرض ترويجي جديد", expanded=False): 
        key_prefix = f"create_offer_{section_key}"
        st.markdown("#### 🎯 بيانات العرض الأساسية")
        col1, col2 = st.columns(2)
        with col1:
            offer_name = st.text_input("📝 اسم العرض:", placeholder="مثال: عرض العيد المميز", key=f"{key_prefix}_name")
            offer_type = st.selectbox("📊 نوع العرض:", list(OFFER_TYPES_MAP.values()), format_func=lambda x: x, key=f"{key_prefix}_type")
            offer_type_key = [k for k, v in OFFER_TYPES_MAP.items() if v == offer_type][0] if offer_type else "buy_x_get_y"
        with col2:
            applied_to = st.selectbox("🎯 تطبيق العرض على:", list(APPLIED_TO_MAP.values()), format_func=lambda x: x, key=f"{key_prefix}_applied_to")
            applied_to_key = [k for k, v in APPLIED_TO_MAP.items() if v == applied_to][0] if applied_to else "product"
            channel = st.selectbox("📺 قناة النشر:", list(CHANNELS_MAP.values()), format_func=lambda x: x, key=f"{key_prefix}_channel")
            channel_key = [k for k, v in CHANNELS_MAP.items() if v == channel][0] if channel else "browser_and_application"
    
        st.markdown("#### 📅 تواريخ العرض")
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            start_date = st.date_input("📅 تاريخ البدء:", value=datetime.now().date(), key=f"{key_prefix}_start_date")
            start_time = st.time_input("⏰ وقت البدء:", value=datetime.now().time().replace(hour=0, minute=0, second=0), key=f"{key_prefix}_start_time")
        with col_date2:
            end_date = st.date_input("📅 تاريخ الانتهاء:", value=datetime.now().date() + timedelta(days=30), key=f"{key_prefix}_end_date")
            end_time = st.time_input("⏰ وقت الانتهاء:", value=datetime.now().time().replace(hour=23, minute=59, second=59), key=f"{key_prefix}_end_time")
        start_datetime = datetime.combine(start_date, start_time).strftime('%Y-%m-%d %H:%M:%S')
        end_datetime = datetime.combine(end_date, end_time).strftime('%Y-%m-%d %H:%M:%S')
    
        st.markdown("#### ⚙️ خيارات متقدمة")
        col_opt1, col_opt2, col_opt3 = st.columns(3)
        with col_opt1:
            with_coupon = st.checkbox("🔖 تطبيق مع كوبون", value=False, key=f"{key_prefix}_with_coupon")
        with col_opt2:
            max_discount = st.number_input("💰 الحد الأقصى للخصم (SAR):", min_value=0.0, value=0.0, step=5.0, key=f"{key_prefix}_max_discount")
        with col_opt3:
            min_purchase = st.number_input("💵 الحد الأدنى للشراء (SAR):", min_value=0.0, value=0.0, step=10.0, key=f"{key_prefix}_min_purchase")
    
        st.markdown("#### 🛒 شروط العرض (اشتر X واحصل على Y)")
        col_buy1, col_buy2, col_buy3 = st.columns(3)
        with col_buy1:
            buy_type = st.selectbox("نوع الشراء (X):", ["منتج", "تصنيف", "ماركة"], key=f"{key_prefix}_buy_type")
            buy_type_key = {"منتج": "product", "تصنيف": "category", "ماركة": "brand"}.get(buy_type, "product")
        with col_buy2:
            buy_quantity = st.number_input("كمية الشراء (X):", min_value=1, value=1, step=1, key=f"{key_prefix}_buy_quantity")
        with col_buy3:
            if buy_type == "منتج":
                buy_items = st.multiselect("اختر المنتجات:", options=[f"{p.get('name')} (SKU: {p.get('sku')})" for p in st.session_state.get("all_products", []) if p.get('sku')], key=f"{key_prefix}_buy_products")
                buy_ids = []
                for item in buy_items:
                    for p in st.session_state.get("all_products", []):
                        if f"{p.get('name')} (SKU: {p.get('sku')})" == item:
                            buy_ids.append(p.get('id')); break
            elif buy_type == "تصنيف":
                buy_items = st.multiselect("اختر التصنيفات:", options=[f"{c.get('name')} (ID: {c.get('id')})" for c in st.session_state.get("all_categories", [])], key=f"{key_prefix}_buy_categories")
                buy_ids = [int(c.split('ID: ')[1].replace(')', '')) for c in buy_items]
            else:
                buy_items = st.multiselect("اختر الماركات:", options=[f"{b.get('name')} (ID: {b.get('id')})" for b in st.session_state.get("all_brands", [])], key=f"{key_prefix}_buy_brands")
                buy_ids = [int(b.split('ID: ')[1].replace(')', '')) for b in buy_items]
    
        st.markdown("#### 🎁 العرض (يحصل على Y)")
        col_get1, col_get2, col_get3 = st.columns(3)
        with col_get1:
            get_type = st.selectbox("نوع العرض (Y):", ["منتج", "تصنيف", "ماركة"], key=f"{key_prefix}_get_type")
            get_type_key = {"منتج": "product", "تصنيف": "category", "ماركة": "brand"}.get(get_type, "product")
        with col_get2:
            get_quantity = st.number_input("كمية العرض (Y):", min_value=1, value=1, step=1, key=f"{key_prefix}_get_quantity")
        with col_get3:
            discount_type = st.selectbox("نوع الخصم:", ["خصم بنسبة", "منتج مجاني", "مبلغ ثابت"], key=f"{key_prefix}_discount_type")
            discount_type_key = {"خصم بنسبة": "percentage", "منتج مجاني": "free-product", "مبلغ ثابت": "fixed_amount"}.get(discount_type, "percentage")
            discount_amount = 0.0
            if discount_type != "منتج مجاني":
                discount_amount = st.number_input("قيمة الخصم:", min_value=0.0, value=10.0, step=1.0, key=f"{key_prefix}_discount_amount")
    
        if get_type == "منتج":
            get_items = st.multiselect("اختر المنتجات للعرض:", options=[f"{p.get('name')} (SKU: {p.get('sku')})" for p in st.session_state.get("all_products", []) if p.get('sku')], key=f"{key_prefix}_get_products")
            get_ids = []
            for item in get_items:
                for p in st.session_state.get("all_products", []):
                    if f"{p.get('name')} (SKU: {p.get('sku')})" == item:
                        get_ids.append(p.get('id')); break
        elif get_type == "تصنيف":
            get_items = st.multiselect("اختر التصنيفات للعرض:", options=[f"{c.get('name')} (ID: {c.get('id')})" for c in st.session_state.get("all_categories", [])], key=f"{key_prefix}_get_categories")
            get_ids = [int(c.split('ID: ')[1].replace(')', '')) for c in get_items]
        else:
            get_items = st.multiselect("اختر الماركات للعرض:", options=[f"{b.get('name')} (ID: {b.get('id')})" for b in st.session_state.get("all_brands", [])], key=f"{key_prefix}_get_brands")
            get_ids = [int(b.split('ID: ')[1].replace(')', '')) for b in get_items]
    
        offer_message = st.text_input("💬 رسالة العرض:", placeholder="تسوق الآن واستمتع بالخصم!", key=f"{key_prefix}_message")
    
        if st.button("🚀 إنشاء العرض الجديد", type="primary", use_container_width=True, key=f"{key_prefix}_submit"):
            if not offer_name:
                st.error("⚠️ الرجاء إدخال اسم العرض")
                return
            payload = {
                "name": offer_name, "offer_type": offer_type_key, "applied_channel": channel_key, "applied_to": applied_to_key,
                "start_date": start_datetime, "expiry_date": end_datetime, "status": "active", "applied_with_coupon": with_coupon,
                "max_discount_amount": max_discount, "min_purchase_amount": min_purchase, "min_items_count": 0, "message": offer_message,
                "buy": {"type": buy_type_key, "quantity": buy_quantity}, "get": {"type": get_type_key, "quantity": get_quantity, "discount_type": discount_type_key}
            }
            if buy_ids:
                if buy_type_key == "product": payload["buy"]["products"] = buy_ids
                elif buy_type_key == "category": payload["buy"]["categories"] = buy_ids
                elif buy_type_key == "brand": payload["buy"]["brands"] = buy_ids
            if get_ids:
                if get_type_key == "product": payload["get"]["products"] = get_ids
                elif get_type_key == "category": payload["get"]["categories"] = get_ids
                elif get_type_key == "brand": payload["get"]["brands"] = get_ids
            if discount_amount > 0: payload["get"]["discount_amount"] = discount_amount
        
            with st.spinner("جاري إنشاء العرض..."):
                res = safe_api_request("POST", SALLA_API_URL, headers, json=payload)
                if res:
                    st.success("✅ تم إنشاء العرض بنجاح!")
                    st.rerun()
                else: st.error("❌ فشل إنشاء العرض")

def get_advanced_export_excel(offers_list, all_products_list):
    """دالة تصدير مطابقة لنموذج سلة مع التنسيقات وأعمدة SKU الإضافية"""
    import pandas as pd
    import io
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # 1. إنشاء قاموس لربط ID بـ SKU من المنتجات المتوفرة
    id_to_sku = {}
    for p in all_products_list:
        pid = str(p.get('id', ''))
        sku = str(p.get('sku', 'لا يوجد'))
        if pid: id_to_sku[pid] = sku
        
    rows = []
    for o in offers_list:
        buy_obj = o.get('buy', {}) or {}
        get_obj = o.get('get', {}) or {}
        
        # استخراج IDs الشراء
        buy_ids = []
        for p in buy_obj.get('products', []):
            pid = str(p.get('id', p) if isinstance(p, dict) else p)
            if pid.isdigit(): buy_ids.append(pid)
        for p in o.get('products', []): # للأنواع المباشرة
            pid = str(p.get('id', p) if isinstance(p, dict) else p)
            if pid.isdigit() and pid not in buy_ids: buy_ids.append(pid)
            
        # استخراج IDs العرض
        get_ids = []
        for p in get_obj.get('products', []):
            pid = str(p.get('id', p) if isinstance(p, dict) else p)
            if pid.isdigit(): get_ids.append(pid)
            
        # تحويل IDs إلى SKUs
        buy_skus = ",".join([id_to_sku.get(pid, 'غير_معروف') for pid in buy_ids])
        get_skus = ",".join([id_to_sku.get(pid, 'غير_معروف') for pid in get_ids])
        
        # تحديد نوع الخصم وقيمته
        disc_type_raw = get_obj.get('discount_type', '')
        if not disc_type_raw:
            o_type = o.get('offer_type', '')
            disc_type_raw = 'percentage' if o_type == 'percentage' else ('fixed_amount' if o_type == 'fixed_amount' else o_type)
            
        disc_type_ar = "خصم بنسبة" if disc_type_raw == "percentage" else ("مبلغ ثابت" if disc_type_raw == "fixed_amount" else ("منتج مجاني" if disc_type_raw in ["free-product", "buy_x_get_y"] else disc_type_raw))
        disc_amt = get_obj.get('discount_amount', o.get('max_discount_amount', o.get('discount_amount', 0)))

        # بناء الصف بالترتيب القياسي لسلة + أعمدتك الجديدة
        rows.append({
            "الإجراء": "تحديث",
            "معرف العرض": str(o.get('id', '')),
            "اسم العرض": str(o.get('name', '')),
            "نوع العرض": "اذا اشترى العميل X يحصل على Y" if o.get('offer_type') == 'buy_x_get_y' else str(o.get('offer_type', '')),
            "المنصة": "متصفح وتطبيق المتجر",
            "تطبيق على": "منتجات مختارة",
            "تاريخ البدء": str(o.get('start_date', '')),
            "تاريخ الانتهاء": str(o.get('expiry_date', '')),
            "تطبيق مع كوبون": "نعم" if o.get('applied_with_coupon') else "لا",
            "الحد الأقصى للخصم": float(o.get('max_discount_amount', 0) or 0),
            "الحد الأدنى للشراء": float(o.get('min_purchase_amount', 0) or 0),
            "الحد الأدنى للكمية": int(o.get('min_items_count', 0) or 0),
            "مجموعات العملاء": "",
            "نوع شراء X": "منتج",
            "كمية شراء X": int(buy_obj.get('quantity', 1)),
            "عناصر شراء X (IDs)": ",".join(buy_ids),
            "عناصر شراء X (SKUs)": buy_skus,  # ✅ العمود الجديد
            "نوع عرض Y": "منتج",
            "كمية عرض Y": int(get_obj.get('quantity', 1)),
            "عناصر عرض Y (IDs)": ",".join(get_ids),
            "عناصر عرض Y (SKUs)": get_skus,  # ✅ العمود الجديد
            "نوع الخصم": disc_type_ar,
            "قيمة الخصم": float(disc_amt or 0),
            "رسالة العرض": str(o.get('message', '')),
            "حالة العرض": "نشط" if o.get('status') == 'active' else "غير نشط"
        })
        
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    
    # 2. تطبيق تنسيقات Excel الأنيقة (نفس تنسيق سلة القديم)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salla Offers"
    
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
        
    # تنسيق رأس الجدول (Header)
    header_fill = PatternFill(start_color="0F1C2E", end_color="0F1C2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # توسيع الأعمدة قليلاً
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # إضافة الفلتر التلقائي (AutoFilter)
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(buf)
    return buf.getvalue()
    
def render_offers_page():
    # ✅ تأسيس الذاكرة الجوهرية قبل تشغيل أي شيء في الصفحة
    if "featured_offer_groups" not in st.session_state:
        st.session_state["featured_offer_groups"] = {}
    if "sound_playing" not in st.session_state:
        st.session_state["sound_playing"] = True
    if "qa_action" not in st.session_state: 
        st.session_state.qa_action = None

    st.markdown("""
    <div style="background: linear-gradient(135deg, #0F1C2E 0%, #00EBCF 100%); padding: 15px 25px; border-radius: 12px; color: white; margin-bottom: 20px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
        <h2 style="color: white; margin: 0;">📊 مركز إدارة العروض الخاصة المتقدم</h2>
    </div>
    """, unsafe_allow_html=True)
    
    headers = get_headers()
    if not headers: return

    # ==========================================
    # 🌟 CSS الأزرار الجانبية (مع ترتيب صحيح)
    # ==========================================
    st.markdown("""
    <style>
        div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) { display: none !important; margin: 0 !important; padding: 0 !important; }
        div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"] {
            position: fixed !important; right: -240px !important; width: 280px !important; background: linear-gradient(135deg, #1E293B 0%, #0F1C2E 100%) !important;
            padding: 5px 10px !important; border-radius: 20px 0 0 20px !important; border: 2px solid #00EBCF !important; border-right: none !important;
            z-index: 999999 !important; transition: right 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important; box-shadow: -4px 4px 12px rgba(0,0,0,0.3) !important;
        }
        div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"]:hover { right: 0px !important; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-1"]) + div[data-testid="stElementContainer"] { top: 120px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-2"]) + div[data-testid="stElementContainer"] { top: 185px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-3"]) + div[data-testid="stElementContainer"] { top: 250px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-4"]) + div[data-testid="stElementContainer"] { top: 315px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-5"]) + div[data-testid="stElementContainer"] { top: 380px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-6"]) + div[data-testid="stElementContainer"] { top: 445px; }
        div[data-testid="stElementContainer"]:has(span[id="qa-marker-7"]) + div[data-testid="stElementContainer"] { top: 510px; }
        div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"] button { width: 100% !important; text-align: right !important; padding-right: 35px !important; font-size: 14px !important; font-weight: bold !important; background: transparent !important; border: none !important; color: white !important; box-shadow: none !important; }
        div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"]::before { content: "👈"; position: absolute; left: 10px; top: 50%; transform: translateY(-50%); font-size: 18px; pointer-events: none; }
        
        .filter-container { background: linear-gradient(135deg, #1a2332 0%, #0f1c2e 100%); padding: 15px 20px; border-radius: 12px; border: 1px solid #2d3a4a; margin-bottom: 15px; }
        .filter-title { color: #00EBCF; font-weight: bold; font-size: 13px; text-align: center; margin-bottom: 8px; display: block; }
        .filter-container .stRadio > div { justify-content: center !important; }
        .filter-container .stRadio label { color: #cbd5e1 !important; }
        .filter-container .stRadio label div[data-testid="stMarkdownContainer"] p { color: #cbd5e1 !important; }
        .filter-container .stRadio [data-testid="stBaseButton-selected"] { background-color: #00EBCF !important; color: #0f1c2e !important; }
        
        .mobile-toggle-btn { position: fixed; right: 0px; top: 50%; transform: translateY(-50%); background: #00EBCF; color: #0f1c2e; border: none; border-radius: 8px 0 0 8px; padding: 10px 6px; font-size: 14px; cursor: pointer; z-index: 999998; writing-mode: vertical-rl; font-weight: bold; box-shadow: -2px 2px 10px rgba(0,0,0,0.3); display: none; }
        @media (pointer: coarse) {
            .mobile-toggle-btn { display: block !important; }
            div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"] { right: 0px !important; width: 200px !important; opacity: 0.85 !important; }
            div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"] button { font-size: 12px !important; padding-right: 10px !important; padding: 4px 8px !important; }
            div[data-testid="stElementContainer"]:has(span[id^="qa-marker-"]) + div[data-testid="stElementContainer"]::before { display: none !important; }
        }
    </style>
    <button class="mobile-toggle-btn" onclick="var btns = document.querySelectorAll('div[data-testid=\\'stElementContainer\\']:has(span[id^=\\'qa-marker-\\']) + div[data-testid=\\'stElementContainer\\']'); btns.forEach(function(el) { if (el.style.right === '0px' || el.style.right === '0') { el.style.right = '-200px'; } else { el.style.right = '0px'; } });">⚡ إجراءات</button>
    """, unsafe_allow_html=True)

    def fetch_all_pages_local(url_base, loading_text="جاري التحميل..."):
        all_data = []
        page = 1
        total_pages = 1
        status_text = st.empty()
        progress_bar = st.progress(0)
        while True:
            status_text.info(f"📥 {loading_text} (صفحة {page} من {total_pages if page > 1 else '...'}) | تم تحميل {len(all_data)} عنصر")
            url = f"{url_base}?per_page=60&page={page}" if "?" not in url_base else f"{url_base}&per_page=60&page={page}"
            res = safe_api_request("GET", url, headers)
            if not res or not res.get("data"): break
            if page == 1: total_pages = res.get("pagination", {}).get("totalPages", 1)
            all_data.extend(res["data"])
            progress_bar.progress(min(page / total_pages, 1.0))
            if page >= total_pages: break
            page += 1
        progress_bar.empty()
        status_text.empty()
        return all_data
    
    def render_dynamic_selection(label, selection_type, existing_ids, key_prefix):
        options = {}
        if selection_type == "product":
            for p in st.session_state.get("all_products", []): options[f"🆔 {p.get('id')} - SKU:{p.get('sku','')} - {p.get('name','')}"] = p.get('id')
        elif selection_type == "category":
            for c in st.session_state.get("all_categories", []): options[f"📁 {c.get('name','')} - (ID:{c.get('id')})"] = c.get('id')
        elif selection_type == "brand":
            for b in st.session_state.get("all_brands", []): options[f"🏢 {b.get('name','')} - (ID:{b.get('id')})"] = b.get('id')
        selected_labels = []
        options_inv = {v: k for k, v in options.items()}
        for eid in existing_ids:
            if eid in options_inv: selected_labels.append(options_inv[eid])
            else:
                fallback = f"ID: {eid} (غير متوفر)"
                options[fallback] = eid
                selected_labels.append(fallback)
        selected = st.multiselect(label, options=list(options.keys()), default=selected_labels, key=key_prefix)
        return [options[s] for s in selected]
    
    if "all_products" not in st.session_state: st.session_state["all_products"] = []
    if "all_categories" not in st.session_state: st.session_state["all_categories"] = []
    if "all_brands" not in st.session_state: st.session_state["all_brands"] = []
    
    if not st.session_state["all_products"] or not st.session_state["all_categories"] or not st.session_state["all_brands"]:
        with st.spinner("🔄 جاري تهيئة البيانات المساعدة للعروض..."):
            if not st.session_state["all_categories"]: st.session_state["all_categories"] = fetch_all_pages_local("https://api.salla.dev/admin/v2/categories", "جاري سحب التصنيفات")
            if not st.session_state["all_brands"]: st.session_state["all_brands"] = fetch_all_pages_local("https://api.salla.dev/admin/v2/brands", "جاري سحب الماركات التجارية")
            if not st.session_state["all_products"]: st.session_state["all_products"] = fetch_all_pages_local("https://api.salla.dev/admin/v2/products", "جاري سحب قائمة المنتجات")

    # ⚡ استدعاء البيانات من الذاكرة المركزية مباشرة بدون تكرار التحميل
    if "all_offers" not in st.session_state:
        st.warning("⚠️ يرجى مزامنة البيانات من لوحة التحكم الرئيسية أولاً.")
        return
    
    raw_offers = st.session_state["all_offers"]
    
    render_expiry_alerts(raw_offers, headers)
    render_create_offer_section(headers, section_key="main")
    
    if "qa_action" not in st.session_state: st.session_state.qa_action = None

    st.markdown('<span id="qa-marker-1"></span>', unsafe_allow_html=True)
    if st.button("⏹️ إيقاف جميع العروض المفعلة", key="btn_qa_1"):
        active_offers = [o for o in raw_offers if o.get('status') == 'active']
        if not active_offers: st.warning("لا توجد عروض مفعلة")
        else:
            with st.spinner("جاري الإيقاف..."):
                c = sum(1 for o in active_offers if safe_api_request("PUT", f"{SALLA_API_URL}/{o.get('id')}/status", headers, json={"status": "inactive"}))
                st.success(f"تم إيقاف {c} عرض!")
        st.session_state.qa_action = None

    st.markdown('<span id="qa-marker-2"></span>', unsafe_allow_html=True)
    if st.button("📅 إدارة تواريخ الانتهاء", key="btn_qa_2"):
        st.session_state.qa_action = "end_dates"
        st.rerun()

    st.markdown('<span id="qa-marker-3"></span>', unsafe_allow_html=True)
    if st.button("📅 إدارة تواريخ البداية", key="btn_qa_3"):
        st.session_state.qa_action = "start_dates"
        st.rerun()

    st.markdown('<span id="qa-marker-4"></span>', unsafe_allow_html=True)
    if st.button("🚀 تطبيق العرض مع كوبون (الكل)", key="btn_qa_4"):
        with st.spinner("جاري تحديث كافة العروض..."):
            active_offers = [o for o in raw_offers if o.get("status") == "active"]
            c = 0
            for offer_summary in active_offers:
                offer_id = offer_summary.get("id")
                full_offer_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
                if not full_offer_res or not full_offer_res.get("data"): continue
                payload = rebuild_offer_payload(full_offer_res["data"], {"applied_with_coupon": True})
                if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json=payload): c += 1
            st.success(f"تم تفعيل دمج الكوبونات لـ {c} عرض!")
        st.session_state.qa_action = None

    st.markdown('<span id="qa-marker-5"></span>', unsafe_allow_html=True)
    if st.button("📝 مسودة تأسيس العروض", key="btn_qa_5"):
        st.session_state.qa_action = "draft"
        st.rerun()

    st.markdown('<span id="qa-marker-6"></span>', unsafe_allow_html=True)
    if st.button("➕ عرض جديد", key="btn_qa_6"):
        st.session_state.qa_action = "create_offer"
        st.rerun()

    if st.session_state.qa_action == "create_offer":
        with st.container(border=True):
            col_t, col_c = st.columns([5, 1])
            with col_c:
                if st.button("❌ إغلاق", use_container_width=True, type="primary", key="close_create_offer"):
                    st.session_state.qa_action = None
                    st.rerun()
            with col_t:
                st.markdown("### ➕ إنشاء عرض جديد")
                render_create_offer_section(headers, section_key="popup")
                
    # ✅ الزر الجديد للعروض المميزة (المجموعات)
    st.markdown('<span id="qa-marker-7"></span>', unsafe_allow_html=True)
    if st.button("⭐ مجموعات العروض المميزة", key="btn_qa_7"): st.session_state.qa_action = "featured_groups"; st.rerun()

    # ==========================================
    # ⚙️ عرض الشاشات المنبثقة للإجراءات
    # ==========================================
    if st.session_state.qa_action in ["end_dates", "start_dates", "draft", "featured_groups"]:
        with st.container(border=True):
            col_t, col_c = st.columns([5, 1])
            with col_c:
                if st.button("❌ إغلاق اللوحة", use_container_width=True, type="primary"):
                    st.session_state.qa_action = None
                    st.rerun()

            if st.session_state.qa_action == "featured_groups":
                with col_t: st.markdown("### ⭐ إدارة مجموعات العروض المميزة")
                st.info("💡 يمكنك من هنا تجميع عدة عروض تحت اسم واحد (مثلاً: عروض الشتاء، عروض الفلاش) لتسهيل فلترتها وتنفيذ إجراءات سريعة عليها لاحقاً.")
                
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    st.markdown("#### ➕ إنشاء مجموعة جديدة")
                    new_g_name = st.text_input("اسم المجموعة الجديدة:")
                    off_opts = {f"🎯 {o['name']} (ID: {o['id']})": o['id'] for o in raw_offers}
                    sel_offs = st.multiselect("اختر العروض المراد إضافتها للمجموعة:", options=list(off_opts.keys()))
                    if st.button("💾 حفظ المجموعة", type="primary"):
                        if not new_g_name: st.error("الرجاء كتابة اسم المجموعة")
                        elif not sel_offs: st.error("الرجاء اختيار عرض واحد على الأقل")
                        else:
                            st.session_state["featured_offer_groups"][new_g_name] = [off_opts[k] for k in sel_offs]
                            st.success("✅ تم حفظ المجموعة بنجاح!")
                            st.rerun()
                with col_g2:
                    st.markdown("#### 📋 المجموعات الحالية")
                    if st.session_state.get("featured_offer_groups"):
                        for g_name, g_ids in st.session_state.get("featured_offer_groups", {}).items():
                            with st.expander(f"📁 {g_name} ({len(g_ids)} عروض)"):
                                for oid in g_ids:
                                    offer_name = next((o['name'] for o in raw_offers if o['id'] == oid), "غير معروف")
                                    st.markdown(f"- `{offer_name}`")
                                if st.button("🗑️ حذف المجموعة", key=f"del_g_{g_name}", type="primary"):
                                    del st.session_state["featured_offer_groups"][g_name]
                                    st.rerun()
                    else:
                        st.warning("لا توجد مجموعات مميزة حالياً.")
                        
            if st.session_state.qa_action == "end_dates":
                with col_t: st.markdown("### 📅 إدارة وتمديد تواريخ الانتهاء")
                target_scope_end = st.radio("استهداف العروض للتمديد/الإنهاء:", ["تنتهي في تاريخ محدد", "جميع العروض المتاحة"], key="b_end_scope")
                if target_scope_end == "تنتهي في تاريخ محدد":
                    col_date1, col_time1 = st.columns(2)
                    with col_date1: target_date = st.date_input("تاريخ الانتهاء الحالي:", value=datetime.now().date(), key="be_t_date")
                    with col_time1: target_time = st.time_input("الوقت:", value=datetime.now().time().replace(minute=59, second=59), key="be_t_time")
                    target_str = datetime.combine(target_date, target_time).strftime('%Y-%m-%d')
                    matching_offers_end = [o for o in raw_offers if o.get('status') == 'active' and o.get('expiry_date', '') and o.get('expiry_date', '').startswith(target_str)]
                else: 
                    matching_offers_end = [o for o in raw_offers if o.get('status') == 'active']

                action_type = st.radio("الإجراء المطلوب تنفيذه:", ["تاريخ جديد للتمديد", "إلغاء التفعيل", "مسح العناوين الترويجية لمنتجاتها"], key="be_action")
                
                if action_type == "تاريخ جديد للتمديد":
                    col_date2, col_time2 = st.columns(2)
                    with col_date2: new_date = st.date_input("التاريخ الجديد:", value=datetime.now().date() + timedelta(days=30), key="be_n_date")
                    with col_time2: new_time = st.time_input("الوقت الجديد:", value=datetime.now().time().replace(minute=59, second=59), key="be_n_time")
                    new_expiry_str = datetime.combine(new_date, new_time).strftime('%Y-%m-%d %H:%M:%S')
                    btn_lbl = "🔄 تطبيق التمديد"
                elif action_type == "إلغاء التفعيل": btn_lbl = "🛑 تطبيق الإيقاف"
                else: btn_lbl = "🧹 مسح العناوين الترويجية للمنتجات"
                
                if st.button(btn_lbl, use_container_width=True, type="primary"):
                    if not matching_offers_end: st.warning("لا توجد عروض مطابقة (أو ربما جميعها متوقفة)")
                    else:
                        with st.spinner("جاري المعالجة..."):
                            if action_type == "مسح العناوين الترويجية لمنتجاتها":
                                product_ids = set()
                                st.info("جاري استخراج المنتجات المشمولة من العروض...")
                                for offer in matching_offers_end:
                                    offer_id = offer.get('id')
                                    full_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
                                    if full_res and full_res.get('data'):
                                        off_data = full_res['data']
                                        for p in off_data.get('buy', {}).get('products', []):
                                            pid = p.get('id', p) if isinstance(p, dict) else p
                                            if str(pid).isdigit(): product_ids.add(str(pid))
                                        for p in off_data.get('get', {}).get('products', []):
                                            pid = p.get('id', p) if isinstance(p, dict) else p
                                            if str(pid).isdigit(): product_ids.add(str(pid))
                                c_prods = 0
                                st.info(f"تم العثور على {len(product_ids)} منتج مختلف. جاري التحديث الآمن وإزالة العناوين...")
                                for pid in product_ids:
                                    if update_product_promotions_secure(int(pid), "", "", headers): c_prods += 1
                                st.success(f"✅ تم مسح وإفراغ العناوين الترويجية لـ {c_prods} منتج بنجاح!")
                                st.cache_data.clear(); st.rerun()
                            else:
                                c = 0
                                for offer in matching_offers_end:
                                    offer_id = offer.get('id')
                                    if action_type == "إلغاء التفعيل":
                                        if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}/status", headers, json={"status": "inactive"}): c += 1
                                    else:
                                        full_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
                                        if full_res and full_res.get('data'):
                                            payload = rebuild_offer_payload(full_res['data'], {"expiry_date": new_expiry_str})
                                            if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json=payload): c += 1
                                action_word = "إيقاف" if action_type == "إلغاء التفعيل" else "تمديد"
                                st.success(f"✅ تم تنفيذ إجراء الـ {action_word} لـ {c} عرض بنجاح!")
                                st.cache_data.clear(); st.rerun()

            elif st.session_state.qa_action == "start_dates":
                with col_t: st.markdown("### 📅 إدارة وتعديل تواريخ البداية")
                target_scope_start = st.radio("استهداف العروض للبدء:", ["تبدأ في تاريخ محدد", "جميع العروض المتاحة"], key="bs_scope")
                if target_scope_start == "تبدأ في تاريخ محدد":
                    col_sd1, col_st1 = st.columns(2)
                    with col_sd1: target_s_date = st.date_input("تاريخ البداية الحالي:", value=datetime.now().date(), key="bs_t_date")
                    with col_st1: target_s_time = st.time_input("الوقت:", value=datetime.now().time().replace(minute=0, second=0), key="bs_t_time")
                    target_s_str = datetime.combine(target_s_date, target_s_time).strftime('%Y-%m-%d')
                    matching_offers_start = [o for o in raw_offers if o.get('start_date', '') and o.get('start_date', '').startswith(target_s_str)]
                else: matching_offers_start = raw_offers
                
                if st.button("🕒 تعيين الوقت الحالي", key="btn_now_start_bulk"):
                    st.session_state["bs_n_date"] = (datetime.now() + timedelta(hours=3)).date()
                    st.session_state["bs_n_time"] = (datetime.now() + timedelta(hours=3)).time()
                col_nsd, col_nst = st.columns(2)
                with col_nsd: new_s_date = st.date_input("تاريخ البداية الجديد:", value=st.session_state.get("bs_n_date", datetime.now().date()), key="bs_n_date")
                with col_nst: new_s_time = st.time_input("الوقت الجديد:", value=st.session_state.get("bs_n_time", datetime.now().time().replace(minute=0, second=0)), key="bs_n_time")
                new_start_str = datetime.combine(new_s_date, new_s_time).strftime('%Y-%m-%d %H:%M:%S')
                
                if st.button("🚀 تطبيق تاريخ البداية", use_container_width=True, type="primary"):
                    if not matching_offers_start: st.warning("لا توجد عروض مطابقة")
                    else:
                        with st.spinner("جاري التعديل..."):
                            c = 0
                            for offer in matching_offers_start:
                                offer_id = offer.get('id')
                                full_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
                                if full_res and full_res.get('data'):
                                    payload = rebuild_offer_payload(full_res['data'], {"start_date": new_start_str})
                                    if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json=payload): c += 1
                            st.success(f"تم تعديل البداية لـ {c} عرض!")

            elif st.session_state.qa_action == "draft":
                with col_t: st.markdown("### 📝 مسودة تأسيس العروض")
                st.info("استخدم هذه الأداة لرفع ملف المسودة (ملف قراءة العروض) وسيقوم النظام بإنشاء ملف إكسيل للعروض جاهز للاستيراد في سلة.")
                
                def get_draft_template():
                    import openpyxl
                    from openpyxl.worksheet.datavalidation import DataValidation
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    draft_headers = ["رقم المنتج sku", "اسم المنتج", "سعر بيع المنتج", "نسبة الضريبة", "العلامة التجارية", "Agent", "التصنيفات", "اسم العرض", "كمية المنتجات لتطبيق العرض", "نسبة الخصم", "قيمة الخصم شامل", "السعر قبل شامل", "السعر بعد شامل", "هل عرض مجمع؟", "تاريخ البدء", "تاريخ الانتهاء", "نوع الخصم"]
                    example_row = ["2746", "سنافي 20 مج حبوب 4 ق", 88.95, 0.15, "N.A", "SPIMACO", "MEDICINE", "2حبة بسعر 99.95 ريال", 2, 0.5, 77.95, 177.9, 99.95, "عرض_مجمع_1", "2026-08-01 00:00:00", "2026-08-30 23:59:59", "خصم بنسبة"]
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "مسودة العروض"
                    ws.append(draft_headers)
                    ws.append(example_row)
                    header_fill = PatternFill(start_color="0F1C2E", end_color="0F1C2E", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)
                    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
                    for col in range(1, len(draft_headers) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
                    dv = DataValidation(type="list", formula1='"منتج مجاني,خصم بنسبة"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add("Q2:Q1000")
                    buf = io.BytesIO()
                    wb.save(buf)
                    return buf.getvalue()

                st.download_button("📥 تنزيل نموذج مسودة التأسيس", data=get_draft_template(), file_name="Draft_Offers_Template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                uploaded_draft = st.file_uploader("📂 رفع ملف مسودة العروض المعبأ:", type=["xlsx", "csv"], key="draft_up")
                
                if uploaded_draft:
                    if st.button("⚙️ معالجة المسودة وتوليد ملف العروض الجاهز لرفعه", type="primary"):
                        try:
                            df_draft = pd.read_excel(uploaded_draft) if uploaded_draft.name.endswith('.xlsx') else pd.read_csv(uploaded_draft)
                            required_columns = ['رقم المنتج sku', 'اسم المنتج', 'نوع الخصم']
                            missing_cols = [col for col in required_columns if col not in df_draft.columns]
                            if missing_cols:
                                st.error(f"❌ الأعمدة التالية غير موجودة في الملف: {', '.join(missing_cols)}")
                                st.info("تأكد من استخدام القالب الصحيح أو أعد تحميل الملف.")
                                return
                            
                            sku_map = {str(p.get('sku', '')).strip(): str(p.get('id', '')) for p in st.session_state["all_products"] if p.get('sku')}
                            offer_name_map = {str(o.get('name', '')).strip(): str(o.get('id', '')) for o in raw_offers}
                            output_rows = []
                            
                            if 'هل عرض مجمع؟' in df_draft.columns: group_col = 'هل عرض مجمع؟'
                            elif 'عرض مجمع' in df_draft.columns: group_col = 'عرض مجمع'
                            else: group_col = None; st.info("ℹ️ لم يتم العثور على عمود 'هل عرض مجمع؟'، سيتم معاملة كل صف كعرض منفصل.")
                            
                            offer_name_col = 'اسم العرض' if 'اسم العرض' in df_draft.columns else None
                            
                            if group_col and group_col in df_draft.columns:
                                df_draft[group_col] = df_draft[group_col].fillna('')
                                empty_mask = df_draft[group_col] == ''
                                df_draft.loc[empty_mask, group_col] = df_draft.loc[empty_mask].index.astype(str) + "_single_offer"
                                
                                for group_name, group_df in df_draft.groupby(group_col):
                                    first_row = group_df.iloc[0]
                                    skus = [str(x).strip() for x in group_df['رقم المنتج sku'].tolist() if pd.notna(x) and str(x).strip()]
                                    ids = [sku_map.get(s, '') for s in skus if sku_map.get(s, '')]
                                    sku_str = "-".join(skus) if skus else "UnknownSKU"
                                    id_str = ",".join(ids)
                                    
                                    if offer_name_col and offer_name_col in group_df.columns: offer_msg = str(first_row.get(offer_name_col, ''))
                                    else: offer_msg = str(first_row.get('اسم العرض', '')) if 'اسم العرض' in df_draft.columns else ''
                                    if not offer_msg: offer_msg = f"{sku_str} / عرض"
                                    
                                    offer_name = f"{sku_str} / {offer_msg}"
                                    action = "تحديث" if offer_name in offer_name_map else "إنشاء"
                                    offer_id = offer_name_map.get(offer_name, "")
                                    start_date = str(first_row.get('تاريخ البدء', '')) if pd.notna(first_row.get('تاريخ البدء')) else ''
                                    end_date = str(first_row.get('تاريخ الانتهاء', '')) if pd.notna(first_row.get('تاريخ الانتهاء')) else ''
                                    
                                    discount_type_raw = str(first_row.get('نوع الخصم', '')).strip()
                                    salla_discount_type = "منتج مجاني" if discount_type_raw == "منتج مجاني" else "خصم بنسبة"
                                    discount_amount = 0.0; buy_qty = 1; get_qty = 1
                                    
                                    if salla_discount_type == "منتج مجاني":
                                        match = re.search(r'(\d+)\s*(?:حبة)?\s*\+\s*(\d+)', offer_msg)
                                        if match: buy_qty = int(match.group(1)); get_qty = int(match.group(2))
                                    else:
                                        match_price = re.search(r'(\d+)\s*حبة بسعر', offer_msg)
                                        if match_price:
                                            buy_qty = int(match_price.group(1)) - 1; get_qty = 1
                                            L = float(first_row.get('السعر قبل شامل', 0)) if pd.notna(first_row.get('السعر قبل شامل')) else 0
                                            M = float(first_row.get('السعر بعد شامل', 0)) if pd.notna(first_row.get('السعر بعد شامل')) else 0
                                            C = float(first_row.get('سعر بيع المنتج', 0)) if pd.notna(first_row.get('سعر بيع المنتج')) else 0
                                            if C > 0: discount_amount = round(((L - M) / C) * 100, 2)
                                        else:
                                            match_pct = re.search(r'خصم\s*(\d+(\.\d+)?)%', offer_msg)
                                            if match_pct: discount_amount = float(match_pct.group(1))
                                            match_buy_get = re.search(r'على الحبة (الثانية|الثالثة|الرابعة)', offer_msg)
                                            if match_buy_get:
                                                buy_qty = {"الثانية": 1, "الثالثة": 2, "الرابعة": 3}.get(match_buy_get.group(1), 1)
                                                get_qty = 1

                                    row = {
                                        "الإجراء": action, "معرف العرض": offer_id, "اسم العرض": offer_name, "نوع العرض": "اذا اشترى العميل X يحصل على Y",
                                        "المنصة": "متصفح وتطبيق المتجر", "تطبيق على": "منتجات مختارة", "تاريخ البدء": start_date, "تاريخ الانتهاء": end_date,
                                        "تطبيق مع كوبون": "نعم", "الحد الأقصى للخصم": 0, "الحد الأدنى للشراء": 0, "الحد الأدنى للكمية": 0,
                                        "مجموعات العملاء": "", "نوع شراء X": "منتج", "كمية شراء X": buy_qty, "عناصر شراء X (IDs)": id_str,
                                        "نوع عرض Y": "منتج", "كمية عرض Y": get_qty, "عناصر عرض Y (IDs)": id_str, "نوع الخصم": salla_discount_type, 
                                        "قيمة الخصم": discount_amount, "رسالة العرض": offer_msg, "حالة العرض": "نشط"
                                    }
                                    output_rows.append(row)
                            else:
                                for idx, row in df_draft.iterrows():
                                    sku = str(row.get('رقم المنتج sku', '')).strip()
                                    if not sku: continue
                                    sku_str = sku
                                    id_str = sku_map.get(sku, '')
                                    
                                    if offer_name_col and offer_name_col in df_draft.columns: offer_msg = str(row.get(offer_name_col, ''))
                                    else: offer_msg = str(row.get('اسم العرض', '')) if 'اسم العرض' in df_draft.columns else ''
                                    if not offer_msg: offer_msg = f"{sku_str} / عرض"
                                    
                                    offer_name = f"{sku_str} / {offer_msg}"
                                    action = "تحديث" if offer_name in offer_name_map else "إنشاء"
                                    offer_id = offer_name_map.get(offer_name, "")
                                    start_date = str(row.get('تاريخ البدء', '')) if pd.notna(row.get('تاريخ البدء')) else ''
                                    end_date = str(row.get('تاريخ الانتهاء', '')) if pd.notna(row.get('تاريخ الانتهاء')) else ''
                                    
                                    discount_type_raw = str(row.get('نوع الخصم', '')).strip()
                                    salla_discount_type = "منتج مجاني" if discount_type_raw == "منتج مجاني" else "خصم بنسبة"
                                    discount_amount = 0.0; buy_qty = 1; get_qty = 1
                                    
                                    if salla_discount_type == "منتج مجاني":
                                        match = re.search(r'(\d+)\s*(?:حبة)?\s*\+\s*(\d+)', offer_msg)
                                        if match: buy_qty = int(match.group(1)); get_qty = int(match.group(2))
                                    else:
                                        match_price = re.search(r'(\d+)\s*حبة بسعر', offer_msg)
                                        if match_price:
                                            buy_qty = int(match_price.group(1)) - 1; get_qty = 1
                                            L = float(row.get('السعر قبل شامل', 0)) if pd.notna(row.get('السعر قبل شامل')) else 0
                                            M = float(row.get('السعر بعد شامل', 0)) if pd.notna(row.get('السعر بعد شامل')) else 0
                                            C = float(row.get('سعر بيع المنتج', 0)) if pd.notna(row.get('سعر بيع المنتج')) else 0
                                            if C > 0: discount_amount = round(((L - M) / C) * 100, 2)
                                        else:
                                            match_pct = re.search(r'خصم\s*(\d+(\.\d+)?)%', offer_msg)
                                            if match_pct: discount_amount = float(match_pct.group(1))
                                            match_buy_get = re.search(r'على الحبة (الثانية|الثالثة|الرابعة)', offer_msg)
                                            if match_buy_get:
                                                buy_qty = {"الثانية": 1, "الثالثة": 2, "الرابعة": 3}.get(match_buy_get.group(1), 1)
                                                get_qty = 1

                                    row_data = {
                                        "الإجراء": action, "معرف العرض": offer_id, "اسم العرض": offer_name, "نوع العرض": "اذا اشترى العميل X يحصل على Y",
                                        "المنصة": "متصفح وتطبيق المتجر", "تطبيق على": "منتجات مختارة", "تاريخ البدء": start_date, "تاريخ الانتهاء": end_date,
                                        "تطبيق مع كوبون": "نعم", "الحد الأقصى للخصم": 0, "الحد الأدنى للشراء": 0, "الحد الأدنى للكمية": 0,
                                        "مجموعات العملاء": "", "نوع شراء X": "منتج", "كمية شراء X": buy_qty, "عناصر شراء X (IDs)": id_str,
                                        "نوع عرض Y": "منتج", "كمية عرض Y": get_qty, "عناصر عرض Y (IDs)": id_str, "نوع الخصم": salla_discount_type, 
                                        "قيمة الخصم": discount_amount, "رسالة العرض": offer_msg, "حالة العرض": "نشط"
                                    }
                                    output_rows.append(row_data)
                                
                            df_out = pd.DataFrame(output_rows)
                            st.success("✅ تمت المعالجة بنجاح! راجع النتائج بالأسفل ثم قم بتحميل الملف الجاهز لرفعه لسلة.")
                            st.dataframe(df_out, use_container_width=True)
                            buf = io.BytesIO()
                            df_out.to_excel(buf, index=False)
                            st.download_button("📥 تحميل ملف العروض الجاهز للرفع 🚀", data=buf.getvalue(), file_name="Generated_Salla_Offers.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
                        except Exception as e: 
                            st.error(f"خطأ أثناء معالجة المسودة: {str(e)}")

    st.divider()

    # ==========================================
    # --- رفع واستيراد ملف العروض الجاهز للسلة ---
    # ==========================================
    st.info("💡 **تنبيه هام قبل التصدير:** سلة تُرجع بيانات العروض مختصرة. لضمان ظهور (أرقام المنتجات، SKUs، ونسب الخصم) بشكل دقيق وكامل في ملف الإكسيل، يرجى الضغط على زر (سحب التفاصيل الدقيقة) أولاً.")
    col_dl, col_ex, col_sync_details = st.columns([1, 1, 1.2])
    with col_dl: 
        st.download_button("📥 تنزيل قالب سلة الافتراضي للعروض", data=generate_salla_excel_template(), file_name="Salla_Offers_Template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_sync_details:
        if st.button("🔄 سحب التفاصيل الدقيقة (للتصدير)", use_container_width=True):
            import time # ✅ تم إضافة الاستدعاء هنا مباشرة لمنع خطأ UnboundLocalError
            with st.spinner("جاري سحب التفاصيل الكاملة لجميع العروض (يستغرق بضع ثوانٍ)..."):
                progress_bar = st.progress(0)
                status_text = st.empty()
                total_o = len(st.session_state["all_offers"])
                for i, o in enumerate(st.session_state["all_offers"]):
                    status_text.info(f"📥 سحب بيانات العرض {i+1} من {total_o}...")
                    det_res = safe_api_request("GET", f"{SALLA_API_URL}/{o['id']}", headers)
                    if det_res and det_res.get("data"):
                        st.session_state["all_offers"][i] = det_res["data"]
                    progress_bar.progress((i + 1) / total_o)
                    time.sleep(0.2) # حماية من حظر الـ API
                status_text.success("✅ تم تحديث جميع تفاصيل العروض بنجاح! يمكنك التصدير الآن.")
                time.sleep(1)
                st.rerun()
    with col_ex:
        if raw_offers: 
            # ✅ استخدام دالة التصدير المتقدمة المخصصة
            st.download_button("📥 تصدير قائمة العروض الحالية", data=get_advanced_export_excel(raw_offers, st.session_state.get('all_products', [])), file_name=f"Offers_Export_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")

    uploaded_file = st.file_uploader("📂 رفع ملف العروض الجاهز (XLSX) لرفعه للمتجر مباشرة:", type=["xlsx"])
    if uploaded_file:
        try:
            df_user = pd.read_excel(uploaded_file)
            st.dataframe(df_user, use_container_width=True)
            if st.button("🚀 تأكيد معالجة ونشر ملف العروض المرفوعة", use_container_width=True):
                res = process_excel_import(df_user)
                for m in res["success"]: st.success(m)
                for m in res["errors"]: st.error(m)
                st.rerun()
        except Exception as e: 
            st.error(f"خطأ في القراءة: {str(e)}")

    # ==========================================
    # 🔍 الفلاتر الخاصة بالعروض
    # ==========================================
    with st.container(border=True):
        st.markdown("<div style='background: #0F1C2E; color: white; padding: 10px; border-radius: 8px; text-align: center; font-weight: bold;'>🔍 أدوات التصفية والبحث المتقدمة</div>", unsafe_allow_html=True)
        
        # استخراج تواريخ الانتهاء الفريدة من العروض
        avail_dates = set()
        for o in raw_offers:
            ed = safe_parse_date(o.get('expiry_date'))
            if ed: avail_dates.add(ed.strftime('%Y-%m-%d'))
        date_options = ["الكل", "بدون تاريخ"] + sorted(list(avail_dates))
        
        # ✅ استخراج الماركات المتاحة من العروض المشمولة والمنتجات
        avail_brands = set()
        po_map = st.session_state.get("product_offers_map", {})
        for p in st.session_state.get("all_products", []):
            if po_map.get(str(p.get('id', ''))):
                brand = p.get('brand')
                if isinstance(brand, dict) and brand.get('name'):
                    avail_brands.add(brand.get('name'))
        brand_options = sorted(list(avail_brands))
        
        col_search, col_status, col_brands = st.columns([2, 1, 1])
        with col_search: search_offer = st.text_input("🔎 ابحث باسم العرض أو بالمعرف:", key="filter_search_input")
        with col_status: status_filter = st.selectbox("📌 حالة العرض:", ["الكل", "نشط", "غير نشط"], key="filter_status_select")
        # ✅ فلتر الماركات المشمولة في العرض
        with col_brands: brands_filter = st.multiselect("🏢 عروض تشمل منتجات للماركات:", options=brand_options, placeholder="اختر ماركة أو أكثر...", key="filter_brands")
    
        col_date, col_feat, col_over = st.columns(3)
        with col_date: filter_date_str = st.selectbox("📅 تاريخ الانتهاء:", date_options, key="filter_date_select")
        with col_feat: filter_featured = st.selectbox("⭐ العروض المميزة:", ["الكل"] + list(st.session_state.get("featured_offer_groups", {}).keys()), key="filter_featured")
        with col_over: 
            filter_overlap = st.checkbox("🔄 فحص التداخل (منتجات مكررة)", key="f_overlap")
            filter_discounted_only = st.checkbox("💰 عروض بها منتجات مخفضة", key="f_disc_prod")
            
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1: type_filter = st.selectbox("📊 نوع العرض:", ["الكل"] + list(OFFER_TYPES_MAP.values()), key="type_filter")
        with col_f2: channel_filter = st.selectbox("📺 قناة النشر:", ["الكل"] + list(CHANNELS_MAP.values()), key="channel_filter")
        with col_f3: applied_filter = st.selectbox("🎯 تطبيق على:", ["الكل"] + list(APPLIED_TO_MAP.values()), key="applied_filter")
        
        
        

    now_ksa = datetime.now() + timedelta(hours=3)
    overlapping_offer_ids = set()

    if filter_overlap:
        with st.spinner("🔄 جاري تحليل تداخل المنتجات..."):
            product_offer_map = {}
            for o in raw_offers:
                o_id = o.get('id')
                if o.get('status') != 'active': continue
                full_res = safe_api_request("GET", f"{SALLA_API_URL}/{o_id}", headers)
                if full_res and full_res.get('data'):
                    p_ids = set()
                    for p in full_res['data'].get('buy', {}).get('products', []): p_ids.add(str(p.get('id', p) if isinstance(p, dict) else p))
                    for p in full_res['data'].get('get', {}).get('products', []): p_ids.add(str(p.get('id', p) if isinstance(p, dict) else p))
                    for pid in p_ids:
                        if pid not in product_offer_map: product_offer_map[pid] = []
                        product_offer_map[pid].append(o_id)
            for pid, o_ids in product_offer_map.items():
                if len(o_ids) > 1: overlapping_offer_ids.update(o_ids)
            if not overlapping_offer_ids: st.success("✅ لا يوجد تداخل في منتجات العروض النشطة.")
            else: st.warning(f"⚠️ تم العثور على {len(overlapping_offer_ids)} عرض متداخل.")

    filtered_offers = []
    
    # ✅ إنشاء قاموس عكسي (Offer ID -> List of Product Objects) لتسهيل الفلترة السريعة
    offer_to_products = {}
    for p in st.session_state.get("all_products", []):
        p_id_str = str(p.get('id', ''))
        for offer_info in po_map.get(p_id_str, []):
            o_id_str = str(offer_info.get('id', ''))
            if o_id_str not in offer_to_products:
                offer_to_products[o_id_str] = []
            offer_to_products[o_id_str].append(p)
            
    for offer in raw_offers:
        offer_id = str(offer.get('id', 'N/A'))
        offer_name = offer.get('name', 'عرض بدون اسم')
        status = offer.get('status', 'inactive')
        start_date = safe_parse_date(offer.get('start_date'))
        exp_date_obj = safe_parse_date(offer.get('expiry_date'))
        exp_date_str_val = exp_date_obj.strftime('%Y-%m-%d') if exp_date_obj else ""
    
        if search_offer:
            if search_offer.lower() not in offer_name.lower() and search_offer not in str(offer_id): continue
        if status_filter == "نشط" and status != "active": continue
        if status_filter == "غير نشط" and status == "active": continue
        if filter_overlap and int(offer_id) not in overlapping_offer_ids: continue

        # ✅ تطبيق فلتر "الماركات" (يقبل العرض إذا احتوى على منتج واحد على الأقل للماركة المحددة)
        if brands_filter:
            offer_brands = set()
            for p in offer_to_products.get(offer_id, []):
                brand = p.get('brand')
                if isinstance(brand, dict) and brand.get('name'):
                    offer_brands.add(brand.get('name'))
                    
            if not any(b in brands_filter for b in offer_brands):
                continue
                
        if filter_date_str != "الكل":
            if filter_date_str == "بدون تاريخ" and exp_date_str_val != "": continue
            elif filter_date_str != "بدون تاريخ" and exp_date_str_val != filter_date_str: continue
            
        if filter_featured != "الكل":
            if int(offer_id) not in st.session_state.get("featured_offer_groups", {}).get(filter_featured, []): continue
        if type_filter != "الكل":
            offer_type_ar = OFFER_TYPES_MAP.get(offer.get('offer_type', ''), '')
            if offer_type_ar != type_filter: continue
        if channel_filter != "الكل":
            channel_ar = CHANNELS_MAP.get(offer.get('applied_channel', ''), '')
            if channel_ar != channel_filter: continue
        if applied_filter != "الكل":
            applied_ar = APPLIED_TO_MAP.get(offer.get('applied_to', ''), '')
            if applied_ar != applied_filter: continue
            
        # ✅ تطبيق فلتر "عروض بها منتجات مخفضة" (عبر الـ Checkbox)
        if filter_discounted_only:
            has_discounted_product = False
            for p in offer_to_products.get(offer_id, []):
                pr = safe_float(p.get('price', 0))
                reg = safe_float(p.get('regular_price', 0))
                sal = safe_float(p.get('sale_price', 0))
                if sal > 0 and sal < (reg if reg > 0 else pr):
                    has_discounted_product = True
                    break
            
            # إذا كان الـ Checkbox مفعلاً والعرض لا يحتوي على منتجات مخفضة، نقوم بتخطيه
            if not has_discounted_product: continue
    
        filtered_offers.append(offer)

    st.session_state["filtered_offers"] = filtered_offers

    st.markdown(f"<div style='background: #f0f4f8; padding: 8px 16px; border-radius: 8px; margin-bottom: 14px; border-right: 4px solid #00b4d8;'><strong>📊 عدد العروض المطابقة للبحث: {len(filtered_offers)} عرض</strong></div>", unsafe_allow_html=True)

    # ==========================================
    # ✅ 3. أزرار التنزيل و (إجراءات العروض المفلترة)
    # ==========================================
    if filtered_offers:
        col_dl_filt, col_act_filt = st.columns(2)
        with col_dl_filt:
            # ✅ استخدام دالة التصدير المتقدمة المخصصة هنا أيضاً
            st.download_button("📥 تحميل العروض المفلترة (Excel)", data=get_advanced_export_excel(filtered_offers, st.session_state.get('all_products', [])), file_name=f"filtered_offers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_filtered_offers", type="primary", use_container_width=True)
        with col_act_filt:
            with st.popover("⚙️ إجراءات العروض المفلترة", use_container_width=True):
                st.markdown(f"<div style='text-align:center; margin-bottom:10px;'><b>تطبيق إجراء جماعي على ({len(filtered_offers)}) عرض</b></div>", unsafe_allow_html=True)
                
                bulk_action = st.radio("اختر الإجراء المطلوب تنفيذه دفعة واحدة:", [
                    "▶️ تفعيل العروض",
                    "📅 تمديد العروض",
                    "🛑 إيقاف ومسح العناوين الترويجية من المنتجات",
                    "🗑️ حذف العروض نهائياً"
                ], key="bulk_action_offers")
                
                new_expiry_str = None
                if bulk_action == "📅 تمديد العروض":
                    cd1, cd2 = st.columns(2)
                    with cd1: nd = st.date_input("التاريخ الجديد:")
                    with cd2: nt = st.time_input("الوقت الجديد:")
                    new_expiry_str = datetime.combine(nd, nt).strftime('%Y-%m-%d %H:%M:%S')
                
                if "حذف" in bulk_action:
                    st.error("🚨 سيتم الحذف نهائياً من المتجر!")
                    confirm_msg = "☑️ أوافق على الحذف"
                else:
                    confirm_msg = "☑️ تأكيد تنفيذ الإجراء المختار"
                    
                confirm_bulk = st.checkbox(confirm_msg, key="confirm_bulk_action_offers")
                
                if st.button("🚀 تنفيذ الإجراء", type="primary", disabled=not confirm_bulk, use_container_width=True, key="execute_bulk_action_offers"):
                    # ✅ إضافة شريط التقدم التفاعلي وعداد العروض
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    success_c = 0
                    total_offers = len(filtered_offers)
                    
                    for idx, off in enumerate(filtered_offers):
                        oid = off['id']
                        status_text.info(f"⏳ جاري تنفيذ الإجراء على العرض {idx+1} من {total_offers}...")
                        
                        if "تفعيل" in bulk_action:
                            if safe_api_request("PUT", f"{SALLA_API_URL}/{oid}/status", headers, json={"status": "active"}): success_c += 1
                        elif "تمديد" in bulk_action:
                            full = safe_api_request("GET", f"{SALLA_API_URL}/{oid}", headers)
                            if full and full.get('data'):
                                payload = rebuild_offer_payload(full['data'], {"expiry_date": new_expiry_str})
                                if safe_api_request("PUT", f"{SALLA_API_URL}/{oid}", headers, json=payload): success_c += 1
                        elif "إيقاف ومسح" in bulk_action:
                            # 1. إيقاف العرض
                            safe_api_request("PUT", f"{SALLA_API_URL}/{oid}/status", headers, json={"status": "inactive"})
                            # 2. مسح العناوين الترويجية
                            full = safe_api_request("GET", f"{SALLA_API_URL}/{oid}", headers)
                            if full and full.get('data'):
                                pids = set()
                                for p in full['data'].get('buy', {}).get('products', []):
                                    if str(p.get('id', p)).isdigit(): pids.add(int(p.get('id', p)))
                                for p in full['data'].get('get', {}).get('products', []):
                                    if str(p.get('id', p)).isdigit(): pids.add(int(p.get('id', p)))
                                for pid in pids:
                                    update_product_promotions_secure(pid, "", "", headers)
                            success_c += 1
                        elif "حذف" in bulk_action:
                            if safe_api_request("DELETE", f"{SALLA_API_URL}/{oid}", headers): success_c += 1
                            
                        # ✅ تحديث شريط التقدم وإضافة مهلة زمنية للحماية من حظر API سلة
                        progress_bar.progress((idx + 1) / total_offers)
                        import time; time.sleep(0.3)
                        
                    status_text.success(f"✅ تم تنفيذ الإجراء على {success_c} عرض بنجاح!")
                    import time; time.sleep(1.5)
                    if "all_offers" in st.session_state: del st.session_state["all_offers"] # لإجبار التحديث
                    st.rerun()
                        
    # ==========================================
    # 📄 عرض بطاقات العروض مع ترقيم الصفحات
    # ==========================================
    
    items_per_page = 10
    total_pages = max(1, (len(filtered_offers) + items_per_page - 1) // items_per_page)
    
    if "offers_page" not in st.session_state: st.session_state["offers_page"] = 1
    if st.session_state["offers_page"] > total_pages: st.session_state["offers_page"] = total_pages
    
    start_idx = (st.session_state["offers_page"] - 1) * items_per_page
    end_idx = start_idx + items_per_page
    displayed_offers = filtered_offers[start_idx:end_idx]
    
    def render_pagination_top():
        col_prev, col_page, col_next = st.columns([1, 2, 1])
        with col_prev:
            if st.button("⬅️ السابق", disabled=st.session_state["offers_page"] == 1, use_container_width=True, key="offers_prev_top"):
                st.session_state["offers_page"] -= 1; st.rerun()
        with col_page: st.markdown(f"<h4 style='text-align:center;'>📄 صفحة {st.session_state['offers_page']} من {total_pages}</h4>", unsafe_allow_html=True)
        with col_next:
            if st.button("التالي ➡️", disabled=st.session_state["offers_page"] == total_pages, use_container_width=True, key="offers_next_top"):
                st.session_state["offers_page"] += 1; st.rerun()

    def render_pagination_bottom():
        col_prev, col_page, col_next = st.columns([1, 2, 1])
        with col_prev:
            if st.button("⬅️ السابق", disabled=st.session_state["offers_page"] == 1, use_container_width=True, key="offers_prev_bottom"):
                st.session_state["offers_page"] -= 1; st.rerun()
        with col_page: st.markdown(f"<h4 style='text-align:center;'>📄 صفحة {st.session_state['offers_page']} من {total_pages}</h4>", unsafe_allow_html=True)
        with col_next:
            if st.button("التالي ➡️", disabled=st.session_state["offers_page"] == total_pages, use_container_width=True, key="offers_next_bottom"):
                st.session_state["offers_page"] += 1; st.rerun()
    
    st.markdown("---")
    render_pagination_top()
    st.markdown("---")
    
    inv_type_map = {"product": "منتجات", "category": "تصنيفات", "brand": "ماركات"}
    type_options_ar = ["منتجات", "تصنيفات", "ماركات"]
    type_map = {"منتجات": "product", "تصنيفات": "category", "ماركات": "brand"}
    
    def get_promo_badge(pid):
        for pr in st.session_state.get("all_products", []):
            if str(pr.get('id')) == str(pid):
                promo_obj = pr.get('promotion')
                promo_title = pr.get('promotion_title', "")
                if isinstance(promo_obj, dict): promo_title = promo_obj.get('title', promo_title)
                if promo_title: return f"<span style='color:#b45309; font-size:11px; background:#fef3c7; padding:2px 6px; border-radius:4px; margin-right:4px;'>🔖 {promo_title}</span>"
        return ""
    
    for idx, offer in enumerate(displayed_offers):
        offer_id = offer.get('id', 'N/A')
        offer_name = offer.get('name', 'عرض بدون اسم')
        status = offer.get('status', 'inactive')
        
        with st.spinner(f"جاري جلب تفاصيل العرض: {offer_name}..."):
            detailed_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
            offer_data = detailed_res.get("data", offer) if detailed_res else offer

        o_type_raw = offer_data.get('offer_type', '')
        o_channel_raw = offer_data.get('applied_channel', 'browser_and_application')
        o_applied_raw = offer_data.get('applied_to', 'product')
        start_date = safe_parse_date(offer_data.get('start_date'))
        exp_date = safe_parse_date(offer_data.get('expiry_date'))
        
        badge = "🟢 نشط بالمتجر" if status == "active" else "🔴 متوقف حالياً"
        if start_date and start_date > now_ksa: exp_badge = "⏳ لم يبدأ بعد"
        elif exp_date and exp_date < now_ksa: exp_badge = "⚠️ منتهي الصلاحية"
        else: exp_badge = "⏳ ساري الصلاحية"
        
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #0f1c2e 0%, #1a365d 100%); padding: 14px 20px; border-radius: 12px 12px 0px 0px; margin-top: 25px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px; border-bottom: 3px solid #00b4d8;">
            <span style="color: #ffffff; font-weight: bold; font-size: 16px;">🎯 {offer_name} (ID: {offer_id})</span>
            <div style="display: flex; gap: 8px; flex-wrap: wrap;">
                <span style="background: rgba(255,255,255,0.2); color: #fff; padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight:600;">{badge}</span>
                <span style="background: rgba(255,193,7,0.25); color: #ffca28; padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight:600;">{exp_badge}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        cx, cy = st.columns(2)
        with cx:
            st.markdown(f"⚙️ **نوع العرض:** `{OFFER_TYPES_MAP.get(o_type_raw, o_type_raw)}`")
            st.markdown(f"📺 **قناة نشر العرض:** `{CHANNELS_MAP.get(o_channel_raw, o_channel_raw)}`")
            st.markdown(f"🎯 **يتم تطبيق العرض على:** `{APPLIED_TO_MAP.get(o_applied_raw, o_applied_raw)}`")
            st.markdown(f"📅 **توقيت بدء العرض:** `{offer_data.get('start_date', 'غير محدد')}`")
            st.markdown(f"📅 **توقيت انتهاء العرض:** `{offer_data.get('expiry_date', 'بدون تاريخ (مستمر)')}`")
        with cy:
            st.markdown(f"🛡️ **الحد الأقصى للخصم:** `{offer_data.get('max_discount_amount', 0)} SAR` | 💵 **الحد الأدنى للشراء:** `{offer_data.get('min_purchase_amount', 0)} SAR`")
            c_groups_raw = offer_data.get('customer_groups', [])
            c_groups_rendered = ", ".join([str(g.get('name', g.get('id', g))) if isinstance(g, dict) else str(g) for g in c_groups_raw]) if c_groups_raw else "كل المجموعات"
            st.markdown(f"👥 **مجموعة العملاء المستهدفة:** `{c_groups_rendered}`")
            st.markdown(f"**🔖 تطبيق العرض مع كوبون التخفيض؟** `{'نعم' if offer_data.get('applied_with_coupon') else 'لا يطبق'}`")
            st.markdown(f"**📢 نص رسالة العرض:** *{offer_data.get('message', 'لا توجد رسالة مرفقة')}*")
                
        st.markdown("<hr style='margin: 15px 0; border-top: 1px dashed #e2e8f0;'>", unsafe_allow_html=True)
        col_x, col_y = st.columns(2)
            
        with col_x:
            st.markdown("<b style='color:#0f1c2e;'>🛒 مجموعة الشراء (X) - [إذا اشترى العميل]:</b>", unsafe_allow_html=True)
            buy_obj = offer_data.get('buy', {})
            b_type_raw = buy_obj.get("type", "product")
            if isinstance(b_type_raw, dict): b_type_raw = b_type_raw.get("id", "product")
            st.markdown(f"<div style='margin-bottom:8px; font-size:13px; color:#64748b;'>مطبق على: <b>{inv_type_map.get(b_type_raw, 'منتجات')}</b></div>", unsafe_allow_html=True)
                
            buy_html = "<div style='background:#f8fafc; padding:12px; border-radius:8px; border:1px solid #e2e8f0; max-height: 180px; overflow-y: auto;'><ul style='margin:0; padding-right:15px; font-size:13px; line-height:1.6;'>";
            has_items = False
            for p in buy_obj.get('products', []):
                p_id = p.get('id', p) if isinstance(p, dict) else p
                p_name = p.get('name', 'بدون اسم') if isinstance(p, dict) else 'منتج'
                p_sku = p.get('sku', 'لا يوجد') if isinstance(p, dict) else 'لا يوجد'
                promo_badge = get_promo_badge(p_id)
                buy_html += f"<li style='margin-bottom:8px;'>📦 <b>{p_name}</b><br><span style='color:#64748b; font-size:11px; background:#e2e8f0; padding:2px 6px; border-radius:4px;'>SKU: {p_sku}</span> <span style='color:#64748b; font-size:11px; background:#e2e8f0; padding:2px 6px; border-radius:4px;'>ID: {p_id}</span> {promo_badge}</li>"
                has_items = True
            for c in buy_obj.get('categories', []):
                c_id = c.get('id', c) if isinstance(c, dict) else c
                c_name = c.get('name', 'بدون اسم') if isinstance(c, dict) else 'تصنيف'
                buy_html += f"<li style='margin-bottom:8px;'>📁 <b>{c_name}</b><br><span style='color:#64748b; font-size:11px; background:#e2e8f0; padding:2px 6px; border-radius:4px;'>ID: {c_id}</span></li>"
                has_items = True
            for b in buy_obj.get('brands', []):
                b_id = b.get('id', b) if isinstance(b, dict) else b
                b_name = b.get('name', 'بدون اسم') if isinstance(b, dict) else 'ماركة'
                buy_html += f"<li style='margin-bottom:8px;'>🏢 <b>{b_name}</b><br><span style='color:#64748b; font-size:11px; background:#e2e8f0; padding:2px 6px; border-radius:4px;'>ID: {b_id}</span></li>"
                has_items = True
            buy_html += "</ul></div>"
            if has_items: st.markdown(buy_html, unsafe_allow_html=True)
            else: st.info("جميع الأصناف المشمولة")
            st.caption(f"الكمية المطلوبة: {buy_obj.get('quantity', 1)} قطعة")
                
        with col_y:
            st.markdown("<b style='color:#0f1c2e;'>🎁 مجموعة المنح والهدية (Y) - [يحصل على]:</b>", unsafe_allow_html=True)
            get_obj = offer_data.get('get', {})               
            g_type_raw = get_obj.get("type", "product")
            if isinstance(g_type_raw, dict): g_type_raw = g_type_raw.get("id", "product")
            st.markdown(f"<div style='margin-bottom:8px; font-size:13px; color:#64748b;'>مطبق على: <b>{inv_type_map.get(g_type_raw, 'منتجات')}</b></div>", unsafe_allow_html=True)
                
            get_html = "<div style='background:#f0fdf4; padding:12px; border-radius:8px; border:1px solid #bbf7d0; max-height: 180px; overflow-y: auto;'><ul style='margin:0; padding-right:15px; font-size:13px; line-height:1.6;'>";
            has_items_y = False
            for p in get_obj.get('products', []):
                p_id = p.get('id', p) if isinstance(p, dict) else p
                p_name = p.get('name', 'بدون اسم') if isinstance(p, dict) else 'منتج'
                p_sku = p.get('sku', 'لا يوجد') if isinstance(p, dict) else 'لا يوجد'
                promo_badge = get_promo_badge(p_id)
                get_html += f"<li style='margin-bottom:8px;'>📦 <b>{p_name}</b><br><span style='color:#166534; font-size:11px; background:#dcfce7; padding:2px 6px; border-radius:4px;'>SKU: {p_sku}</span> <span style='color:#166534; font-size:11px; background:#dcfce7; padding:2px 6px; border-radius:4px;'>ID: {p_id}</span> {promo_badge}</li>"
                has_items_y = True
            for c in get_obj.get('categories', []):
                c_id = c.get('id', c) if isinstance(c, dict) else c
                c_name = c.get('name', 'بدون اسم') if isinstance(c, dict) else 'تصنيف'
                get_html += f"<li style='margin-bottom:8px;'>📁 <b>{c_name}</b><br><span style='color:#166534; font-size:11px; background:#dcfce7; padding:2px 6px; border-radius:4px;'>ID: {c_id}</span></li>"
                has_items_y = True
            for b in get_obj.get('brands', []):
                b_id = b.get('id', b) if isinstance(b, dict) else b
                b_name = b.get('name', 'بدون اسم') if isinstance(b, dict) else 'ماركة'
                get_html += f"<li style='margin-bottom:8px;'>🏢 <b>{b_name}</b><br><span style='color:#166534; font-size:11px; background:#dcfce7; padding:2px 6px; border-radius:4px;'>ID: {b_id}</span></li>"
                has_items_y = True
            get_html += "</ul></div>"
            if has_items_y: st.markdown(get_html, unsafe_allow_html=True)
            else: st.success("جميع الأصناف المشمولة")
            st.caption(f"كمية المنح/الخصم: {get_obj.get('quantity', 1)} قطعة")
            if get_obj.get('discount_amount'): st.markdown(f"🔥 **قيمة/نسبة الخصم :** `{get_obj.get('discount_amount')}`")

        st.markdown("<br>", unsafe_allow_html=True)
        b1, b2, b3 = st.columns(3)
        with b1:
            t_status = "inactive" if status == "active" else "active"
            lbl = "🛑 إيقاف العرض" if status == "active" else "▶️ إعادة تفعيل العرض"
            if st.button(lbl, key=f"t_st_{offer_id}_{idx}", use_container_width=True):
                if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}/status", headers, json={"status": t_status}):
                    for i, o in enumerate(st.session_state["all_offers"]):
                        if str(o.get('id')) == str(offer_id): st.session_state["all_offers"][i]['status'] = t_status
                    st.rerun()
        with b2:
            if st.button("🔖 عكس تطبيق العرض مع الكوبون ⏯", key=f"t_cp_{offer_id}_{idx}", use_container_width=True):
                new_coupon_status = not offer_data.get('applied_with_coupon', False)
                if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json={"applied_with_coupon": new_coupon_status}):
                    for i, o in enumerate(st.session_state["all_offers"]):
                        if str(o.get('id')) == str(offer_id): st.session_state["all_offers"][i]['applied_with_coupon'] = new_coupon_status
                    st.rerun()
        with b3:
            if st.button("🗑️ حذف العرض بالكامل", key=f"t_dl_{offer_id}_{idx}", use_container_width=True, type="primary"):
                if safe_api_request("DELETE", f"{SALLA_API_URL}/{offer_id}", headers):
                    st.session_state["all_offers"] = [o for o in st.session_state["all_offers"] if str(o.get('id')) != str(offer_id)]
                    st.rerun()

        with st.expander("✏️ تعديل ومراجعة العرض الترويجي", expanded=False):
            ed_name = st.text_input("إسم العرض:", value=offer_name, key=f"ed_n_{offer_id}_{idx}")
            ed_msg = st.text_input("رسالة العرض:", value=offer_data.get('message', ''), key=f"ed_m_{offer_id}_{idx}")
                
            ec1, ec2, ec3 = st.columns(3)
            with ec1:
                current_type_idx = list(OFFER_TYPES_MAP.keys()).index(o_type_raw) if o_type_raw in OFFER_TYPES_MAP else 0
                ed_type_ar = st.selectbox("نوع العرض:", list(OFFER_TYPES_MAP.values()), index=current_type_idx, key=f"ed_t_ar_{offer_id}_{idx}")
                ed_applied_ar = st.selectbox("تطبيق العرض على:", list(APPLIED_TO_MAP.values()), index=list(APPLIED_TO_MAP.keys()).index(o_applied_raw) if o_applied_raw in APPLIED_TO_MAP else 0, key=f"ed_app_ar_{offer_id}_{idx}")
            with ec2:
                current_chan_idx = list(CHANNELS_MAP.keys()).index(o_channel_raw) if o_channel_raw in CHANNELS_MAP else 0
                ed_chan_ar = st.selectbox("منصة النشر:", list(CHANNELS_MAP.values()), index=current_chan_idx, key=f"ed_ch_ar_{offer_id}_{idx}")
                ed_status = st.selectbox("حالة العرض:", ["active", "inactive"], index=0 if status == "active" else 1, format_func=lambda x: "مفعل" if x == "active" else "مسودة", key=f"ed_status_field_{offer_id}_{idx}")
            with ec3:
                ed_coupon = st.selectbox("تطبيق مع كوبون؟", ["لا", "نعم"], index=1 if offer_data.get('applied_with_coupon') else 0, key=f"ed_c_{offer_id}_{idx}")

            selected_ed_type_key = [k for k, v in OFFER_TYPES_MAP.items() if v == ed_type_ar][0]
            selected_ed_chan_key = [k for k, v in CHANNELS_MAP.items() if v == ed_chan_ar][0]
            selected_ed_app_key = [k for k, v in APPLIED_TO_MAP.items() if v == ed_applied_ar][0]
            ed_cust_groups = st.text_input("مجموعة العملاء (IDs):", value=",".join([str(g.get('id', g)) if isinstance(g, dict) else str(g) for g in offer_data.get('customer_groups', [])]), key=f"ed_cg_{offer_id}_{idx}")

            ecc1, ecc2, ecc3 = st.columns(3)
            with ecc1: ed_max_discount = st.number_input("أقصى خصم (SAR):", min_value=0.0, value=safe_float(offer_data.get('max_discount_amount', 0.0)), key=f"ed_max_d_{offer_id}_{idx}")
            with ecc2: ed_min_purchase = st.number_input("أدنى شراء (SAR):", min_value=0.0, value=safe_float(offer_data.get('min_purchase_amount', 0.0)), key=f"ed_min_p_{offer_id}_{idx}")
            with ecc3: ed_min_items = st.number_input("أدنى كمية:", min_value=0, value=int(safe_float(offer_data.get('min_items_count', 0.0))), key=f"ed_min_i_{offer_id}_{idx}")

            if selected_ed_type_key == "buy_x_get_y":
                eq1, eq2 = st.columns(2)
                with eq1:
                    ed_buy_type_ar = st.selectbox("نوع شراء X:", type_options_ar, index=type_options_ar.index(inv_type_map.get(b_type_raw, "منتجات")), key=f"ed_bt_{offer_id}_{idx}")
                    ed_buy_type = type_map[ed_buy_type_ar]
                    ed_buy_qty = st.number_input("كمية الشراء (X):", min_value=1, value=int(buy_obj.get('quantity', 1)), key=f"ed_bq_{offer_id}_{idx}")
                    existing_buy_ids = [i.get('id', i) if isinstance(i, dict) else i for i in buy_obj.get({'product':'products','category':'categories','brand':'brands'}.get(ed_buy_type), [])]
                    ed_buy_selected_ids = render_dynamic_selection(f"تعديل {ed_buy_type_ar} الشراء (X):", ed_buy_type, existing_buy_ids, f"ed_buy_X_{offer_id}_{idx}")
                    
                with eq2:
                    ed_get_type_ar = st.selectbox("نوع عرض Y:", type_options_ar, index=type_options_ar.index(inv_type_map.get(g_type_raw, "منتجات")), key=f"ed_gt_{offer_id}_{idx}")
                    ed_get_type = type_map[ed_get_type_ar]
                    ed_get_qty = st.number_input("كمية العرض (Y):", min_value=1, value=int(get_obj.get('quantity', 1)), key=f"ed_gq_{offer_id}_{idx}")
                    existing_get_ids = [i.get('id', i) if isinstance(i, dict) else i for i in get_obj.get({'product':'products','category':'categories','brand':'brands'}.get(ed_get_type), [])]
                    ed_get_selected_ids = render_dynamic_selection(f"تعديل {ed_get_type_ar} الممنوحة (Y):", ed_get_type, existing_get_ids, f"ed_get_Y_{offer_id}_{idx}")
                    
                ed_discount_type_ar = st.selectbox("نوع الخصم Y:", ["منتج مجاني", "خصم بنسبة"], index=1 if get_obj.get('discount_type', 'free-product') == 'percentage' else 0, key=f"ed_dt_ar_{offer_id}_{idx}")
                if ed_discount_type_ar == "خصم بنسبة":
                    ed_disc_amt = st.number_input("نسبة الخصم Y (%):", min_value=1.0, max_value=100.0, value=safe_float(get_obj.get('discount_amount', 50.0)), key=f"ed_da_{offer_id}_{idx}")
                    ed_disc_type = "percentage"
                else:
                    ed_disc_amt = 0.0; ed_disc_type = "free-product"
            else:
                eq1, eq2 = st.columns(2)
                with eq1:
                    ed_disc_amt = st.number_input("قيمة/نسبة الخصم:", min_value=0.0, value=safe_float(get_obj.get('discount_amount', 10.0)), key=f"ed_da_direct_{offer_id}_{idx}")
                    ed_buy_type_ar = st.selectbox("النوع:", type_options_ar, index=type_options_ar.index(inv_type_map.get(b_type_raw, "منتجات")), key=f"ed_bt_direct_{offer_id}_{idx}")
                    ed_buy_type = type_map[ed_buy_type_ar]
                    existing_buy_ids = [i.get('id', i) if isinstance(i, dict) else i for i in buy_obj.get({'product':'products','category':'categories','brand':'brands'}.get(ed_buy_type), [])]
                    ed_buy_selected_ids = render_dynamic_selection(f"العناصر:", ed_buy_type, existing_buy_ids, f"ed_buy_direct_{offer_id}_{idx}")
                with eq2: st.caption("مباشر على الخيارات دون اشتراط هدايا")
                ed_buy_qty = 1; ed_get_type = "product"; ed_get_qty = 1; ed_get_selected_ids = []; ed_disc_type = "percentage" if selected_ed_type_key == "percentage" else "fixed_amount"

            col_ed_start_date, col_ed_start_time = st.columns(2)
            with col_ed_start_date: ed_start_date_val = st.date_input("بدء - تاريخ:", value=safe_parse_date(offer_data.get('start_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))).date() if safe_parse_date(offer_data.get('start_date')) else datetime.now().date(), key=f"ed_s_date_{offer_id}_{idx}")
            with col_ed_start_time: ed_start_time_val = st.time_input("بدء - وقت:", value=safe_parse_date(offer_data.get('start_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))).time() if safe_parse_date(offer_data.get('start_date')) else datetime.now().time(), key=f"ed_start_time_{offer_id}_{idx}", step=60)
            ed_start = datetime.combine(ed_start_date_val, ed_start_time_val).strftime('%Y-%m-%d %H:%M:%S')
                
            col_ed_end_date, col_ed_end_time = st.columns(2)
            with col_ed_end_date: ed_end_date_val = st.date_input("انتهاء - تاريخ:", value=safe_parse_date(offer_data.get('expiry_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))).date() if safe_parse_date(offer_data.get('expiry_date')) else (datetime.now() + timedelta(days=30)).date(), key=f"ed_e_date_{offer_id}_{idx}")
            with col_ed_end_time: ed_end_time_val = st.time_input("انتهاء - وقت:", value=safe_parse_date(offer_data.get('expiry_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))).time() if safe_parse_date(offer_data.get('expiry_date')) else datetime.now().time().replace(hour=23, minute=59, second=59), key=f"ed_end_time_{offer_id}_{idx}", step=60)
            ed_end = datetime.combine(ed_end_date_val, ed_end_time_val).strftime('%Y-%m-%d %H:%M:%S')
                
            if st.button("💾 اعتماد وحفظ التحديث", key=f"sv_of_{offer_id}_{idx}", type="primary", use_container_width=True):
                try:
                    cg_p_list = [int(g.strip()) for g in ed_cust_groups.split(",") if g.strip().isdigit()] if ed_cust_groups.strip() else []
                    update_payload = {
                        "name": ed_name, "message": ed_msg, "start_date": ed_start, "expiry_date": ed_end,
                        "status": ed_status, "offer_type": selected_ed_type_key, "applied_channel": selected_ed_chan_key, "applied_to": selected_ed_app_key,
                        "applied_with_coupon": ed_coupon == "نعم", "max_discount_amount": float(ed_max_discount), "min_purchase_amount": float(ed_min_purchase), "min_items_count": int(ed_min_items),
                        "customer_groups": cg_p_list, "buy": {"type": ed_buy_type, "quantity": int(ed_buy_qty)}, "get": {"type": ed_get_type, "quantity": int(ed_get_qty), "discount_type": ed_disc_type}
                    }
                    buy_cat = {'product':'products', 'category':'categories', 'brand':'brands'}[ed_buy_type]
                    if ed_buy_selected_ids: update_payload["buy"][buy_cat] = ed_buy_selected_ids
                    if selected_ed_type_key == "buy_x_get_y":
                        get_cat = {'product':'products', 'category':'categories', 'brand':'brands'}[ed_get_type]
                        if ed_get_selected_ids: update_payload["get"][get_cat] = ed_get_selected_ids
                    if ed_disc_amt > 0: update_payload["get"]["discount_amount"] = float(ed_disc_amt)
                    if safe_api_request("PUT", f"{SALLA_API_URL}/{offer_id}", headers, json=update_payload):
                        st.success("تم التحديث بنجاح!")
                        # ⚡ جلب التحديث وحقنه في الذاكرة فوراً
                        fresh_res = safe_api_request("GET", f"{SALLA_API_URL}/{offer_id}", headers)
                        if fresh_res and fresh_res.get('data'):
                            for i, o in enumerate(st.session_state["all_offers"]):
                                if str(o.get('id')) == str(offer_id): st.session_state["all_offers"][i] = fresh_res['data']
                        st.rerun()
                except Exception as e: st.error(f"خطأ: {str(e)}")
        st.markdown("</div>", unsafe_allow_html=True)
    
    st.markdown("---")
    render_pagination_bottom()
    st.markdown("---")

def rebuild_offer_payload(existing_data, overrides=None):
    """بناء آمن وذكي للطلب (Payload) يعالج شروط سلة الصارمة ويمنع أخطاء 422 و 500"""
    if overrides is None: overrides = {}
    
    def extract_ids(items):
        if not items: return []
        return [item.get('id', item) if isinstance(item, dict) else item for item in items]
        
    buy_data = existing_data.get('buy', {}) or {}
    get_data = existing_data.get('get', {}) or {}

    payload = {
        "name": existing_data.get('name', 'عرض بدون اسم'),
        "offer_type": existing_data.get('offer_type', 'buy_x_get_y'),
        "applied_channel": existing_data.get('applied_channel', 'browser_and_application'),
        "applied_to": existing_data.get('applied_to', 'product'),
        "start_date": existing_data.get('start_date', ''),
        "expiry_date": existing_data.get('expiry_date', ''),
        "status": existing_data.get('status', 'active'),
        "applied_with_coupon": existing_data.get('applied_with_coupon', False),
        "max_discount_amount": float(existing_data.get('max_discount_amount') or 0),
        "min_purchase_amount": float(existing_data.get('min_purchase_amount') or 0),
        "min_items_count": int(existing_data.get('min_items_count') or 0),
        "message": existing_data.get('message', ''),
        "buy": {
            "type": buy_data.get("type", "product"),
            "quantity": int(buy_data.get("quantity", 1))
        },
        "get": {
            "type": get_data.get("type", "product"),
            "quantity": int(get_data.get("quantity", 1)),
            "discount_type": get_data.get("discount_type", "percentage")
        }
    }
    
    if existing_data.get('customer_groups'):
        payload["customer_groups"] = extract_ids(existing_data.get('customer_groups'))

    # ✅ الإصلاح الجذري لمعالجة خطأ 422 (توزيع المنتجات بذكاء)
    applied_to = payload["applied_to"]
    
    # تجميع كافة المعرفات من أي مكان في العرض الأصلي
    all_pids = extract_ids(existing_data.get('products', [])) + extract_ids(buy_data.get('products', [])) + extract_ids(get_data.get('products', []))
    all_pids = list(set(all_pids)) # إزالة التكرار
    
    all_cids = extract_ids(existing_data.get('categories', [])) + extract_ids(buy_data.get('categories', [])) + extract_ids(get_data.get('categories', []))
    all_cids = list(set(all_cids))
    
    all_bids = extract_ids(existing_data.get('brands', [])) + extract_ids(buy_data.get('brands', [])) + extract_ids(get_data.get('brands', []))
    all_bids = list(set(all_bids))

    # توزيع المعرفات إجبارياً على Buy و Get لترضي سلة
    if applied_to == 'product' and all_pids:
        payload['buy']['products'] = all_pids
        payload['get']['products'] = all_pids
    elif applied_to == 'category' and all_cids:
        payload['buy']['categories'] = all_cids
        payload['get']['categories'] = all_cids
    elif applied_to == 'brand' and all_bids:
        payload['buy']['brands'] = all_bids
        payload['get']['brands'] = all_bids

    if 'discount_amount' in get_data: 
        payload['get']['discount_amount'] = float(get_data['discount_amount'])

    for key, value in overrides.items():
        if key in payload: payload[key] = value

    return payload
