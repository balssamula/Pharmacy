import streamlit as st
import pandas as pd
import io
import requests
import json
from datetime import datetime
from typing import Optional, List, Dict, Any
import logging
import traceback
import re

# --- إعدادات التسجيل ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- إعدادات الصفحة ---
st.set_page_config(
    page_title="منظومة بلسم الرقمية لإدارة العروض",
    layout="wide",
    page_icon="🎁",
    initial_sidebar_state="expanded"
)

# ==========================================
# CSS البسيط والمتوافق مع Streamlit
# ==========================================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');
    
    * {
        font-family: 'Cairo', sans-serif !important;
    }
    
    [data-testid="stSidebar"] {
        background-color: #0f1c2e !important;
        padding: 20px 15px !important;
    }
    
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }
    
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3 {
        color: #00b4d8 !important;
    }
    
    .stButton button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }
    
    .stButton button:hover {
        transform: scale(1.02) !important;
        box-shadow: 0 4px 15px rgba(0,0,0,0.15) !important;
    }
    
    .stTextInput input, .stNumberInput input {
        border-radius: 8px !important;
        border: 2px solid #e2e8f0 !important;
    }
    
    .stTextInput input:focus, .stNumberInput input:focus {
        border-color: #00b4d8 !important;
        box-shadow: 0 0 0 3px rgba(0, 180, 216, 0.15) !important;
    }
    
    .stSelectbox select {
        border-radius: 8px !important;
        border: 2px solid #e2e8f0 !important;
    }
    
    .stExpander {
        border-radius: 10px !important;
        border: 1px solid #e8edf2 !important;
    }
    
    .stDataFrame {
        border-radius: 10px !important;
        overflow: hidden !important;
    }
    
    /* تنسيق شريط الحالة */
    .sticky-bar {
        background: linear-gradient(135deg, #0a1628 0%, #1a2d4a 100%);
        padding: 12px 24px;
        border-radius: 10px;
        margin-bottom: 20px;
        border-bottom: 3px solid #00b4d8;
        display: flex;
        justify-content: space-between;
        align-items: center;
        flex-wrap: wrap;
        gap: 10px;
    }
    
    .sticky-bar .title {
        color: #ffffff;
        font-weight: 700;
        font-size: 16px;
    }
    
    .sticky-bar .status {
        color: #00b4d8;
        font-weight: 600;
        font-size: 13px;
        background: rgba(0, 180, 216, 0.12);
        padding: 4px 14px;
        border-radius: 20px;
        border: 1px solid rgba(0, 180, 216, 0.3);
    }
    
    /* تنسيق حالة المنتج في منتصف الحاوية */
    .status-center {
        display: flex;
        justify-content: center;
        align-items: center;
        height: 100%;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# دوال مساعدة
# ==========================================

def safe_parse_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        try:
            return datetime.strptime(date_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            return None

def parse_products_cleanly(product_list: Optional[List]) -> str:
    if not product_list or not isinstance(product_list, list):
        return "كل منتجات المتجر"
    
    clean_elements = []
    for p in product_list:
        try:
            if isinstance(p, dict):
                name = p.get('name', 'منتج مشمول')
                sku = p.get('sku', 'بدون SKU')
                product_id = p.get('id', 'بدون ID')
                clean_elements.append(f"• {name} (SKU: {sku}) [ID: {product_id}]")
            else:
                clean_elements.append(f"• معرف منتج رقم: {p}")
        except Exception:
            clean_elements.append("• منتج غير معرف")
    
    return "\n".join(clean_elements) if clean_elements else "لا توجد منتجات"

def get_product_price(product: Dict) -> float:
    try:
        price = product.get('price', {})
        if isinstance(price, dict):
            return float(price.get('amount', 0))
        return float(price) if price else 0.0
    except (ValueError, TypeError):
        return 0.0

def safe_api_request(method: str, url: str, headers: Dict, **kwargs) -> Optional[Dict]:
    try:
        response = requests.request(method, url, headers=headers, timeout=30, **kwargs)
        
        if response.status_code >= 400:
            error_detail = ""
            try:
                error_data = response.json()
                error_detail = json.dumps(error_data, ensure_ascii=False)
            except:
                error_detail = response.text[:500]
            
            if response.status_code != 404:
                st.error(f"⚠️ خطأ {response.status_code}: {error_detail}")
            logger.error(f"API Error {response.status_code}: {error_detail}")
            return None
        
        return response.json()
        
    except requests.exceptions.RequestException as e:
        st.error(f"⚠️ خطأ في الاتصال: {str(e)}")
        logger.error(f"Request Error: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"⚠️ خطأ في تحليل البيانات: {str(e)}")
        return None

def get_headers():
    token = st.session_state.get('access_token', '')
    if not token:
        st.warning("⚠️ الرجاء إدخال مفتاح الربط (Access Token)")
        return None
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

# ==========================================
# دوال التصدير
# ==========================================

def export_offers_to_excel(offers: List[Dict]) -> bytes:
    try:
        data = []
        for offer in offers:
            buy_products = []
            for p in offer.get('buy', {}).get('products', []):
                if isinstance(p, dict):
                    buy_products.append(str(p.get('id', '')))
                else:
                    buy_products.append(str(p))
            
            get_products = []
            for p in offer.get('get', {}).get('products', []):
                if isinstance(p, dict):
                    get_products.append(str(p.get('id', '')))
                else:
                    get_products.append(str(p))
            
            data.append({
                'المعرف': offer.get('id', ''),
                'اسم العرض': offer.get('name', ''),
                'النوع': offer.get('offer_type', ''),
                'الحالة': 'مفعل' if offer.get('status') == 'active' else 'غير مفعل',
                'مع كوبون': 'نعم' if offer.get('applied_with_coupon', False) else 'لا',
                'تاريخ البدء': offer.get('start_date', ''),
                'تاريخ الانتهاء': offer.get('expiry_date', ''),
                'منتجات الشراء': ', '.join(buy_products),
                'كمية الشراء': offer.get('buy', {}).get('quantity', 1),
                'منتجات الهدية': ', '.join(get_products),
                'كمية الهدية': offer.get('get', {}).get('quantity', 1),
                'الرسالة': offer.get('message', '')
            })
        
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='قائمة العروض')
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        st.error(f"⚠️ خطأ في تصدير العروض: {str(e)}")
        return b""

def export_products_to_excel(products: List[Dict]) -> bytes:
    try:
        data = []
        for p in products:
            price = get_product_price(p)
            sale_price = p.get('sale_price', {})
            if isinstance(sale_price, dict):
                sale_price = sale_price.get('amount', 0)
            elif isinstance(sale_price, (int, float)):
                sale_price = sale_price
            else:
                sale_price = 0
            
            data.append({
                'المعرف': p.get('id', ''),
                'الاسم': p.get('name', ''),
                'SKU': p.get('sku', ''),
                'السعر': price,
                'السعر المخفض': sale_price if sale_price > 0 else '',
                'المخزون': p.get('quantity', 0),
                'المبيعات': p.get('sold_quantity', 0),
                'الحالة': 'معروض' if p.get('status') == 'sale' else 'مخفي',
                'الرابط': p.get('url', '')
            })
        
        df = pd.DataFrame(data)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='قائمة المنتجات')
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        st.error(f"⚠️ خطأ في تصدير المنتجات: {str(e)}")
        return b""

# ==========================================
# دالة معالجة استيراد الإكسيل
# ==========================================

def process_excel_import(df: pd.DataFrame) -> Dict:
    results = {"success": [], "errors": []}
    headers = get_headers()
    
    if not headers:
        results["errors"].append("❌ الرجاء إدخال مفتاح الربط أولاً")
        return results
    
    for idx, row in df.iterrows():
        try:
            action = str(row.get('Action', 'create')).strip().lower()
            offer_id = row.get('Offer_ID')
            
            if offer_id and isinstance(offer_id, float):
                offer_id = int(offer_id) if offer_id.is_integer() else None
            
            offer_name = str(row.get('Offer_Name', 'عرض جديد')).strip()
            offer_type = str(row.get('Offer_Type', 'buy_x_get_y')).strip()
            applied_channel = str(row.get('Applied_Channel', 'browser_and_application')).strip()
            applied_to = str(row.get('Applied_To', 'product')).strip()
            
            offer_status = str(row.get('Offer_Status', 'active')).strip().lower()
            if offer_status not in ['active', 'inactive']:
                offer_status = 'active'
            
            start_date = row.get('Start_Date_Time')
            if pd.isna(start_date):
                start_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(start_date, datetime):
                start_date = start_date.strftime('%Y-%m-%d %H:%M:%S')
            else:
                start_date = str(start_date)
            
            expiry_date = row.get('Expiry_Date_Time')
            if pd.isna(expiry_date):
                expiry_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(expiry_date, datetime):
                expiry_date = expiry_date.strftime('%Y-%m-%d %H:%M:%S')
            else:
                expiry_date = str(expiry_date)
            
            buy_products_str = str(row.get('Buy_Products_IDs', '')).strip()
            buy_products = []
            if buy_products_str and buy_products_str != 'nan':
                for p in re.split(r'[,\s;]+', buy_products_str):
                    p = p.strip()
                    if p and p.isdigit():
                        buy_products.append(int(p))
            
            get_products_str = str(row.get('Get_Products_IDs', '')).strip()
            get_products = []
            if get_products_str and get_products_str != 'nan':
                for p in re.split(r'[,\s;]+', get_products_str):
                    p = p.strip()
                    if p and p.isdigit():
                        get_products.append(int(p))
            
            offer_data = {
                "name": offer_name,
                "offer_type": offer_type,
                "applied_channel": applied_channel,
                "applied_to": applied_to,
                "start_date": start_date,
                "expiry_date": expiry_date,
                "message": str(row.get('Offer_Message', '')).strip(),
                "status": offer_status
            }
            
            with_coupon = str(row.get('With_Coupon', 'لا')).strip()
            offer_data["applied_with_coupon"] = with_coupon == 'نعم'
            
            buy_qty = 1
            try:
                buy_qty_val = row.get('Buy_Quantity')
                if pd.notna(buy_qty_val):
                    buy_qty = int(float(buy_qty_val))
            except:
                pass
            
            buy_type = str(row.get('Buy_Type', 'product')).strip()
            offer_data["buy"] = {
                "type": buy_type,
                "quantity": buy_qty
            }
            if buy_products:
                offer_data["buy"]["products"] = buy_products
            
            get_qty = 1
            try:
                get_qty_val = row.get('Get_Quantity')
                if pd.notna(get_qty_val):
                    get_qty = int(float(get_qty_val))
            except:
                pass
            
            get_type = str(row.get('Get_Type', 'product')).strip()
            discount_type = str(row.get('Discount_Type', 'percentage')).strip()
            
            offer_data["get"] = {
                "type": get_type,
                "quantity": get_qty,
                "discount_type": discount_type
            }
            
            try:
                discount_amount = row.get('Discount_Amount')
                if pd.notna(discount_amount) and float(discount_amount) > 0:
                    offer_data["get"]["discount_amount"] = float(discount_amount)
            except:
                pass
            
            if get_products:
                offer_data["get"]["products"] = get_products
            
            if action == 'create':
                response = safe_api_request(
                    "POST", 
                    "https://api.salla.dev/admin/v2/specialoffers", 
                    headers, 
                    json=offer_data
                )
                if response:
                    status_text = "مفعل" if offer_status == "active" else "غير مفعل"
                    results["success"].append(f"✅ تم إنشاء العرض: {offer_name} (الحالة: {status_text})")
                else:
                    results["errors"].append(f"❌ فشل إنشاء العرض: {offer_name}")
                    
            elif action == 'update' and offer_id:
                response = safe_api_request(
                    "PUT", 
                    f"https://api.salla.dev/admin/v2/specialoffers/{offer_id}", 
                    headers, 
                    json=offer_data
                )
                if response:
                    results["success"].append(f"✅ تم تحديث العرض ID: {offer_id}")
                else:
                    results["errors"].append(f"❌ فشل تحديث العرض ID: {offer_id}")
                    
            elif action == 'delete' and offer_id:
                response = safe_api_request(
                    "DELETE", 
                    f"https://api.salla.dev/admin/v2/specialoffers/{offer_id}", 
                    headers
                )
                if response:
                    results["success"].append(f"✅ تم حذف العرض ID: {offer_id}")
                else:
                    results["errors"].append(f"❌ فشل حذف العرض ID: {offer_id}")
                    
            elif action in ['active', 'inactive'] and offer_id:
                status = "active" if action == 'active' else "inactive"
                response = safe_api_request(
                    "PUT", 
                    f"https://api.salla.dev/admin/v2/specialoffers/{offer_id}/status", 
                    headers, 
                    json={"status": status}
                )
                if response:
                    results["success"].append(f"✅ تم تغيير حالة العرض ID: {offer_id} إلى {status}")
                else:
                    results["errors"].append(f"❌ فشل تغيير حالة العرض ID: {offer_id}")
                    
            else:
                results["errors"].append(f"⚠️ إجراء غير معروف: {action}")
                
        except Exception as e:
            results["errors"].append(f"❌ خطأ في الصف {idx+1}: {str(e)}")
            logger.error(f"Import error: {traceback.format_exc()}")
    
    return results

# ==========================================
# دالة إنشاء نموذج الإكسيل
# ==========================================

def generate_salla_excel_template() -> bytes:
    try:
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl import Workbook
            from openpyxl.styles import numbers
        except ImportError:
            import subprocess
            subprocess.check_call(["pip", "install", "openpyxl"])
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl import Workbook
            from openpyxl.styles import numbers
        
        output = io.BytesIO()
        
        columns = [
            "Action", "Offer_ID", "Offer_Name", "Offer_Type", "Applied_Channel",
            "Applied_To", "With_Coupon", "Offer_Status", "Start_Date_Time", "Expiry_Date_Time", 
            "Buy_Type", "Buy_Quantity", "Buy_Products_IDs", 
            "Get_Type", "Get_Quantity", "Discount_Type", 
            "Discount_Amount", "Get_Products_IDs", "Offer_Message"
        ]
        
        sample_data = [
            ["create", "", "عرض ترويجي جديد", "buy_x_get_y", "browser_and_application",
             "product", "نعم", "active", "2026-06-22 12:00:00", "2026-07-22 23:59:59",
             "product", 1, "1298176905", 
             "product", 1, "percentage", 50, "1298176905", "خصم 50% على الحبة الثانية"],
        ]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "قائمة العروض"
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
        
        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        data_font = Font(name="Segoe UI", size=10)
        data_alignment = Alignment(horizontal="right", vertical="center")
        
        for row in range(3, len(sample_data) + 3):
            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value or '')
                    if len(val) > max_length:
                        max_length = len(val)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # القوائم المنسدلة
        dv_action = DataValidation(
            type="list",
            formula1='"create,update,active,inactive,delete"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="قيمة غير صحيحة",
            error="الرجاء اختيار: create, update, active, inactive, delete"
        )
        ws.add_data_validation(dv_action)
        dv_action.add("A3:A100")
        
        dv_offer_type = DataValidation(
            type="list",
            formula1='"buy_x_get_y,percentage,fixed_amount,discounts_table,tiered_offer,cart_offer,special_price"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="نوع عرض غير صحيح",
            error="الرجاء اختيار نوع العرض المناسب"
        )
        ws.add_data_validation(dv_offer_type)
        dv_offer_type.add("D3:D100")
        
        dv_channel = DataValidation(
            type="list",
            formula1='"browser,browser_and_application"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="قناة غير صحيحة",
            error="الرجاء اختيار: browser أو browser_and_application"
        )
        ws.add_data_validation(dv_channel)
        dv_channel.add("E3:E100")
        
        dv_applied_to = DataValidation(
            type="list",
            formula1='"order,product,category,paymentMethod"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="نوع التطبيق غير صحيح",
            error="الرجاء اختيار: order, product, category, paymentMethod"
        )
        ws.add_data_validation(dv_applied_to)
        dv_applied_to.add("F3:F100")
        
        dv_coupon = DataValidation(
            type="list",
            formula1='"نعم,لا"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="قيمة غير صحيحة",
            error="الرجاء اختيار نعم أو لا"
        )
        ws.add_data_validation(dv_coupon)
        dv_coupon.add("G3:G100")
        
        dv_offer_status = DataValidation(
            type="list",
            formula1='"active,inactive"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="حالة العرض غير صحيحة",
            error="الرجاء اختيار: active (مفعل) أو inactive (غير مفعل)"
        )
        ws.add_data_validation(dv_offer_status)
        dv_offer_status.add("H3:H100")
        
        dv_disc_type = DataValidation(
            type="list",
            formula1='"percentage,free-product"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="نوع خصم غير صحيح",
            error="الرجاء اختيار: percentage أو free-product"
        )
        ws.add_data_validation(dv_disc_type)
        dv_disc_type.add("P3:P100")
        
        # تنسيق التواريخ
        for row in range(3, 100):
            for col in ['I', 'J']:
                cell = ws[f"{col}{row}"]
                cell.number_format = numbers.FORMAT_DATE_DATETIME
        
        # تعليمات
        ws.insert_rows(1)
        ws.merge_cells('A1:S1')
        instructions_cell = ws.cell(row=1, column=1)
        instructions_cell.value = """
📋 تعليمات التعبئة:
- Action: create (إنشاء), update (تحديث), delete (حذف), active (تفعيل), inactive (إيقاف)
- Applied_To: order (طلب), product (منتج), category (تصنيف), paymentMethod (طريقة دفع) - مطلوب!
- With_Coupon: نعم (تطبيق مع كوبون) أو لا (بدون كوبون)
- Offer_Status: active (مفعل) أو inactive (غير مفعل) - حالة العرض
- Offer_ID: مطلوب للتحديث والحذف (استخدم أرقام صحيحة بدون نقاط عشرية)
- التواريخ: استخدم الصيغة YYYY-MM-DD HH:mm:ss
- المنتجات: يمكن إدخال أكثر من معرف بفاصلة مثل: 123,456,789
"""
        instructions_cell.font = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
        instructions_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 140
        
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        st.error(f"⚠️ خطأ في إنشاء النموذج: {str(e)}")
        columns = [
            "Action", "Offer_ID", "Offer_Name", "Offer_Type", "Applied_Channel",
            "Applied_To", "With_Coupon", "Offer_Status", "Start_Date_Time", "Expiry_Date_Time",
            "Buy_Type", "Buy_Quantity", "Buy_Products_IDs",
            "Get_Type", "Get_Quantity", "Discount_Type",
            "Discount_Amount", "Get_Products_IDs", "Offer_Message"
        ]
        df = pd.DataFrame(columns=columns)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='قائمة العروض')
        buffer.seek(0)
        return buffer.getvalue()

# ==========================================
# دالة إنشاء عرض جديد
# ==========================================

def create_new_offer(offer_data: Dict) -> bool:
    headers = get_headers()
    if not headers:
        return False
    response = safe_api_request(
        "POST",
        "https://api.salla.dev/admin/v2/specialoffers",
        headers,
        json=offer_data
    )
    return response is not None

# ==========================================
# دالة تحديث حالة المنتج
# ==========================================

def update_product_status(product_id: int, status: str) -> bool:
    """تحديث حالة المنتج - إصلاح مشكلة price و sale_price"""
    headers = get_headers()
    if not headers:
        return False
    
    # جلب المنتج الحالي
    current = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current or not current.get('data'):
        st.error("❌ فشل في جلب بيانات المنتج")
        return False
    
    product_data = current['data']
    
    # تحديث الحالة فقط - لا نغير الأسعار
    update_payload = {
        "status": status
    }
    
    # إضافة الأسعار الحالية لتجنب خطأ 422
    if 'price' in product_data:
        price_val = product_data.get('price')
        if isinstance(price_val, dict):
            update_payload['price'] = price_val.get('amount', 0)
        else:
            update_payload['price'] = float(price_val) if price_val else 0
    
    if 'sale_price' in product_data:
        sale_val = product_data.get('sale_price')
        if isinstance(sale_val, dict):
            update_payload['sale_price'] = sale_val.get('amount', 0)
        else:
            update_payload['sale_price'] = float(sale_val) if sale_val else 0
    
    # إرسال التحديث
    response = safe_api_request(
        "PUT",
        f"https://api.salla.dev/admin/v2/products/{product_id}",
        headers,
        json=update_payload
    )
    
    return response is not None

# ==========================================
# دالة تحديث العنوان الترويجي
# ==========================================

def update_product_promotion(product_id: int, promotion_title: str) -> bool:
    """تحديث العنوان الترويجي للمنتج - إصلاح مشكلة price و sale_price"""
    headers = get_headers()
    if not headers:
        return False
    
    # جلب المنتج الحالي
    current = safe_api_request("GET", f"https://api.salla.dev/admin/v2/products/{product_id}", headers)
    if not current or not current.get('data'):
        st.error("❌ فشل في جلب بيانات المنتج")
        return False
    
    product_data = current['data']
    
    # بناء payload التحديث
    update_payload = {
        "promotion_title": promotion_title,
        "status": product_data.get('status', 'sale')
    }
    
    # إضافة الأسعار الحالية لتجنب خطأ 422
    if 'price' in product_data:
        price_val = product_data.get('price')
        if isinstance(price_val, dict):
            update_payload['price'] = price_val.get('amount', 0)
        else:
            update_payload['price'] = float(price_val) if price_val else 0
    
    if 'sale_price' in product_data:
        sale_val = product_data.get('sale_price')
        if isinstance(sale_val, dict):
            update_payload['sale_price'] = sale_val.get('amount', 0)
        else:
            update_payload['sale_price'] = float(sale_val) if sale_val else 0
    
    # إرسال التحديث
    response = safe_api_request(
        "PUT",
        f"https://api.salla.dev/admin/v2/products/{product_id}",
        headers,
        json=update_payload
    )
    
    return response is not None

# ==========================================
# إدارة جلسة الدخول
# ==========================================

def init_session_state():
    defaults = {
        "admin_password": "admin123",
        "logged_in": False,
        "access_token": "",
        "setup_completed": True
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

init_session_state()

# ==========================================
# شاشة الدخول
# ==========================================

if not st.session_state["logged_in"]:
    # استخدام st.columns لتوسيط المحتوى بدلاً من HTML
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("---")
        st.markdown("## 🛡️ منظومة بلسم")
        st.markdown("##### تسجيل الدخول الآمن إلى لوحة التحكم")
        st.markdown("---")
        
        st.text_input("🔑 مفتاح الربط (Access Token):", type="password", key="login_token", help="أدخل التوكن الخاص بتطبيقك")
        st.text_input("👤 اسم المستخدم:", value="admin", key="lg_un")
        st.text_input("🔒 كلمة المرور:", type="password", key="lg_pw")
        
        if st.button("🚀 دخول آمن للمنظومة", use_container_width=True, key="submit_login"):
            token = st.session_state.get("login_token", "").strip()
            if token:
                st.session_state["access_token"] = token
            if st.session_state.get("lg_un") == "admin" and st.session_state.get("lg_pw") == st.session_state["admin_password"]:
                st.session_state["logged_in"] = True
                st.rerun()
            else:
                st.error("❌ بيانات الدخول خاطئة.")
    st.stop()

# ==========================================
# إعدادات API
# ==========================================

SALLA_API_URL = "https://api.salla.dev/admin/v2/specialoffers"

# ==========================================
# الشريط العلوي
# ==========================================

st.markdown("---")
st.markdown("### 🛡️ لوحة التحكم الإدارية لصيدليات بلسم العُلا")
st.markdown("✅ الاتصال موثق ومستقر")
st.markdown("---")

# أزرار التحكم العلوية
col1, col2, col3 = st.columns([1.5, 1.5, 5])

with col1:
    with st.popover("🔑 تعديل مفتاح الربط"):
        current_token = st.session_state.get("access_token", "")
        new_tok = st.text_input("أدخل التوكن الجديد:", value=current_token, type="password")
        if st.button("تحديث التوكن", use_container_width=True):
            if new_tok.strip():
                st.session_state["access_token"] = new_tok.strip()
                st.success("✅ تم تحديث التوكن بنجاح!")
                st.rerun()
            else:
                st.warning("⚠️ الرجاء إدخال توكن صحيح")

with col2:
    with st.popover("🔒 تعديل كلمة المرور"):
        new_pwd = st.text_input("أدخل كلمة المرور الجديدة:", type="password")
        if st.button("تحديث الباسورد", use_container_width=True):
            if new_pwd.strip():
                st.session_state["admin_password"] = new_pwd.strip()
                st.success("✅ تم تحديث كلمة المرور بنجاح!")
            else:
                st.warning("⚠️ الرجاء إدخال كلمة مرور صحيحة")

st.divider()

# ==========================================
# القائمة الجانبية
# ==========================================

st.sidebar.markdown("---")
st.sidebar.markdown("### 🏥 بوابة بلسم الرقمية")
st.sidebar.markdown("منظومة إدارة العروض والمنتجات")
st.sidebar.markdown("---")

page = st.sidebar.radio(
    "📋 تصفح الأقسام التنفيذية:",
    [
        "📊 لوحة تصفية وإدارة العروض الحالية",
        "📦 مركز جرد المنتجات ومعرفات الـ IDs"
    ],
    index=0
)

st.sidebar.markdown("---")

if st.sidebar.button("🔄 تحديث البيانات والصفحة", use_container_width=True, key="refresh_page_btn"):
    st.rerun()

if st.sidebar.button("🚪 تسجيل الخروج", use_container_width=True, key="logout_sidebar"):
    st.session_state["logged_in"] = False
    st.session_state["access_token"] = ""
    st.rerun()

with st.sidebar.expander("ℹ️ معلومات النظام", expanded=False):
    st.caption(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    st.caption("🔗 https://api.salla.dev/admin/v2")
    st.caption("📊 الحالة: متصل")

# ==========================================
# الشاشة الأولى: لوحة العروض
# ==========================================

if page == "📊 لوحة تصفية وإدارة العروض الحالية":
    st.markdown("""
        ## 📊 لوحة إدارة العروض الاحترافية
        إدارة شاملة للعروض مع إمكانية التصفية والبحث والتعديل الفوري
    """)
    
    # --- نموذج الاستيراد ---
    col_info, col_btn = st.columns([3, 1])
    with col_info:
        st.info("📥 قم بتنزيل النموذج الاحترافي وتعبئة البيانات بالصيغ المحددة")
    with col_btn:
        st.download_button(
            label="📥 تحميل النموذج",
            data=generate_salla_excel_template(),
            file_name="Salla_Offers_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    uploaded_file = st.file_uploader(
        "📂 اختر ملف العروض بصيغة XLSX للاستيراد الجماعي:",
        type=["xlsx"],
        help="قم بتحميل ملف الإكسيل المعبأ وفق النموذج"
    )
    
    if uploaded_file:
        try:
            df_user = pd.read_excel(uploaded_file)
            
            date_columns = ['Start_Date_Time', 'Expiry_Date_Time']
            for col in date_columns:
                if col in df_user.columns:
                    df_user[col] = df_user[col].apply(
                        lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, (datetime, pd.Timestamp)) else str(x) if pd.notna(x) else ''
                    )
            
            numeric_columns = ['Offer_ID', 'Buy_Quantity', 'Get_Quantity', 'Discount_Amount']
            for col in numeric_columns:
                if col in df_user.columns:
                    df_user[col] = pd.to_numeric(df_user[col], errors='coerce')
            
            st.success(f"✅ تم تحميل الملف! يحتوي على {len(df_user)} عرض")
            st.dataframe(df_user, use_container_width=True, height=300)
            
            if st.button("🚀 تأكيد النشر الجماعي", use_container_width=True, type="primary"):
                with st.spinner("🔄 جاري معالجة العروض..."):
                    results = process_excel_import(df_user)
                
                if results["success"]:
                    for msg in results["success"]:
                        st.success(msg)
                if results["errors"]:
                    for msg in results["errors"]:
                        st.error(msg)
                
                if results["success"]:
                    st.balloons()
                    st.rerun()
                    
        except Exception as e:
            st.error(f"⚠️ خطأ في قراءة الملف: {str(e)}")
            st.code(traceback.format_exc())
    
    st.divider()
    
    # --- زر إنشاء عرض جديد ---
    with st.expander("➕ إنشاء عرض جديد", expanded=False):
        st.markdown("### إنشاء عرض جديد")
        
        col1, col2 = st.columns(2)
        with col1:
            new_offer_name = st.text_input("اسم العرض:", placeholder="أدخل اسم العرض")
            new_offer_type = st.selectbox(
                "نوع العرض:",
                ["buy_x_get_y", "percentage", "fixed_amount", "discounts_table", "tiered_offer", "cart_offer", "special_price"],
                key="new_offer_type"
            )
            new_applied_to = st.selectbox(
                "تطبيق العرض على:",
                ["product", "order", "category", "paymentMethod"],
                key="new_applied_to"
            )
            new_with_coupon = st.selectbox(
                "تطبيق مع كوبون:",
                ["نعم", "لا"],
                key="new_with_coupon"
            )
        
        with col2:
            new_offer_status = st.selectbox(
                "حالة العرض:",
                ["active", "inactive"],
                key="new_offer_status",
                format_func=lambda x: "مفعل" if x == "active" else "غير مفعل"
            )
            new_start_date = st.text_input(
                "تاريخ البدء:",
                value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                help="صيغة: YYYY-MM-DD HH:mm:ss"
            )
            new_expiry_date = st.text_input(
                "تاريخ الانتهاء:",
                value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                help="صيغة: YYYY-MM-DD HH:mm:ss"
            )
            new_message = st.text_input("الرسالة الترويجية:", placeholder="رسالة العرض")
        
        st.markdown("---")
        st.markdown("#### 🛒 تفاصيل العرض")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            new_buy_type = st.selectbox(
                "نوع الشراء:",
                ["product", "category"],
                key="new_buy_type"
            )
            new_buy_qty = st.number_input(
                "كمية الشراء المطلوبة:",
                min_value=1,
                value=1,
                key="new_buy_qty"
            )
        
        with col2:
            new_get_type = st.selectbox(
                "نوع الهدية:",
                ["product", "category"],
                key="new_get_type"
            )
            new_get_qty = st.number_input(
                "كمية الهدية:",
                min_value=1,
                value=1,
                key="new_get_qty"
            )
        
        with col3:
            new_discount_type = st.selectbox(
                "نوع الخصم:",
                ["percentage", "free-product"],
                key="new_discount_type"
            )
            new_discount_amount = st.number_input(
                "قيمة الخصم:",
                min_value=0.0,
                value=0.0,
                step=0.5,
                key="new_discount_amount",
                help="نسبة مئوية (مثال: 50) أو 0 للهدية المجانية"
            )
        
        new_buy_products = st.text_input(
            "معرفات منتجات الشراء (بفاصلة):",
            placeholder="مثال: 1298176905, 1444615766",
            help="أدخل معرفات المنتجات مفصولة بفاصلة"
        )
        new_get_products = st.text_input(
            "معرفات منتجات الهدية (بفاصلة):",
            placeholder="مثال: 1298176905, 1444615766",
            help="أدخل معرفات المنتجات مفصولة بفاصلة"
        )
        
        if st.button("✅ إنشاء العرض", use_container_width=True, type="primary"):
            try:
                buy_products = []
                if new_buy_products.strip():
                    for p in re.split(r'[,\s;]+', new_buy_products.strip()):
                        p = p.strip()
                        if p and p.isdigit():
                            buy_products.append(int(p))
                
                get_products = []
                if new_get_products.strip():
                    for p in re.split(r'[,\s;]+', new_get_products.strip()):
                        p = p.strip()
                        if p and p.isdigit():
                            get_products.append(int(p))
                
                offer_data = {
                    "name": new_offer_name,
                    "offer_type": new_offer_type,
                    "applied_channel": "browser_and_application",
                    "applied_to": new_applied_to,
                    "start_date": new_start_date,
                    "expiry_date": new_expiry_date,
                    "message": new_message,
                    "status": new_offer_status,
                    "applied_with_coupon": new_with_coupon == "نعم",
                    "buy": {
                        "type": new_buy_type,
                        "quantity": int(new_buy_qty)
                    },
                    "get": {
                        "type": new_get_type,
                        "quantity": int(new_get_qty),
                        "discount_type": new_discount_type
                    }
                }
                
                if buy_products:
                    offer_data["buy"]["products"] = buy_products
                if get_products:
                    offer_data["get"]["products"] = get_products
                if new_discount_amount > 0 and new_discount_type == "percentage":
                    offer_data["get"]["discount_amount"] = float(new_discount_amount)
                
                with st.spinner("🔄 جاري إنشاء العرض..."):
                    if create_new_offer(offer_data):
                        st.success("✅ تم إنشاء العرض بنجاح!")
                        st.rerun()
                    else:
                        st.error("❌ فشل إنشاء العرض. يرجى التحقق من البيانات.")
                        
            except Exception as e:
                st.error(f"❌ خطأ: {str(e)}")
    
    st.divider()
    
    # --- جلب العروض ---
    with st.spinner("🔄 جاري تحميل العروض..."):
        headers = get_headers()
        if headers:
            res = safe_api_request("GET", SALLA_API_URL, headers)
        else:
            res = None
    
    if res and res.get("data"):
        raw_offers = res["data"]
        
        # --- أزرار التصدير ---
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("📥 تصدير العروض", use_container_width=True):
                excel_data = export_offers_to_excel(raw_offers)
                if excel_data:
                    st.download_button(
                        label="📥 تحميل",
                        data=excel_data,
                        file_name=f"offers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_offers"
                    )
        
        # --- الفلترة ---
        with st.expander("🔍 خيارات البحث والفلترة المتقدمة", expanded=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                search_offer = st.text_input(
                    "🔎 بحث باسم العرض أو المعرف:",
                    placeholder="أدخل نص البحث..."
                )
            with col2:
                offer_status_filter = st.selectbox(
                    "📌 حالة العرض:",
                    ["الكل", "نشط", "غير نشط", "منتهي الصلاحية", "لم يبدأ بعد"]
                )
            with col3:
                offer_type_filter = st.selectbox(
                    "🏷️ نوع العرض:",
                    ["الكل", "buy_x_get_y", "percentage", "fixed_amount", "discounts_table", "tiered_offer", "cart_offer", "special_price"]
                )
        
        now = datetime.now()
        filtered_offers = []
        
        for offer in raw_offers:
            match = True
            start_date = safe_parse_date(offer.get('start_date'))
            expiry_date = safe_parse_date(offer.get('expiry_date'))
            
            if search_offer:
                search_lower = search_offer.lower()
                offer_name = offer.get('name', '').lower()
                offer_id = str(offer.get('id', ''))
                
                buy_products = offer.get('buy', {}).get('products', [])
                get_products = offer.get('get', {}).get('products', [])
                all_ids = []
                for p in buy_products + get_products:
                    if isinstance(p, dict):
                        all_ids.append(str(p.get('id', '')))
                        all_ids.append(str(p.get('sku', '')).lower())
                    else:
                        all_ids.append(str(p))
                
                if search_lower not in offer_name and search_lower not in offer_id:
                    if not any(search_lower in pid for pid in all_ids):
                        match = False
            
            if offer_type_filter != "الكل" and offer.get('offer_type') != offer_type_filter:
                match = False
            
            if offer_status_filter != "الكل":
                current_status = offer.get('status', '')
                if offer_status_filter == "نشط" and current_status != "active":
                    match = False
                elif offer_status_filter == "غير نشط" and current_status != "inactive":
                    match = False
                elif offer_status_filter == "منتهي الصلاحية" and (not expiry_date or expiry_date >= now):
                    match = False
                elif offer_status_filter == "لم يبدأ بعد" and (not start_date or start_date <= now):
                    match = False
            
            if match:
                filtered_offers.append(offer)
        
        st.markdown(f"""
            <div style="background: #f0f4f8; padding: 8px 16px; border-radius: 8px; margin-bottom: 14px; border-right: 4px solid #00b4d8;">
                <strong>📊 عدد العروض: {len(filtered_offers)} عرض</strong>
            </div>
        """, unsafe_allow_html=True)
        
        # عرض العروض
        for idx, offer in enumerate(filtered_offers):
            offer_id = offer.get('id', 'N/A')
            offer_status = offer.get('status', 'inactive')
            applied_with_coupon = offer.get('applied_with_coupon', False)
            
            # تنسيق شارات الحالة
            status_text = "🟢 مفعل" if offer_status == "active" else "🔴 غير مفعل"
            coupon_text = "🔖 مع كوبون" if applied_with_coupon else "🔖 بدون كوبون"
            
            with st.container():
                st.markdown(f"""
                    <div style="background: #ffffff; padding: 14px 18px; border-radius: 12px; 
                                border-right: 5px solid #2a9d8f; border-left: 1px solid #e8edf2; 
                                border-top: 1px solid #e8edf2; border-bottom: 1px solid #e8edf2; 
                                margin-bottom: 14px; box-shadow: 0 2px 8px rgba(0,0,0,0.04);">
                """, unsafe_allow_html=True)
                
                col1, col2, col3, col4, col5 = st.columns([2.5, 1.2, 1.2, 1.2, 1.2])
                
                with col1:
                    st.markdown(f"""
                        <div>
                            <h4 style="margin: 0 0 4px 0; color: #0f1c2e; font-size: 16px;">🎯 {offer.get('name', 'عرض بدون اسم')}</h4>
                            <span style="color: #6c757d; font-size: 12px;">🆔 ID: {offer_id}</span>
                            <br>
                            <span style="color: #6c757d; font-size: 12px;">📅 {offer.get('start_date', 'غير محدد')} → {offer.get('expiry_date', 'غير محدد')}</span>
                            <br>
                            <span style="display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; 
                                  background: {'#d4edda' if offer_status == 'active' else '#f8d7da'}; 
                                  color: {'#155724' if offer_status == 'active' else '#721c24'};">
                                {status_text}
                            </span>
                            <span style="display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; 
                                  background: {'#d4edda' if applied_with_coupon else '#f8d7da'}; 
                                  color: {'#155724' if applied_with_coupon else '#721c24'}; margin-right: 5px;">
                                {coupon_text}
                            </span>
                        </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"**🏷️ النوع:** {offer.get('offer_type', 'غير محدد')}")
                    st.caption(f"📌 {offer.get('applied_to', 'غير محدد')}")
                
                with col3:
                    if st.button("🔄 تبديل الكوبون", key=f"toggle_coupon_{offer_id}_{idx}", use_container_width=True):
                        with st.spinner("🔄 جاري تحديث..."):
                            # استخدام PATCH بدلاً من PUT لتجنب مشاكل الحقول المطلوبة
                            update_res = safe_api_request(
                                "PUT",
                                f"{SALLA_API_URL}/{offer_id}",
                                get_headers(),
                                json={"applied_with_coupon": not applied_with_coupon}
                            )
                            if update_res:
                                st.success(f"✅ تم تحديث حالة الكوبون!")
                                st.rerun()
                
                with col4:
                    target_status = "inactive" if offer_status == "active" else "active"
                    btn_label = "⏸️ إيقاف" if offer_status == "active" else "▶️ تفعيل"
                    if st.button(btn_label, key=f"toggle_status_{offer_id}_{idx}", use_container_width=True):
                        with st.spinner("🔄 جاري تحديث الحالة..."):
                            update_res = safe_api_request(
                                "PUT",
                                f"{SALLA_API_URL}/{offer_id}/status",
                                get_headers(),
                                json={"status": target_status}
                            )
                            if update_res:
                                st.success("✅ تم تحديث الحالة!")
                                st.rerun()
                
                with col5:
                    if st.button("🗑️ حذف", key=f"delete_offer_{offer_id}_{idx}", use_container_width=True, type="primary"):
                        with st.spinner("🔄 جاري الحذف..."):
                            del_res = safe_api_request("DELETE", f"{SALLA_API_URL}/{offer_id}", get_headers())
                            if del_res is not None:
                                st.success("✅ تم حذف العرض!")
                                st.rerun()
                
                # تفاصيل العرض
                with st.expander("🔽 تفاصيل العرض المتقدمة", expanded=False):
                    st.markdown("---")
                    col_left, col_right = st.columns(2)
                    
                    with col_left:
                        st.markdown("**🛒 منتجات الشراء (X):**")
                        buy_products = offer.get('buy', {}).get('products', [])
                        st.text(parse_products_cleanly(buy_products))
                        st.markdown(f"**📦 كمية الشراء:** `{offer.get('buy', {}).get('quantity', 1)}`")
                    
                    with col_right:
                        st.markdown("**🎁 منتجات الهدية (Y):**")
                        get_products = offer.get('get', {}).get('products', [])
                        st.text(parse_products_cleanly(get_products))
                        st.markdown(f"**🎯 كمية الهدية:** `{offer.get('get', {}).get('quantity', 1)}`")
                    
                    st.markdown("---")
                    st.markdown("#### ✏️ تعديل تفاصيل العرض")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        ed_name = st.text_input("اسم العرض:", value=offer.get('name', ''), key=f"edit_name_{offer_id}")
                        ed_msg = st.text_input("الرسالة الترويجية:", value=offer.get('message', ''), key=f"edit_msg_{offer_id}")
                    
                    with col2:
                        ed_type = st.selectbox(
                            "نوع العرض:",
                            ["buy_x_get_y", "percentage", "fixed_amount", "discounts_table", "tiered_offer", "cart_offer", "special_price"],
                            index=["buy_x_get_y", "percentage", "fixed_amount", "discounts_table", "tiered_offer", "cart_offer", "special_price"].index(offer.get('offer_type', 'buy_x_get_y')) if offer.get('offer_type', 'buy_x_get_y') in ["buy_x_get_y", "percentage", "fixed_amount", "discounts_table", "tiered_offer", "cart_offer", "special_price"] else 0,
                            key=f"edit_type_{offer_id}"
                        )
                        
                        ed_applied_to = st.selectbox(
                            "تطبيق العرض على:",
                            ["order", "product", "category", "paymentMethod"],
                            index=["order", "product", "category", "paymentMethod"].index(offer.get('applied_to', 'product')) if offer.get('applied_to', 'product') in ["order", "product", "category", "paymentMethod"] else 1,
                            key=f"edit_applied_to_{offer_id}"
                        )
                        
                        ed_coupon = st.selectbox(
                            "تطبيق مع كوبون:",
                            ["نعم", "لا"],
                            index=0 if offer.get('applied_with_coupon', False) else 1,
                            key=f"edit_coupon_{offer_id}"
                        )
                        
                        ed_offer_status = st.selectbox(
                            "حالة العرض:",
                            ["active", "inactive"],
                            index=0 if offer.get('status') == 'active' else 1,
                            key=f"edit_offer_status_{offer_id}"
                        )
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        ed_start = st.text_input(
                            "تاريخ البدء:",
                            value=offer.get('start_date', ''),
                            key=f"edit_start_{offer_id}",
                            help="صيغة: YYYY-MM-DD HH:mm:ss"
                        )
                    with col2:
                        ed_end = st.text_input(
                            "تاريخ الانتهاء:",
                            value=offer.get('expiry_date', ''),
                            key=f"edit_end_{offer_id}",
                            help="صيغة: YYYY-MM-DD HH:mm:ss"
                        )
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        ed_buy_q = st.number_input(
                            "كمية الشراء المطلوبة:",
                            value=int(offer.get('buy', {}).get('quantity', 1)),
                            min_value=1,
                            key=f"edit_buy_q_{offer_id}"
                        )
                    with col2:
                        ed_get_q = st.number_input(
                            "كمية الهدية:",
                            value=int(offer.get('get', {}).get('quantity', 1)),
                            min_value=1,
                            key=f"edit_get_q_{offer_id}"
                        )
                    
                    if st.button("💾 حفظ التحديثات", key=f"save_offer_{offer_id}", use_container_width=True, type="primary"):
                        try:
                            update_payload = {
                                "name": ed_name,
                                "message": ed_msg,
                                "start_date": ed_start,
                                "expiry_date": ed_end,
                                "offer_type": ed_type,
                                "applied_to": ed_applied_to,
                                "applied_with_coupon": ed_coupon == "نعم",
                                "status": ed_offer_status,
                                "buy": {
                                    "type": offer.get('buy', {}).get('type', 'product'),
                                    "quantity": int(ed_buy_q)
                                },
                                "get": {
                                    "type": offer.get('get', {}).get('type', 'product'),
                                    "quantity": int(ed_get_q),
                                    "discount_type": offer.get('get', {}).get('discount_type', 'free-product')
                                }
                            }
                            
                            buy_products_ids = []
                            for p in offer.get('buy', {}).get('products', []):
                                if isinstance(p, dict):
                                    buy_products_ids.append(p.get('id'))
                                else:
                                    buy_products_ids.append(p)
                            if buy_products_ids:
                                update_payload["buy"]["products"] = buy_products_ids
                            
                            get_products_ids = []
                            for p in offer.get('get', {}).get('products', []):
                                if isinstance(p, dict):
                                    get_products_ids.append(p.get('id'))
                                else:
                                    get_products_ids.append(p)
                            if get_products_ids:
                                update_payload["get"]["products"] = get_products_ids
                            
                            with st.spinner("🔄 جاري حفظ التغييرات..."):
                                update_res = safe_api_request(
                                    "PUT",
                                    f"{SALLA_API_URL}/{offer_id}",
                                    get_headers(),
                                    json=update_payload
                                )
                                if update_res:
                                    st.success("✅ تم تحديث العرض!")
                                    st.rerun()
                        except Exception as e:
                            st.error(f"❌ خطأ: {str(e)}")
                
                st.markdown("</div>", unsafe_allow_html=True)
    
    else:
        if headers:
            st.warning("⚠️ لا توجد عروض حالياً في المتجر")
        else:
            st.warning("⚠️ الرجاء إدخال مفتاح الربط أولاً")

# ==========================================
# الشاشة الثانية: مركز جرد المنتجات
# ==========================================

elif page == "📦 مركز جرد المنتجات ومعرفات الـ IDs":
    st.markdown("""
        ## 📦 مركز جرد المنتجات
        إدارة المنتجات وحالة الظهور وعرض العروض المرتبطة
    """)
    
    # --- زر تصدير المنتجات ---
    if st.button("📥 تصدير المنتجات", use_container_width=False):
        with st.spinner("🔄 جاري تحميل المنتجات للتصدير..."):
            headers = get_headers()
            if headers:
                products_res = safe_api_request("GET", "https://api.salla.dev/admin/v2/products", headers)
                if products_res and products_res.get("data"):
                    excel_data = export_products_to_excel(products_res["data"])
                    if excel_data:
                        st.download_button(
                            label="📥 تحميل",
                            data=excel_data,
                            file_name=f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_products"
                        )
    
    with st.spinner("🔄 جاري تحميل المنتجات والعروض..."):
        headers = get_headers()
        if headers:
            products_res = safe_api_request("GET", "https://api.salla.dev/admin/v2/products", headers)
            offers_res = safe_api_request("GET", SALLA_API_URL, headers)
        else:
            products_res = None
            offers_res = None
    
    if products_res and products_res.get("data") and offers_res:
        products = products_res["data"]
        offers = offers_res.get("data", [])
        
        st.info(f"📊 عدد المنتجات: {len(products)} منتج")
        
        # بناء خريطة العروض
        offer_map = {}
        for offer in offers:
            buy_products = offer.get('buy', {}).get('products', [])
            get_products = offer.get('get', {}).get('products', [])
            for p in buy_products + get_products:
                if isinstance(p, dict):
                    product_id = p.get('id')
                    if product_id:
                        offer_map[product_id] = offer
                else:
                    offer_map[p] = offer
        
        search_query = st.text_input(
            "🔍 ابحث هنا بواسطة اسم المنتج أو الـ SKU:",
            placeholder="أدخل اسم المنتج أو SKU..."
        )
        
        filtered_products = []
        if search_query:
            search_lower = search_query.lower()
            for p in products:
                if (search_lower in p.get('name', '').lower() or 
                    search_lower in str(p.get('sku', '')).lower() or
                    search_lower in str(p.get('id', ''))):
                    filtered_products.append(p)
        else:
            filtered_products = products[:20]
            if len(products) > 20:
                st.info("📌 عرض أول 20 منتج. استخدم البحث لعرض المزيد.")
        
        if not filtered_products:
            st.warning("⚠️ لم يتم العثور على منتجات تطابق البحث")
        else:
            for idx, p in enumerate(filtered_products):
                p_id = p.get('id', 'N/A')
                p_name = p.get('name', 'منتج بدون اسم')
                p_sku = p.get('sku', 'لا يوجد')
                
                # العنوان الترويجي
                p_promotion = p.get('promotion_title', '')
                if not p_promotion:
                    promotion = p.get('promotion', {})
                    if isinstance(promotion, dict):
                        p_promotion = promotion.get('title', '')
                
                # حساب الأسعار
                price = get_product_price(p)
                sale_price_obj = p.get('sale_price', {})
                if isinstance(sale_price_obj, dict):
                    sale_price = sale_price_obj.get('amount', 0)
                else:
                    sale_price = float(sale_price_obj) if sale_price_obj else 0
                
                has_discount = sale_price > 0 and sale_price < price
                discount_percent = int(((price - sale_price) / price) * 100) if price > 0 and has_discount else 0
                
                offer = offer_map.get(p_id)
                has_offer = offer is not None
                
                # عرض المنتج باستخدام st.columns بدلاً من HTML
                with st.container():
                    col_title1, col_title2 = st.columns([3, 1])
                    with col_title1:
                        st.markdown(f"### 📦 {p_name}")
                    with col_title2:
                        # حالة المنتج في منتصف الحاوية
                        status = p.get('status', 'sale')
                        status_text = "🟢 معروض" if status == "sale" else "🔴 مخفي"
                        st.markdown(f"<div style='text-align: center; margin-top: 8px;'>{status_text}</div>", unsafe_allow_html=True)
                    
                    st.markdown(f"🏷️ **العنوان الترويجي:** {p_promotion if p_promotion else 'لا يوجد'}")
                    st.markdown(f"🏷️ **SKU:** `{p_sku}` | 🆔 **ID:** `{p_id}`")
                    
                    col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
                    
                    with col1:
                        product_url = p.get('url', '#')
                        st.markdown(f"🔗 [رابط المنتج]({product_url})")
                        if p.get('thumbnail') or p.get('main_image'):
                            st.success("✅ يحتوي على صورة")
                        else:
                            st.warning("⚠️ يحتاج لصورة")
                    
                    with col2:
                        st.markdown("**💰 السعر:**")
                        if has_discount:
                            st.markdown(f"""
                                **السعر المخفض:** {sale_price:,.2f} SAR
                                ~~{price:,.2f} SAR~~
                                **-{discount_percent}%**
                            """)
                        else:
                            st.markdown(f"**{price:,.2f} SAR**")
                        
                        st.markdown(f"**📦 المخزون:** {p.get('quantity', 0)} حبة")
                        st.markdown(f"**📈 المبيعات:** {p.get('sold_quantity', 0)}")
                    
                    with col3:
                        if has_offer:
                            offer_status = offer.get('status', '')
                            offer_id = offer.get('id', '')
                            status_color = "🟢" if offer_status == "active" else "🔴"
                            status_text = "مفعل" if offer_status == "active" else "غير مفعل"
                            coupon_status = "مع كوبون" if offer.get('applied_with_coupon', False) else "بدون كوبون"
                            
                            st.markdown(f"""
                                **🎯 عرض:** {offer.get('name', 'عرض')}
                                🆔 {offer_id}
                                {status_color} {status_text}
                                🔖 {coupon_status}
                            """)
                            
                            col_a, col_b = st.columns(2)
                            with col_a:
                                if offer_status == "active":
                                    if st.button("⏸️ إيقاف", key=f"pause_offer_{p_id}_{idx}", use_container_width=True):
                                        with st.spinner("🔄 جاري إيقاف العرض..."):
                                            update_res = safe_api_request(
                                                "PUT",
                                                f"{SALLA_API_URL}/{offer_id}/status",
                                                get_headers(),
                                                json={"status": "inactive"}
                                            )
                                            if update_res:
                                                st.success("✅ تم إيقاف العرض!")
                                                st.rerun()
                                else:
                                    if st.button("▶️ تفعيل", key=f"activate_offer_{p_id}_{idx}", use_container_width=True):
                                        with st.spinner("🔄 جاري تفعيل العرض..."):
                                            update_res = safe_api_request(
                                                "PUT",
                                                f"{SALLA_API_URL}/{offer_id}/status",
                                                get_headers(),
                                                json={"status": "active"}
                                            )
                                            if update_res:
                                                st.success("✅ تم تفعيل العرض!")
                                                st.rerun()
                        else:
                            st.markdown("**⚪ لا يوجد عرض**")
                            st.button("إضافة عرض", key=f"add_offer_{p_id}_{idx}", disabled=True, use_container_width=True)
                    
                    with col4:
                        if st.button("📋 نسخ ID", key=f"copy_id_{p_id}_{idx}", use_container_width=True):
                            st.toast(f"✅ تم نسخ المعرف: {p_id}")
                        
                        # زر تغيير حالة الظهور
                        current_status = p.get('status', 'sale')
                        btn_label = "👁️ إخفاء" if current_status == "sale" else "👁️ إظهار"
                        btn_type = "primary" if current_status == "sale" else "secondary"
                        
                        if st.button(btn_label, key=f"toggle_status_{p_id}_{idx}", use_container_width=True, type=btn_type):
                            target_status = "hidden" if current_status == "sale" else "sale"
                            with st.spinner("🔄 جاري تحديث الحالة..."):
                                if update_product_status(p_id, target_status):
                                    st.success("✅ تم تحديث حالة المنتج!")
                                    st.rerun()
                                else:
                                    st.error("❌ فشل تحديث حالة المنتج")
                        
                        # زر تعديل العنوان الترويجي
                        if st.button("✏️ تعديل العنوان", key=f"edit_promotion_{p_id}_{idx}", use_container_width=True):
                            st.session_state[f"edit_promotion_{p_id}"] = True
                        
                        if st.session_state.get(f"edit_promotion_{p_id}", False):
                            new_promotion = st.text_input(
                                "العنوان الترويجي الجديد:",
                                value=p_promotion,
                                key=f"promotion_input_{p_id}"
                            )
                            col_save1, col_save2 = st.columns(2)
                            with col_save1:
                                if st.button("💾 حفظ", key=f"save_promotion_{p_id}", use_container_width=True):
                                    if update_product_promotion(p_id, new_promotion):
                                        st.success("✅ تم تحديث العنوان الترويجي!")
                                        st.session_state[f"edit_promotion_{p_id}"] = False
                                        st.rerun()
                                    else:
                                        st.error("❌ فشل تحديث العنوان الترويجي")
                            with col_save2:
                                if st.button("❌ إلغاء", key=f"cancel_promotion_{p_id}", use_container_width=True):
                                    st.session_state[f"edit_promotion_{p_id}"] = False
                                    st.rerun()
                    
                    st.divider()
    
    else:
        if headers:
            st.warning("⚠️ لا توجد منتجات أو فشل في تحميل البيانات.")
        else:
            st.warning("⚠️ الرجاء إدخال مفتاح الربط أولاً")

# ==========================================
# التذييل
# ==========================================

st.markdown("---")
st.markdown("""
    <div style="text-align: center; padding: 16px; color: #6c757d; font-size: 13px;">
        <p>© 2026 منظومة بلسم الرقمية | جميع الحقوق محفوظة</p>
        <p style="font-size: 11px; color: #adb5bd;">تم التطوير باستخدام Streamlit</p>
    </div>
""", unsafe_allow_html=True)
