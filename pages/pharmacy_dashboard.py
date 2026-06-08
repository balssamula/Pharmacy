import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.database import (
    fetch_active_items, get_completed_items, get_tab_completed_counts, 
    get_old_orders, get_old_invoices, get_old_invoices_stats,
    check_duplicate_across_branches, get_all_duplicate_items,
    save_case_note, mark_case_done
)
from utils.helpers import (
    is_cancelled_or_returned_status, is_pending_payment_status, 
    get_branch_number, get_branch_location, get_tab_label, numeric_value,
    get_saudi_time
)
from utils.ui_components import render_metrics, render_completed_table

def to_safe_int(val):
    """تحويل آمن ومطلق لأي قيمة نصية أو فارغة إلى عدد صحيح لمنع الانهيار الحسابي"""
    if pd.isna(val) or str(val).strip() in ["", "nan", "None"]:
        return 0
    try:
        return int(float(str(val).strip()))
    except:
        return 0
        
def export_to_excel(dataframes_dict: dict, pharmacy_name: str) -> bytes:
    """تصدير البيانات إلى ملف Excel مع تنسيق احترافي"""
    output = BytesIO()
    
    tab_colors = {
        "الإضافات والطلبات المفقودة": "4472C4",
        "الإرجاعات والزيادات": "ED7D31",
        "فواتير بعد اخر طلب": "9B59B6",
        "بانتظار الدفع": "3498DB",
        "الملغيات والمسترجعات": "E74C3C",
        "الفواتير القديمة (أرشيف)": "6c757d",
        "الطلبات القديمة (أرشيف)": "6c757d"
    }
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
                worksheet = writer.sheets[sheet_name[:31]]
                
                header_fill = PatternFill(start_color=tab_colors.get(sheet_name, "2A5298"), 
                                         end_color=tab_colors.get(sheet_name, "2A5298"), 
                                         fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for row in range(1, len(df) + 2):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
                
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                empty_df = pd.DataFrame({"ملاحظة": ["لا توجد بيانات في هذا التبويب"]})
                empty_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    
    output.seek(0)
    return output.getvalue()

def render_single_case_card(row, idx, allow_actions, pharmacist_name, pharmacy_name):
    """عرض بطاقة حالة ديناميكية بالكامل تنقاد خلف إشارة الفرق الفعلي"""
    salla_numeric = int(row.get('salla_qty', 0)) if pd.notna(row.get('salla_qty', 0)) else 0
    abc_numeric = int(row.get('abc_qty', 0)) if pd.notna(row.get('abc_qty', 0)) else 0
    diff_value = salla_numeric - abc_numeric
    
    case_type = row.get('case_type', '')

    if diff_value > 0:
        badge_text = "إضافة مخزنية عادية ➕"
        badge_color = "#dff1ff"
        badge_text_color = "#084298"
        diff_style = "color: #28a745; font-weight: bold;"
        required_action = "<span style='color: #28a745; font-weight: bold;'>إضافة</span>"
    elif diff_value < 0:
        badge_text = "إرجاع مخزني عادي 🔄"
        badge_color = "#ffe0df"
        badge_text_color = "#491217"
        diff_style = "color: #dc3545; font-weight: bold;"
        required_action = "<span style='color: #dc3545; font-weight: bold;'>إرجاع</span>"
    else:
        diff_style = "color: #6c757d; font-weight: bold;"
        required_action = "<span style='color: #6c757d; font-weight: bold;'>مطابق</span>"
        
        if case_type == 'orphan_salla':
            badge_text = "طلب مبيعات مفقود الفاتورة 🛒"
            badge_color = "#fff3cd"
            badge_text_color = "#856404"
        elif case_type == 'orphan_abc':
            badge_text = "فاتورة توريد مفقودة الطلب 📄"
            badge_color = "#f8d7da"
            badge_text_color = "#721c24"
        else:
            badge_text = "حالة تسوية عامة"
            badge_color = "#e2e8f0"
            badge_text_color = "#475569"
   
    order_status = row.get('order_status', 'غير متوفرة')
    invoice_date = row.get('invoice_date', '')
    order_date = row.get('order_date', '')

    order_number = str(row.get('order_number', ''))
    sku = str(row.get('sku', ''))
    
    duplicates = []
    try:
        duplicates = check_duplicate_across_branches(order_number, sku, pharmacy_name)
    except:
        pass
    
    with st.container():
        st.markdown(f"""
        <div style="background:#f8f9fa; border-radius:16px; padding:1rem; margin-bottom:1rem; border-right:5px solid #1f7a8c; box-shadow:0 2px 8px rgba(0,0,0,0.05);">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:0.75rem; flex-wrap:wrap;">
                <span style="background:{badge_color}; color:{badge_text_color}; padding:0.25rem 0.75rem; border-radius:20px; font-size:0.8rem; font-weight:bold;">{badge_text}</span>
                <div>
                    <span style="color:#6c757d; font-size:0.75rem;">📅 الطلب: {order_date[:16] if order_date else 'غير محدد'}</span>
                    <span style="color:#6c757d; font-size:0.75rem; margin-right:0.5rem;">📅 الفاتورة: {invoice_date[:16] if invoice_date else 'غير محدد'}</span>
                </div>
            </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([2, 1, 1.5])
        with col1:
            st.markdown(f"""
            - **📋 رقم الطلب:** {row.get('order_number', 'N/A')}
            - **🏷️ SKU:** {row.get('sku', 'N/A')}
            - **📦 المنتج:** {str(row.get('product_name', 'N/A'))[:60]}
            """)
        with col2:
            st.markdown(f"""
            - **🛒 كمية سلة:** {salla_numeric}
            - **📄 كمية ABC:** {abc_numeric}
            - **📊 الفرق:** <span style="{diff_style}">{'+' if diff_value > 0 else ''}{diff_value}</span>
            - **🎯 المطلوب:** {required_action}
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            - **🧾 رقم الفاتورة:** {row.get('invoice_number', 'N/A')}
            - **👤 الصيدلي:** {row.get('abc_pharmacist_name', 'غير معروف')}
            - **⚙️ حالة الطلب:** {order_status}
            """)
            
        if duplicates:
            dup_warning_html = (
                '<div style="background:#fff3cd; border-right:4px solid #ff9800; padding:0.75rem; margin-top:0.75rem; border-radius:10px; margin-bottom:0.75rem; direction:rtl; text-align:right;">'
                '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">'
                '<span style="font-size:1.2rem;">⚠️</span>'
                '<span style="color:#856404; font-weight:bold;">تنبيه هام: هذا الصنف مكرر بموجب نفس رقم الطلب في فروع أخرى!</span>'
                '</div>'
                '<div style="margin-right:1.5rem;">'
            )
            for dup in duplicates:
                dup_warning_html += (
                    '<div style="font-size:0.85rem; margin-bottom:0.3rem; color:#66521a;">'
                    f'🏥 <strong>{dup.get("pharmacy", "غير معروف")}</strong> | الإجراء الحالي: {dup.get("status", "غير معروف")} | تصنيف الحالة: {dup.get("case_type", "غير معروف")}'
                    '</div>'
                )
            dup_warning_html += '</div></div>'
            st.markdown(dup_warning_html, unsafe_allow_html=True)
            
        if row.get('is_duplicate_warning') == 1 or "تنبيه للمراجعة والتدقيق" in str(row.get('case_reason', '')):
            st.markdown(f"""
            <div style="background-color: #fff5f5; border: 1px solid #fc8181; padding: 0.75rem; border-radius: 10px; margin-top: 0.5rem; margin-bottom: 0.5rem;">
                <p style="color: #c53030; margin: 0; font-size: 0.85rem; font-weight: bold; line-height: 1.4;">
                    {row.get('case_reason', '')}
                </p>
            </div>
            """, unsafe_allow_html=True)
            
        st.markdown("</div>", unsafe_allow_html=True)
        
        note_key = f"note_{idx}_{row.get('order_number', '')}_{row.get('sku', '')}"
        note_value = st.text_area("📝 :ملحوظة الصيدلي", value=row.get("pharmacist_note", "") or "", key=note_key, height=60)
        
        btn_col1, btn_col2, btn_col3 = st.columns([1, 1.5, 1.5])
        with btn_col1:
            if st.button("💾 حفظ الملحوظة", key=f"save_{idx}_{note_key}", use_container_width=True):
                save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, note_value)
                st.toast("📋 تم حفظ الملاحظة بنجاح!", icon="💾")
                
        if allow_actions and row.get("status") != "تم":
            if case_type == "branch_conflict":
                with btn_col2:
                    if st.button("📥 تأكيد الإضافة (فرعي الصحيح)", key=f"conf_add_{idx}_{note_key}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, f"[فرع صحيح - تم الإضافة للمخزن] | {note_value}")
                        mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                        st.toast("✅ تم اعتماد الفرع كقيد صحيح وإغلاق التسوية!", icon="📥")
                        st.rerun()
                with btn_col3:
                    if st.button("🔄 تأكيد الإرجاع (فرعي الخطأ)", key=f"conf_ret_{idx}_{note_key}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, f"[فرع مخطئ - جاري عكس الفاتورة على ABC] | {note_value}")
                        mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                        st.toast("🔄 تم تسجيل الخطأ وإصدار قيد الإرجاع المعاكس بنجاح!", icon="🔄")
                        st.rerun()
            elif case_type in {"addition", "orphan_salla", "return", "orphan_abc"}:
                button_label = "✅ تأكيد الإضافة" if case_type in {"addition", "orphan_salla"} else "🔄 تأكيد الإرجاع"
                with btn_col2:
                    if st.button(button_label, key=f"done_{idx}_{note_key}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, note_value)
                        mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                        st.toast("🚀 تم تأكيد وتحديث الحالة بنجاح!", icon="✅")
                        st.rerun()
                    
        st.markdown("---")

def render_old_orders_pharmacy(old_orders_df, pharmacy_name, pharmacist_name):
    if old_orders_df.empty: return
    for idx, row in old_orders_df.iterrows():
        days_old = int(row['days_old'])
        badge = "🔴 قديم جداً" if days_old > 365 else "🟠 قديم" if days_old > 180 else "🟡 يحتاج مراجعة"
        salla_numeric = int(row.get('salla_qty', 0))
        abc_numeric = int(row.get('abc_qty', 0))
        diff_value = salla_numeric - abc_numeric
        diff_style = "color: #28a745; font-weight: bold;" if diff_value > 0 else "color: #dc3545; font-weight: bold;" if diff_value < 0 else "color: #6c757d; font-weight: bold;"
        
        with st.container():
            st.markdown(f"""
            <div style="background:#fff3cd30; border-radius:16px; padding:1rem; margin-bottom:1rem; border-right:5px solid #dc3545; box-shadow:0 2px 8px rgba(0,0,0,0.05);">
                <div style="display:flex; justify-content:space-between; margin-bottom:0.5rem;">
                    <span style="background:#dc3545; color:white; padding:0.2rem 0.8rem; border-radius:20px; font-size:0.8rem; font-weight:bold;">{badge}</span>
                    <span style="color:#6c757d; font-size:0.8rem;">📅 {row['order_date'][:16] if row['order_date'] else ''} | ⏰ {days_old} يوم</span>
                </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                - **📋 رقم الطلب:** {row['order_number']}
                - **🏷️ SKU:** {row['sku']}
                - **📦 المنتج:** {row['product_name'][:60]}
                """)
            with col2:
                st.markdown(f"""
                - **🛒 كمية سلة:** {salla_numeric}
                - **📄 كمية ABC:** {abc_numeric}
                - **📊 الفرق:** <span style="{diff_style}">{diff_value}</span>
                """, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("---")

def render_old_invoices_pharmacy(old_invoices_df, pharmacy_name, pharmacist_name):
    if old_invoices_df.empty: return
    for idx, row in old_invoices_df.iterrows():
        days_old = int(row['days_old'])
        badge = "🔴 قديم جداً" if days_old > 365 else "🟠 قديم" if days_old > 180 else "🟡 يحتاج مراجعة"
        salla_numeric = int(row.get('salla_qty', 0))
        abc_numeric = int(row.get('abc_qty', 0))
        diff_value = salla_numeric - abc_numeric
        diff_style = "color: #28a745; font-weight: bold;" if diff_value > 0 else "color: #dc3545; font-weight: bold;" if diff_value < 0 else "color: #6c757d; font-weight: bold;"
        
        with st.container():
            st.markdown(f"""
            <div style="background:#fff3cd30; border-radius:16px; padding:1rem; margin-bottom:1rem; border-right:5px solid #dc3545; box-shadow:0 2px 8px rgba(0,0,0,0.05);">
                <div style="display:flex; justify-content:space-between; margin-bottom:0.5rem;">
                    <span style="background:#dc3545; color:white; padding:0.2rem 0.8rem; border-radius:20px; font-size:0.8rem; font-weight:bold;">{badge}</span>
                    <span style="color:#6c757d; font-size:0.8rem;">📅 الفاتورة: {row['invoice_date'][:16] if row['invoice_date'] else ''} | ⏰ {days_old} يوم</span>
                </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                - **📋 رقم الفاتورة:** {row['invoice_number']}
                - **🏷️ SKU:** {row['sku']}
                """)
            with col2:
                st.markdown(f"""
                - **🛒 كمية سلة:** {salla_numeric}
                - **📄 كمية ABC:** {abc_numeric}
                - **📊 الفرق:** <span style="{diff_style}">{diff_value}</span>
                """, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("---")

def render_case_cards_pharmacy(df: pd.DataFrame, allow_actions: bool, pharmacist_name: str, pharmacy_name: str):
    if df.empty:
        st.success("🎉 لا توجد حالات في هذا القسم.")
        return
    for idx, row in df.iterrows():
        render_single_case_card(row, idx, allow_actions, pharmacist_name, pharmacy_name)

def show():
    pharmacy_name = st.session_state.username
    pharmacist_name = st.session_state.pharmacist_name or ""
    branch_number = get_branch_number(pharmacy_name)
    branch_location = get_branch_location(branch_number)

    st.markdown(f"""
    <div class="hero">
        <h1>🏥 {pharmacy_name}</h1>
        <p>فرع رقم {branch_number} | الموقع: {branch_location} | الصيدلي: {pharmacist_name}</p>
        <p>🕐 آخر تحديث: {get_saudi_time()}</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("🔄 تحديث الصفحة", use_container_width=True):
            st.rerun()
    with col2:
        if st.button("📥 تصدير التقارير إلى Excel", use_container_width=True):
            st.session_state.show_export_pharmacy = True

    # واجهة العرض السحابي المباشر داخل فرع الصيدلية
    st.markdown('### 🌐 مراقبة مبيعات الفرع الحية (السحاب المشترك)')
    with st.spinner("🔄 جاري مزامنة فواتير الفرع من السحاب..."):
        from utils.api_connectors import fetch_abc_invoices_live
        df_abc_cloud = fetch_abc_invoices_live()
        
    if not df_abc_cloud.empty:
        # تصفية الفواتير لتظهر فقط الفواتير التابعة لهذا الفرع تبعا للاسم المسجل به
        df_branch_cloud = df_abc_cloud[df_abc_cloud['رقم الصيدلية'].astype(str).str.contains(str(pharmacy_name), na=False)]
        if not df_branch_cloud.empty:
            st.success(f"✅ تم مزامنة {len(df_branch_cloud):,} فاتورة خاصة بفرعكم اليوم حياً من النظام المركزي.")
            st.dataframe(df_branch_cloud.head(500), use_container_width=True)
        else:
            st.info("📭 الاتصال مستقر بالسحاب، ولكن لا توجد فواتير مرفوعة تخص فرعكم الحالي اليوم بعد.")
            
    df = fetch_active_items(pharmacy_name, include_hidden=False)
    
    # 💡 [موقع حرج]: جلب الفواتير والطلبات القديمة وتجهيز فلاتر التبويبات في أعلى الدالة لمنع الـ UnboundLocalError
    old_invoices_df = get_old_invoices(pharmacy_name=pharmacy_name, months=6)
    old_orders_df = get_old_orders(pharmacy_name=pharmacy_name, months=6)

    if df.empty:
        st.info("📭 لا توجد حالات نشطة لهذا الفرع حاليًا.")
        return

    is_locked = df['is_item_locked'].iloc[0] == 1 if 'is_item_locked' in df.columns else False
    allow_actions = not is_locked

    active_mask = ~df["order_status"].apply(is_cancelled_or_returned_status)
    active_df = df[active_mask].copy()
    
    if not old_invoices_df.empty and 'item_key' in old_invoices_df.columns:
        active_df = active_df[~active_df['item_key'].isin(old_invoices_df['item_key'])]
    if not old_orders_df.empty and 'item_key' in old_orders_df.columns:
        active_df = active_df[~active_df['item_key'].isin(old_orders_df['item_key'])]

    # 🧠 التصفية الكاملة والمنقاة لجداول التبويبات الحالية والنشطة
    branch_add_df = df[df['case_type'].isin(['addition', 'orphan_salla']) & active_mask].copy()
    if not old_orders_df.empty and 'item_key' in old_orders_df.columns:
        branch_add_df = branch_add_df[~branch_add_df['item_key'].isin(old_orders_df['item_key'])].copy()
        
    total_additions_merged = len(branch_add_df)
    completed_additions_merged = len(branch_add_df[branch_add_df["status"] == "تم"])
    pending_additions_merged = total_additions_merged - completed_additions_merged
    
    branch_ret_df = df[df['case_type'].isin(['return', 'orphan_abc']) & active_mask].copy()
    if not old_invoices_df.empty and 'item_key' in old_invoices_df.columns:
        branch_ret_df = branch_ret_df[~branch_ret_df['item_key'].isin(old_invoices_df['item_key'])].copy()
        
    total_returns_merged = len(branch_ret_df)
    completed_returns_merged = len(branch_ret_df[branch_ret_df["status"] == "تم"])
    pending_returns_merged = total_returns_merged - completed_returns_merged
    
    conflicts_df = df[df["case_type"] == "branch_conflict"].copy()
    total_conflicts = len(conflicts_df)
    completed_conflicts = len(conflicts_df[conflicts_df["status"] == "تم"])

    post_cutoff_df = df[(df["case_type"] == "post_cutoff_abc") & active_mask].copy()
    total_post_cutoff = len(post_cutoff_df)
    completed_post_cutoff = len(post_cutoff_df[post_cutoff_df["status"] == "تم"])
    
    payment_df = df[df["order_status"].apply(is_pending_payment_status) & (df["case_type"] != "branch_conflict") & active_mask].copy()
    total_payment = len(payment_df)
    
    cancelled_df = df[df["order_status"].apply(is_cancelled_or_returned_status) & (df["case_type"] != "branch_conflict")].copy()
    total_cancelled = len(cancelled_df)
    
    completed_df = get_completed_items(pharmacy_name)
    total_completed = len(completed_df)

    # 📊 رسم العدادات الرقمية العلوية النشطة
    # 📊 رسم العدادات الرقمية العلوية النشطة (تم تصحيح اسم متغير الحالات المكتملة)
    total = len(active_df)
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    with col1: st.metric("📊 إجمالي الحالات", total)
    with col2: st.metric("➕ إضافات", len(active_df[active_df["case_type"] == "addition"]))
    with col3: st.metric("➖ إرجاعات", len(active_df[active_df["case_type"] == "return"]))
    with col4: st.metric("📦 طلبات بلا فاتورة", len(active_df[active_df["case_type"] == "orphan_salla"]))
    with col5: st.metric("🧾 فواتير بلا طلب", len(active_df[active_df["case_type"] == "orphan_abc"]))
    with col6: st.metric("⏰ فواتير مخرجات", total_post_cutoff)
    with col7: st.metric("✅ تم إنجازها", total_completed) # 💡 هنا تم الإصلاح

    # 📥 تنفيذ تصدير ملف الـ Excel للفرع بدون أي مشاكل في التعريف المتأخر
    if st.session_state.get('show_export_pharmacy', False):
        excel_sheets = {
            "الاضافات والطلبات المفقودة": branch_add_df,
            "الارجاعات والزيادات": branch_ret_df,
            "فواتير معلقة بين الفروع": conflicts_df,
            "فواتير بعد اخر طلب": post_cutoff_df,
            "بانتظار الدفع": payment_df,
            "الملغيات والمسترجعات": cancelled_df,
            "الفواتير القديمة (أرشيف)": old_invoices_df,
            "الالطلبات القديمة (أرشيف)": old_orders_df
        }
        excel_data = export_to_excel(excel_sheets, pharmacy_name)
        st.download_button(
            label="💾 اضغط هنا لتحميل ملف Excel الموحد للفرع",
            data=excel_data,
            file_name=f"Report_{pharmacy_name.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.session_state.show_export_pharmacy = False

    # تفعيل شاشات الألسنة وبناء التبويبات المرئية
    tab_additions, tab_returns, tab_conflicts, tab_post_cutoff, tab_payment, tab_cancelled, tab_completed, tab_old_orders, tab_old_invoices = st.tabs([
        f"📥 الإضافات والطلبات المفقودة ({completed_additions_merged}/{total_additions_merged})" if total_additions_merged > 0 else "📥 الإضافات والطلبات المفقودة (0)",
        f"📤 الإرجاعات والفواتير المعلقة ({completed_returns_merged}/{total_returns_merged})" if total_returns_merged > 0 else "📤 الإرجاعات والفواتير المعلقة (0)",
        f"📊 فواتير معلقة بين الفروع ({completed_conflicts}/{total_conflicts})" if total_conflicts > 0 else f"📊 فواتير معلقة بين الفروع ({total_conflicts})", 
        f"⏰ فواتير بعد آخر طلب ({completed_post_cutoff}/{total_post_cutoff})" if total_post_cutoff > 0 else f"⏰ فواتير بعد آخر طلب ({total_post_cutoff})",
        f"💰 بانتظار الدفع ({total_payment})", f"⚠️ ملغي/مسترجع ({total_cancelled})", f"✅ تم الانتهاء ({total_completed})",
        f"📅 طلبات قديمة ({len(old_orders_df)})", f"🧾 فواتير قديمة ({len(old_invoices_df)})"
    ])
    
    with tab_additions: render_case_cards_pharmacy(branch_add_df, allow_actions, pharmacist_name, pharmacy_name)
    with tab_returns: render_case_cards_pharmacy(branch_ret_df, allow_actions, pharmacist_name, pharmacy_name)
    with tab_conflicts:
        if not conflicts_df.empty:
            for idx, row in conflicts_df.iterrows(): render_single_case_card(row, idx, allow_actions=True, pharmacist_name=pharmacist_name, pharmacy_name=pharmacy_name)
        else: st.success("🎉 ممتاز! لا توجد فواتير معلقة أو متداخلة مع فروع أخرى لفرعكم الحالي.")
    with tab_post_cutoff: render_case_cards_pharmacy(post_cutoff_df, False, pharmacist_name, pharmacy_name)
    with tab_payment: render_case_cards_pharmacy(payment_df, False, pharmacist_name, pharmacy_name)
    with tab_cancelled: render_case_cards_pharmacy(cancelled_df, False, pharmacist_name, pharmacy_name)
    with tab_completed: render_completed_table(completed_df, is_admin=False)
    with tab_old_orders: render_old_orders_pharmacy(old_orders_df, pharmacy_name, pharmacist_name)
    with tab_old_invoices: render_old_invoices_pharmacy(old_invoices_df, pharmacy_name, pharmacist_name)
