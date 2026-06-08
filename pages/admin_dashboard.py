import streamlit as st
import pandas as pd
# 💡 [إصلاح مشكلة حجم الجدول]: زيادة الحد الأقصى للخلايا التي يمكن تلوينها في الواجهة
pd.set_option("styler.render.max_elements", 5000000)
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from utils.database import (
    get_latest_upload_summary, get_all_sessions, get_session_items, 
    lock_session, unlock_session, activate_session, delete_session,
    fetch_active_items, get_all_last_logins, get_completed_items,
    reopen_case_by_item_key, hide_item_from_pharmacy, unhide_item_from_pharmacy,
    lock_item, unlock_item, save_case_note,
    get_manager_last_login, get_login_history,
    get_old_orders, get_old_orders_stats,
    get_old_invoices, get_old_invoices_stats,
    move_item_to_branch, get_available_branches,
    check_duplicate_across_branches
)
from utils.helpers import (
    is_cancelled_or_returned_status, is_pending_payment_status,
    get_tab_label, numeric_value
)
from utils.excel_processor import process_excel
from utils.api_connectors import fetch_abc_invoices_live

def export_to_excel(dataframes_dict: dict) -> bytes:
    output = BytesIO()
    tab_colors = {
        "الإضافات": "4472C4", "الإرجاعات": "ED7D31",
        "طلبات_بدون_فاتورة": "70AD47", "فواتير_بدون_طلب": "FFC000",
        "فواتير_بعد_آخر_طلب": "9B59B6", "بانتظار_الدفع": "3498DB",
        "ملغي_ومسترجع": "E74C3C", "تم_الانتهاء": "27AE60",
        "مقارنة_الجلسات": "2A5298"
    }
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                worksheet = writer.sheets[sheet_name[:31]]
                header_fill = PatternFill(start_color=tab_colors.get(sheet_name, "2A5298"), 
                                         end_color=tab_colors.get(sheet_name, "2A5298"), fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for row in range(1, len(df) + 2):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if worksheet.cell(row=1, column=col).value == "الفرق":
                            diff_value = cell.value
                            if diff_value:
                                if diff_value > 0:
                                    cell.font = Font(color="008000", bold=True)
                                elif diff_value < 0:
                                    cell.font = Font(color="FF0000", bold=True)
            else:
                empty_df = pd.DataFrame({"ملاحظة": ["لا توجد بيانات"]})
                empty_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    output.seek(0)
    return output.getvalue()

def compare_sessions(session1_id: str, session2_id: str) -> pd.DataFrame:
    import sqlite3
    from utils.database import DB_PATH
    conn = sqlite3.connect(DB_PATH)
    query = """
        SELECT order_number, invoice_number, sku, product_name, pharmacy_name,
               salla_qty, abc_qty, (salla_qty - abc_qty) as difference,
               case_type, order_status, abc_pharmacist_name
        FROM reconciliation_items 
        WHERE upload_batch_id IN (?, ?) AND active = 1
    """
    df = pd.read_sql_query(query, conn, params=(session1_id, session2_id))
    conn.close()
    df['required_action'] = df['difference'].apply(
        lambda x: 'إضافة' if x > 0 else ('إرجاع' if x < 0 else 'مطابق')
    )
    return df.rename(columns={
        "order_number": "رقم الطلب", "invoice_number": "رقم الفاتورة",
        "sku": "SKU", "product_name": "المنتج", "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة", "abc_qty": "كمية ABC",
        "difference": "الفرق", "order_status": "حالة الطلب",
        "abc_pharmacist_name": "الصيدلي"
    })

def styled_dataframe(input_df):
    """تنسيق DataFrame مع تمييز الألوان حسب نوع الحالة"""
    if input_df.empty:
        return None
        
    # 💡 [الإصلاح الجذري]: إعادة ضبط الفهارس لتجنب انهيار Streamlit عند التلوين
    display_df = input_df.copy().reset_index(drop=True)
    
    # 💡 [حماية إضافية]: إزالة أي أعمدة متكررة قد توقف عملية العرض
    display_df = display_df.loc[:, ~display_df.columns.duplicated()]

    def highlight_rows(row):
        # تحويل القيم إلى سلاسل نصية للتعامل الآمن
        case_type = str(row.get('case_type', '')).strip() if 'case_type' in row else ''
        status = str(row.get('status', '')).strip() if 'status' in row else ''
        
        # 1. الحالات المكتملة (أخضر فاتح)
        if status == "تم":
            return ['background-color: #d4edda; color: #155724;'] * len(row)
            
        # 2. تمييز الحالات داخل التبويبات المدمجة
        if case_type in ['orphan_salla', 'orphan_abc']:
            return ['background-color: #fff3cd; color: #856404;'] * len(row)
        elif case_type == 'return':
            return ['background-color: #ffe0df; color: #491217;'] * len(row)
        elif case_type == 'addition':
            return ['background-color: #dff1ff; color: #084298;'] * len(row)
            
        return [''] * len(row)
    
    # توحيد روابط المسميات لمنع سقوط الإحصائيات في الداشبورد الإداري
    if 'case_type' in display_df.columns:
        display_df['label_type'] = display_df['case_type'].map({
            'addition': 'إضافة عادية ➕',
            'return': 'إرجاع عادي 🔄',
            'orphan_salla': 'طلب بدون فاتورة (سلة) 🛒',
            'orphan_abc': 'فاتورة بدون طلب (ABC) 📄',
            'post_cutoff_abc': 'فاتورة بعد آخر طلب ⏰'
        }).fillna(display_df['case_type'])
    
    display_df = display_df.rename(columns={
        "order_number": "رقم الطلب",
        "invoice_number": "رقم الفاتورة",
        "sku": "SKU",
        "product_name": "المنتج",
        "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة",
        "abc_qty": "كمية ABC",
        "difference": "الفرق",
        "order_status": "حالة الطلب",
        "type_label": "نوع الحالة",
        "status": "الحالة"
    })
    
    # التأكد مرة أخرى من عدم وجود أعمدة تحمل نفس الاسم بعد الترجمة
    display_df = display_df.loc[:, ~display_df.columns.duplicated()]
    
    return display_df.style.apply(highlight_rows, axis=1)

def render_table_with_click(df, tab_name, allow_move: bool = True):
    """عرض جدول مع إمكانية تحديد الصف وإظهار إجراءات منبثقة مع حماية الذاكرة التامة"""
    if df.empty:
        st.success("لا توجد بيانات في هذا القسم.")
        return
    
    # 💡 [الحل الجذري والنهائي لمنع انهيار السيرفر والذاكرة]:
    # نحدد سقفاً أقصى للعرض المرئي داخل المتصفح، مع بقاء زر التصدير محتفظاً بالبيانات كاملة
    MAX_ROWS_TO_RENDER = 500
    if len(df) > MAX_ROWS_TO_RENDER:
        st.warning(f"⚠️ يحتوي هذا القسم على {len(df)} صف. لحماية أداء التطبيق ومنع الانهيار المجمّع، تم عرض أول {MAX_ROWS_TO_RENDER} صف فقط بالجدول التفاعلي. يمكنك تصدير الملف بصيغة Excel لقراءة التقرير الكامل بكل سلاسة.")
        display_subset_df = df.head(MAX_ROWS_TO_RENDER).copy()
    else:
        display_subset_df = df.copy()
        
    styled_df = styled_dataframe(display_subset_df)
    if styled_df is not None:
        event = st.dataframe(
            styled_df,
            use_container_width=True,
            height=400,
            selection_mode="single-row",
            on_select="rerun"
        )
        
        if event.selection.rows:
            selected_idx = event.selection.rows[0]
            if 0 <= selected_idx < len(display_subset_df):
                row = display_subset_df.iloc[selected_idx]
                item_key = row.get('item_key', row.get('id', ''))
                if pd.isna(item_key) or item_key == '':
                    item_key = f"old_row_{selected_idx}"
                               
# ========== التحقق من وجود مكررات (النسخة المحمية من المسافات البادئة) ==========
                order_number = str(row.get('order_number', ''))
                sku = str(row.get('sku', ''))
                current_pharmacy = row.get('pharmacy_name', '')
                
                duplicate_warning = ""
                try:
                    from utils.database import check_duplicate_across_branches
                    duplicates = check_duplicate_across_branches(order_number, sku, current_pharmacy)
                    
                    if duplicates:
                        duplicate_warning = (
                            '<div style="background:#fff3cd; border-right:4px solid #ff9800; padding:0.75rem; margin-top:0.75rem; border-radius:10px; margin-bottom:0.75rem; direction:rtl; text-align:right;">'
                            '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">'
                            '<span style="font-size:1.2rem;">⚠️</span>'
                            f'<span style="color:#856404; font-weight:bold;">تنبيه: يوجد نفس المنتج (SKU: {sku}) في فروع أخرى بموجب نفس رقم الطلب!</span>'
                            '</div>'
                            '<div style="margin-right:1.5rem;">'
                        )
                        for dup in duplicates:
                            dup_pharmacy = dup.get('pharmacy', 'غير معروف')
                            dup_status = dup.get('status', 'غير معروف')
                            dup_case = dup.get('case_type', 'غير معروف')
                            dup_invoice = dup.get('invoice_date', '')
                            invoice_str = f' | تاريخ الفاتورة: {dup_invoice[:16]}' if dup_invoice else ''
                            
                            duplicate_warning += (
                                '<div style="font-size:0.85rem; margin-bottom:0.4rem; padding:0.3rem 0; border-bottom:1px dashed #ffe0a3; color:#66521a;">'
                                f'🏥 <strong>{dup_pharmacy}</strong> | الحالة: {dup_status} | النوع: {dup_case}{invoice_str}'
                                '</div>'
                            )
                        duplicate_warning += '</div></div>'
                except Exception as e:
                    pass
                
                # عرض التنبيه بشكل سليم وبدون مسافات بادئة تكسر مفسر الماركداون
                if duplicate_warning:
                    st.markdown(duplicate_warning, unsafe_allow_html=True)
                
                # صندوق الإجراءات الإدارية الذي يليه مباشرة
                st.markdown(f"""
                <div style="background:#f0f2f6;border-radius:10px;padding:1rem;margin-top:1rem;border-right:4px solid #1f7a8c;">
                    <h4 style="margin:0 0 0.5rem 0;">🛠️ إجراءات الصف المحدد (الأرشيف التاريخي)</h4>
                    <p><strong>📋 مستند الحالة:</strong> {row.get('order_number', row.get('invoice_number', 'N/A'))} | 
                    <strong>🏷️ SKU:</strong> {row.get('sku', 'N/A')} | 
                    <strong>📦 المنتج:</strong> {str(row.get('product_name', 'N/A'))[:60]}</p>
                </div>
                """, unsafe_allow_html=True)
                
                # الصف الأول من الأزرار
                col1, col2, col3, col4 = st.columns(4)
                
                # زر الإخفاء/الإظهار
                is_hidden = row.get('hidden_from_pharmacy', 0) == 1
                if col1.button("🙈 إخفاء" if not is_hidden else "👁️ إظهار", key=f"hide_{tab_name}_{selected_idx}", use_container_width=True):
                    if is_hidden:
                        unhide_item_from_pharmacy(item_key)
                    else:
                        hide_item_from_pharmacy(item_key, st.session_state.username)
                    st.rerun()
                
                # زر القفل/الفتح
                is_locked = row.get('is_item_locked', 0) == 1
                if col2.button("🔒 قفل" if not is_locked else "🔓 فتح", key=f"lock_{tab_name}_{selected_idx}", use_container_width=True):
                    if is_locked:
                        unlock_item(item_key)
                    else:
                        lock_item(item_key, st.session_state.username)
                    st.rerun()
                
                # زر إعادة الفتح (للحالات المكتملة فقط)
                if row['status'] == "تم":
                    if col3.button("🔄 إعادة فتح", key=f"reopen_{tab_name}_{selected_idx}", use_container_width=True):
                        reopen_case_by_item_key(item_key)
                        st.rerun()
                
                # الصف الثاني من الأزرار (النقل والملحوظة)
                if allow_move and row['status'] != "تم":
                    st.markdown("---")
                    col_a, col_b, col_c, col_d = st.columns([2, 1, 2, 1])
                    
                    current_branch = row.get('pharmacy_name', '')
                    branches = get_available_branches(current_branch)
                    
                    if branches:
                        selected_branch = col_a.selectbox(
                            "🏥 نقل إلى فرع",
                            branches,
                            key=f"move_branch_{tab_name}_{selected_idx}"
                        )
                        
                        if col_b.button("🚚 نقل", key=f"move_{tab_name}_{selected_idx}", use_container_width=True):
                            if move_item_to_branch(item_key, selected_branch, st.session_state.username):
                                st.success(f"✅ تم نقل العنصر إلى {selected_branch}")
                                st.rerun()
                            else:
                                st.error("❌ فشل نقل العنصر")
                    
                    note = col_c.text_input("📝 ملحوظة", value=row.get('pharmacist_note', ''), key=f"note_{tab_name}_{selected_idx}")
                    
                    if col_d.button("💾 حفظ", key=f"save_note_{tab_name}_{selected_idx}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], row['pharmacy_name'], row['case_type'], note)
                        st.rerun()
                else:
                    # إذا كان النقل غير مسموح، نعرض فقط الملحوظة
                    st.markdown("---")
                    col_a, col_b = st.columns([3, 1])
                    note = col_a.text_input("📝 ملحوظة", value=row.get('pharmacist_note', ''), key=f"note_{tab_name}_{selected_idx}")
                    if col_b.button("💾 حفظ", key=f"save_note_{tab_name}_{selected_idx}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], row['pharmacy_name'], row['case_type'], note)
                        st.rerun()

def render_old_items_table(df, title, is_orders=True):
    if df.empty:
        st.success(f"🎉 لا توجد {title} قديمة (أكثر من 6 أشهر)")
        return
    
    def highlight_old_rows(row):
        if is_cancelled_or_returned_status(row.get('order_status', '')):
            return ['background-color: #ffe5e5; color: #333'] * len(row)
        
        order_date = row.get('order_date', '')
        invoice_date = row.get('invoice_date', '')
        check_date = invoice_date if not is_orders else order_date
        
        if check_date and check_date != '':
            try:
                date_obj = datetime.strptime(str(check_date)[:10], "%Y-%m-%d")
                from datetime import timedelta
                if (datetime.now() - date_obj) > timedelta(days=180):
                    return ['background-color: #1a1a1a; color: white; font-weight: bold'] * len(row)
            except:
                pass
        
        if is_pending_payment_status(row.get('order_status', '')):
            return ['background-color: #fff4d6; color: #333'] * len(row)
        
        return [''] * len(row)
    
    display_df = df.copy()
    display_df = display_df.rename(columns={
        "order_number": "رقم الطلب",
        "invoice_number": "رقم الفاتورة",
        "sku": "SKU",
        "product_name": "المنتج",
        "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة",
        "abc_qty": "كمية ABC",
        "difference": "الفرق",
        "case_label": "نوع الحالة",
        "order_status": "حالة الطلب",
        "order_date": "تاريخ الطلب",
        "invoice_date": "تاريخ الفاتورة",
        "days_old": "عدد الأيام"
    })
    
    styled_df = display_df.style.apply(highlight_old_rows, axis=1)
    st.dataframe(styled_df, use_container_width=True, height=400)
    
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(f"📊 إجمالي {title} القديمة", len(df))
    with col2:
        avg_days = df['days_old'].mean() if 'days_old' in df.columns else 0
        st.metric("📅 متوسط عدد الأيام", f"{avg_days:.0f} يوم")
    with col3:
        max_days = df['days_old'].max() if 'days_old' in df.columns else 0
        st.metric("⏰ أقدم عنصر", f"{max_days:.0f} يوم")

def show():
    st.markdown("""
    <div class="hero">
        <h1>👑 لوحة التحكم الإدارية</h1>
        <p>إدارة الطلبات والفواتير - متابعة الإضافات والإرجاعات - إدارة الجلسات</p>
    </div>
    """, unsafe_allow_html=True)
    
    manager_info = get_manager_last_login()
    login_history = get_login_history(10)
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
        <div class="note-card">
            <strong>👑 آخر دخول للمدير العام:</strong><br>
            📅 {manager_info['last_login']}<br>
            🌐 IP: {manager_info['last_ip']}<br>
            👤 {manager_info['pharmacist_name']}
        </div>
        """, unsafe_allow_html=True)
    with col2:
        if not login_history.empty:
            st.markdown("### 📋 آخر محاولات الدخول")
            st.dataframe(login_history, use_container_width=True)
    
    st.markdown('### 📊 فحص المطابقات الحية')

    # 💡 جلب البيانات تلقائياً من السحاب بدلاً من الرفع اليدوي القديم
    with st.spinner("🔄 جاري سحب وتحديث الفواتير حياً من قاعدة البيانات السحابية..."):
        df_abc_all = fetch_abc_invoices_live()

    if not df_abc_all.empty:
        # 🧠 [فلترة الفرع]: جلب اسم الفرع الحالي (سواء الصيدلي الحالي أو عبر وضع الإشراف للإدارة)
        current_branch = st.session_state.get('supervised_pharmacy') or st.session_state.get('pharmacy_name')
    
        # تصفية الفواتير لتظهر فقط الفواتير الخاصة بالفرع النشط حالياً
        # تأكد أن مسمى العمود هنا يطابق ما يعود من دالة الترجمة (مثلاً: 'رقم الصيدلية')
        df_abc = df_abc_all[df_abc_all['رقم الصيدلية'] == current_branch]
    
        # يكمل كود محرك الفرز القياسي والمطابقة مع طلبات سلة أدناه تلقائياً...
        st.dataframe(df_abc) # تجربة استعراض الجدول للتأكد
    else:
        st.info("📭 لا توجد فواتير مسحوبة من السحاب لهذا النطاق الزمني حتى الآن.")
    
    latest = get_latest_upload_summary()
    if latest:
        batch_id, file_name, uploaded_by, uploaded_at, total_cases, additions, returns, orphan_salla, orphan_abc, post_cutoff, is_locked, session_name = latest
        lock_status = "🔒 مقفلة" if is_locked else "🔓 مفتوحة"
        st.markdown(f"""
        <div class="note-card">
            <strong>📋 الجلسة النشطة:</strong> {session_name or 'غير مسماة'} &nbsp; | &nbsp;
            <strong>الملف:</strong> {file_name} &nbsp; | &nbsp;
            <strong>بواسطة:</strong> {uploaded_by} &nbsp; | &nbsp;
            <strong>التاريخ:</strong> {uploaded_at[:16] if uploaded_at else ''} &nbsp; | &nbsp;
            <strong>الحالة:</strong> {lock_status}
        </div>
        """, unsafe_allow_html=True)
    
    with st.expander("📋 إدارة الجلسات السابقة", expanded=False):
        sessions_df = get_all_sessions()
        if not sessions_df.empty:
            for _, session in sessions_df.iterrows():
                col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 2])
                session_name_val = session.get('session_name', session['upload_batch_id'][:8])
                with col1:
                    st.markdown(f'<div class="session-card"><strong>📅 {session_name_val}</strong><br><small>{session["file_name"][:35]}</small></div>', unsafe_allow_html=True)
                with col2:
                    is_active = session.get('is_active', 0)
                    active_badge = "✅ نشطة" if is_active else "⏸ غير نشطة"
                    st.markdown(f'<div class="session-card"><small>👤 {session["uploaded_by"]}<br>📅 {session["uploaded_at"][:16]}<br>{active_badge}</small></div>', unsafe_allow_html=True)
                with col3:
                    st.markdown(f'<div class="session-card"><small>📊 {int(session.get("total_cases", 0))} حالة<br>➕ {int(session.get("total_additions", 0))}<br>➖ {int(session.get("total_returns", 0))}</small></div>', unsafe_allow_html=True)
                with col4:
                    is_locked = session.get('is_locked', 0)
                    lock_text = "🔒 مقفلة" if is_locked else "🔓 مفتوحة"
                    st.markdown(f'<div class="session-card" style="text-align:center;"><small>{lock_text}</small></div>', unsafe_allow_html=True)
                with col5:
                    btn1, btn2, btn3, btn4 = st.columns(4)
                    with btn1:
                        if not is_locked:
                            if st.button(f"🔒", key=f"lock_{session['upload_batch_id']}"):
                                lock_session(session['upload_batch_id'], st.session_state.username)
                                st.rerun()
                        else:
                            if st.button(f"🔓", key=f"unlock_{session['upload_batch_id']}"):
                                unlock_session(session['upload_batch_id'])
                                st.rerun()
                    with btn2:
                        if not is_active:
                            if st.button(f"⭐", key=f"activate_{session['upload_batch_id']}"):
                                activate_session(session['upload_batch_id'])
                                st.rerun()
                    with btn3:
                        if st.button(f"👁️", key=f"view_{session['upload_batch_id']}"):
                            st.session_state.view_session_id = session['upload_batch_id']
                            st.rerun()
                    with btn4:
                        if st.button(f"🗑️", key=f"delete_{session['upload_batch_id']}"):
                            delete_session(session['upload_batch_id'])
                            st.rerun()
            st.markdown("---")
        else:
            st.info("لا توجد جلسات سابقة")

    active_df = fetch_active_items()
    if not active_df.empty:
        st.markdown("### 📈 التحليل الإحصائي الفوري للجلسة النشطة")
        chart_col1, chart_col2 = st.columns(2)
    
        with chart_col1:
            case_counts = active_df['case_label'].value_counts().reset_index()
            case_counts.columns = ['نوع الحالة', 'العدد']
            fig_pie = px.pie(case_counts, values='العدد', names='نوع الحالة', 
                             hole=0.4, color_discrete_sequence=['#4472C4', '#ED7D31', '#70AD47', '#FFC000'])
            fig_pie.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=280, showlegend=True)
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with chart_col2:
            branch_counts = active_df['pharmacy_name'].value_counts().reset_index()
            branch_counts.columns = ['الفرع', 'عدد الحالات']
            fig_bar = px.bar(branch_counts.head(7), x='عدد الحالات', y='الفرع', orientation='h',
                             color='عدد الحالات', color_continuous_scale='Viridis')
            fig_bar.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=280, coloraxis_showscale=False)
            st.plotly_chart(fig_bar, use_container_width=True)
        
    with st.expander("🔄 مقارنة الجلسات", expanded=False):
        sessions_list = get_all_sessions()
        if not sessions_list.empty:
            session_options = {f"{row['session_name']} ({row['uploaded_at'][:16]})": row['upload_batch_id'] 
                              for _, row in sessions_list.iterrows()}
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                session1 = st.selectbox("اختر الجلسة الأولى", list(session_options.keys()), key="session1")
            with col2:
                session2 = st.selectbox("اختر الجلسة الثانية", list(session_options.keys()), key="session2")
            with col3:
                if st.button("📊 مقارنة", use_container_width=True):
                    with st.spinner("جاري المقارنة..."):
                        comparison_df = compare_sessions(session_options[session1], session_options[session2])
                        st.session_state.comparison_result = comparison_df
                        st.success(f"✅ تمت المقارنة!")
            if st.session_state.get('comparison_result') is not None:
                st.dataframe(st.session_state.comparison_result, use_container_width=True)
                excel_data = export_to_excel({"مقارنة_الجلسات": st.session_state.comparison_result})
                st.download_button("📥 تحميل تقرير المقارنة", data=excel_data, type="primary", 
                    file_name=f"session_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        else:
            st.info("لا توجد جلسات للمقارنة")
    
    if st.session_state.get('view_session_id'):
        with st.expander("📄 عرض الجلسة المحددة", expanded=True):
            session_items = get_session_items(st.session_state.view_session_id)
            if not session_items.empty:
                st.dataframe(session_items, use_container_width=True)
            if st.button("إغلاق العرض", use_container_width=True):
                del st.session_state.view_session_id
                st.rerun()
    
    df = fetch_active_items(include_hidden=True)
    if df.empty:
        st.info("📂 لا توجد بيانات فعالة بعد. ارفع ملف Excel من الأعلى لبدء التحليل.")
        return
    
    active_mask = ~df["order_status"].apply(is_cancelled_or_returned_status)
    active_df = df[active_mask]
    
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    with col1:
        st.metric("📊 إجمالي الحالات", len(active_df))
    with col2:
        st.metric("➕ إضافات", len(active_df[active_df["case_type"] == "addition"]))
    with col3:
        st.metric("➖ إرجاعات", len(active_df[active_df["case_type"] == "return"]))
    with col4:
        st.metric("📦 طلبات بدون فاتورة", len(active_df[active_df["case_type"] == "orphan_salla"]))
    with col5:
        st.metric("🧾 فواتير بدون طلب", len(active_df[active_df["case_type"] == "orphan_abc"]))
    with col6:
        st.metric("⏰ فواتير بعد آخر طلب", len(active_df[active_df["case_type"] == "post_cutoff_abc"]))
    with col7:
        st.metric("✅ تم إنجازها", len(df[df["status"] == "تم"]))
    
    col1, col2 = st.columns([1, 6])
    with col1:
        if st.button("🔄 تحديث الصفحة", use_container_width=True):
            st.rerun()

    # ========== الفلاتر المتقدمة وعزل الجلسات ==========
    sessions_list = get_all_sessions()
    
    # صف الفلاتر الأول: يحتوي على فلتر عزل الجلسات لمنع تكدس صفوف الملفات القديمة
    col_sess, col1, col2 = st.columns(3)
    with col_sess:
        if not sessions_list.empty:
            session_options = {"📋 كل الجلسات التاريخية": "الكل"}
            for _, s_row in sessions_list.iterrows():
                s_name = s_row.get('session_name', s_row['upload_batch_id'][:8]) or "جلسة غير مسماة"
                session_options[f"📂 {s_name} ({str(s_row['uploaded_at'])[:10]})"] = s_row['upload_batch_id']
            selected_session_label = st.selectbox("📂 فلتر الجلسة / الملف المرفوع", list(session_options.keys()))
            selected_session_id = session_options[selected_session_label]
        else:
            selected_session_id = "الكل"
            st.selectbox("📂 فلتر الجلسة / الملف المرفوع", ["لا توجد جلسات مرفوعة"])
            
    with col1:
        branch_options = ["الكل"] + sorted(df["pharmacy_name"].dropna().astype(str).unique().tolist())
        selected_branch = st.selectbox("🏥 فلتر الفرع", branch_options)
    with col2:
        status_filter = st.selectbox("📌 فلتر حالة الإجراء", ["الكل", "قيد المتابعة", "تم"])

    # صف الفلاتر الثاني
    col3, col4, col5, col6 = st.columns(4)
    with col3:
        order_status_options = ["الكل", "تم التوصيل", "ملغي", "مسترجع", "بانتظار الدفع", "تم الاستلام من فرع"]
        selected_order_status = st.selectbox("📋 فلتر حالة الطلب", order_status_options)
    with col4:
        search_order = st.text_input("🔢 رقم الطلب", placeholder="بحث برقم الطلب...")
    with col5:
        search_invoice = st.text_input("🧾 رقم الفاتورة", placeholder="بحث برقم الفاتورة...")
    with col6:
        search_sku = st.text_input("🏷️ SKU", placeholder="بحث بـ SKU...")

    # تطبيق الفلاتر ديناميكياً (تم حذق السطر المتكرر المسبب للمشكلة التاريخية)
    filtered_df = df.copy()
    if selected_session_id != "الكل":
        filtered_df = filtered_df[filtered_df["upload_batch_id"] == selected_session_id]

    if selected_branch != "الكل":
        filtered_df = filtered_df[filtered_df["pharmacy_name"] == selected_branch]
    if status_filter != "الكل":
        filtered_df = filtered_df[filtered_df["status"] == ("تم" if status_filter == "تم" else "قيد المتابعة")]
    if selected_order_status != "الكل":
        if selected_order_status == "تم الاستلام من فرع":
            filtered_df = filtered_df[filtered_df["order_status"].str.contains("تم الاستلام من فرع", na=False)]
        else:
            filtered_df = filtered_df[filtered_df["order_status"] == selected_order_status]
    if search_order:
        filtered_df = filtered_df[filtered_df["order_number"].astype(str).str.contains(search_order, na=False)]
    if search_invoice:
        filtered_df = filtered_df[filtered_df["invoice_number"].astype(str).str.contains(search_invoice, na=False)]
    if search_sku:
        filtered_df = filtered_df[filtered_df["sku"].astype(str).str.contains(search_sku, na=False)]

    active_mask_filtered = ~filtered_df["order_status"].apply(is_cancelled_or_returned_status)
    payment_mask = filtered_df["order_status"].apply(is_pending_payment_status)
    cancelled_mask = filtered_df["order_status"].apply(is_cancelled_or_returned_status)
    
    completed_df = get_completed_items()
    if selected_branch != "الكل":
        completed_df = completed_df[completed_df["pharmacy_name"] == selected_branch]
    
    # ========== العناصر القديمة ==========
    selected_branch_name = None if selected_branch == "الكل" else selected_branch
    
    old_orders_data = get_old_orders(pharmacy_name=selected_branch_name, months=6)
    old_invoices_data = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
    
    old_order_numbers = set(old_orders_data['order_number'].astype(str).tolist()) if not old_orders_data.empty else set()
    old_invoice_numbers = set(old_invoices_data['invoice_number'].astype(str).tolist()) if not old_invoices_data.empty else set()
    
    def exclude_old_items(temp_df, exclude_orders=True, exclude_invoices=True):
        if temp_df.empty:
            return temp_df
        result_temp = temp_df.copy()
        if exclude_orders and 'order_number' in result_temp.columns:
            result_temp = result_temp[~result_temp['order_number'].astype(str).isin(old_order_numbers)]
        if exclude_invoices and 'invoice_number' in result_temp.columns:
            result_temp = result_temp[~result_temp['invoice_number'].astype(str).isin(old_invoice_numbers)]
        return result_temp
    
    # ========== حساب الإحصائيات الصحيحة للتبويبات بناءً على الشروط الدقيقة والمحدثة ==========
    active_mask_filtered = filtered_df["active"] == 1 if "active" in filtered_df.columns else True
    is_cancelled_returned = filtered_df["order_status"].astype(str).str.strip().str.contains("ملغي|مسترجع|cancelled|returned|refunded", na=False, case=False)
    is_pending_payment = filtered_df["order_status"].astype(str).str.strip().str.contains("بانتظار الدفع|لم يتم الدفع|pending|unpaid", na=False, case=False)
    
    # 1️⃣ الإضافات والطلبات المفقودة
    additions_filtered = filtered_df[filtered_df["case_type"].isin(["addition", "orphan_salla"]) & (~is_cancelled_returned) & (~is_pending_payment) & active_mask_filtered].copy()
    additions_filtered = exclude_old_items(additions_filtered)
    total_additions_count = len(additions_filtered)
    completed_additions_count = len(additions_filtered[additions_filtered["status"] == "تم"]) if "status" in additions_filtered.columns else 0
    
    # 2️⃣ الإرجاعات والزيادات
    returns_filtered = filtered_df[filtered_df["case_type"].isin(["return", "orphan_abc"]) & (~is_cancelled_returned) & (~is_pending_payment) & active_mask_filtered].copy()
    returns_filtered = exclude_old_items(returns_filtered)
    total_returns_count = len(returns_filtered)
    completed_returns_count = len(returns_filtered[returns_filtered["status"] == "تم"]) if "status" in returns_filtered.columns else 0
    
    # 3️⃣ فواتير معلقة بين الفروع (الدمج الذكي النشط + الأرشيف المكرر)
    conflicts_filtered = filtered_df[filtered_df["case_type"] == "branch_conflict"].copy()
    conflicts_filtered = exclude_old_items(conflicts_filtered)
    
    old_orders_df = get_old_orders(months=6)
    if not old_orders_df.empty:
        old_orders_filtered_for_tab = old_orders_df.copy()
        if selected_branch != "الكل":
            old_orders_filtered_for_tab = old_orders_filtered_for_tab[old_orders_filtered_for_tab["pharmacy_name"] == selected_branch]
        
        old_conflicts = old_orders_filtered_for_tab[
            (old_orders_filtered_for_tab["case_type"] == "branch_conflict") | 
            (old_orders_filtered_for_tab["order_number"].isin(conflicts_filtered["order_number"]))
        ].copy()
        
        if not old_conflicts.empty:
            conflicts_filtered = pd.concat([conflicts_filtered, old_conflicts], ignore_index=True)
            if 'item_key' in conflicts_filtered.columns:
                conflicts_filtered = conflicts_filtered.drop_duplicates(subset=['item_key'])
                
    total_conflicts = len(conflicts_filtered)
    completed_conflicts = len(conflicts_filtered[conflicts_filtered["status"] == "تم"]) if "status" in conflicts_filtered.columns else 0
                
    # 4️⃣ فواتير بعد آخر طلب
    post_cutoff_filtered = filtered_df[(filtered_df["case_type"] == "post_cutoff_abc") & (~is_cancelled_returned) & active_mask_filtered].copy()
    post_cutoff_filtered = exclude_old_items(post_cutoff_filtered)
    total_post_cutoff = len(post_cutoff_filtered)
    completed_post_cutoff = len(post_cutoff_filtered[post_cutoff_filtered["status"] == "تم"]) if "status" in post_cutoff_filtered.columns else 0
    
    # 5️⃣ بانتظار الدفع
    payment_filtered = filtered_df[is_pending_payment & (filtered_df["case_type"] != "branch_conflict") & active_mask_filtered].copy()
    payment_filtered = exclude_old_items(payment_filtered)
    total_payment = len(payment_filtered)
    
    # 6️⃣ ملغي ومسترجع
    cancelled_filtered = filtered_df[is_cancelled_returned & (filtered_df["case_type"] != "branch_conflict")].copy()
    cancelled_filtered = exclude_old_items(cancelled_filtered)
    total_cancelled = len(cancelled_filtered)
    
    # 7️⃣ المكتملة والأرشيف
    old_invoices_df = get_old_invoices(months=6)
    completed_df_admin = get_completed_items()
    total_completed = len(completed_df_admin) if completed_df_admin is not None else 0

    # =========================================================================
    # 📥 زر تصدير الإكسيل الموحد للإدارة المعرف بشكل قاطع وحصري
    # =========================================================================
    if st.button("📥 تصدير كل التقارير الحالية إلى ملف Excel موحد", key="admin_global_excel_export_btn"):
        excel_data = export_to_excel({
            " canالإضافات والطلبات المفقودة": additions_filtered,
            "الإرجاعات والفواتير المعلقة": returns_filtered,
            "فواتير معلقة بين الفروع": conflicts_filtered,
            "فواتير بعد آخر طلب": post_cutoff_filtered,
            "بانتظار الدفع": payment_filtered,
            "الملغيات والمسترجعات": cancelled_filtered,
            "الطلبات القديمة التاريخية": old_orders_df,
            "الفواتير القديمة التاريخية": old_invoices_df
        })
        st.download_button(
            label="💾 تحميل ملف Excel الموحد",
            data=excel_data,
            type="primary",
            file_name=f"Balsam_Admin_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="admin_excel_download_stream"
        )

    
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] button:nth-child(1) { background-color: #4472C4; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(2) { background-color: #ED7D31; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(3) { background-color: #9B59B6; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(4) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(5) { background-color: #3498DB; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(6) { background-color: #E74C3C; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(7) { background-color: #27AE60; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(8) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(9) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(10) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] { transform: translateY(-2px) !important; box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important; }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="false"] { opacity: 0.85 !important; }
    .stTabs [data-baseweb="tab-list"] button:hover { transform: translateY(-2px) !important; opacity: 1 !important; }
    </style>
    """, unsafe_allow_html=True)
    
    # ========== عناوين التبويبات بصفحة الإدارة (مع توحيد ومطابقة مسميات الفلاتر الإدارية الصحيحة) ==========
    # تم تغيير المتغيرات لتقرأ من جداول الفلترة العلوية للإدارة مباشرة منعا للـ NameError
    total_additions_count = len(additions_filtered)
    completed_additions_count = len(additions_filtered[additions_filtered["status"] == "تم"]) if "status" in additions_filtered.columns else 0
    
    total_returns_count = len(returns_filtered)
    completed_returns_count = len(returns_filtered[returns_filtered["status"] == "تم"]) if "status" in returns_filtered.columns else 0
    
    # =========================================================================
    # 📊 حساب إحصائيات العناصر القديمة وتأمينها لمنع الـ NameError
    # =========================================================================
    # جلب إحصائيات الأعداد من قاعدة البيانات مباشرة لتوفير المتغيرات المطلوبة في التبويبات
    try:
        old_orders_stats_data = get_old_orders_stats()
        old_invoices_stats_data = get_old_invoices_stats()
        
        # فلترة الإحصائيات التاريخية بناءً على الفرع المحدد لتطابق الأرقام مع فلاتر الإدارة
        if selected_branch != "الكل":
            old_orders_stats_data = [r for r in old_orders_stats_data if r[0] == selected_branch]
            old_invoices_stats_data = [r for r in old_invoices_stats_data if r[0] == selected_branch]
            
        total_old_orders_stats = sum(r[1] for r in old_orders_stats_data)
        total_old_invoices_stats = sum(r[1] for r in old_invoices_stats_data)
    except Exception as e:
        total_old_orders_stats = len(old_orders_df) if 'old_orders_df' in locals() else 0
        total_old_invoices_stats = len(old_invoices_df) if 'old_invoices_df' in locals() else 0

    # =========================================================================
    # 📊 بناء التبويبات (Tabs) بشكل متزن ومستقر عددياً ومطابق 100% لمتغيراتك
    # =========================================================================
    tab_additions, tab_returns, tab_conflicts, tab_post_cutoff, tab_payment, tab_cancelled, tab_completed, tab_old_orders, tab_old_invoices, tab_old_stats = st.tabs([
        f"📥 الإضافات والطلبات المفقودة ({completed_additions_count}/{total_additions_count})" if total_additions_count > 0 else "📥 الإضافات والطلبات المفقودة (0)",
        f"📤 الإرجاعات والفواتير المعلقة ({completed_returns_count}/{total_returns_count})" if total_returns_count > 0 else "📤 الإرجاعات والفواتير المعلقة (0)",
        f"📊 فواتير معلقة بين الفروع ({completed_conflicts}/{total_conflicts})" if total_conflicts > 0 else f"📊 فواتير معلقة بين الفروع ({total_conflicts})", 
        f"⏰ فواتير بعد آخر طلب ({completed_post_cutoff}/{total_post_cutoff})" if total_post_cutoff > 0 else f"⏰ فواتير بعد آخر طلب ({total_post_cutoff})",
        f"💰 بانتظار الدفع ({total_payment})",
        f"⚠️ ملغي/مسترجع ({total_cancelled})",
        f"✅ تم الانتهاء ({total_completed})",
        f"📅 طلبات قديمة ({total_old_orders_stats})",   # 💡 تم التأمين والحساب بنجاح
        f"🧾 فواتير قديمة ({total_old_invoices_stats})", # 💡 تم التأمين والحساب بنجاح
        "📊 إحصائيات قديمة"
    ])

       
# =========================================================================
    # 📺 [تم الإصلاح]: ربط تبويبات العرض بالمتغيرات الفردية المنقاة للأعلى
    # =========================================================================
    
    with tab_additions:
        st.markdown(f"### 📥 الإضافات والطلبات المفقودة المستخرجة")
        # تم تغيير المتغير لـ additions_filtered لإنهاء الـ NameError
        if not additions_filtered.empty:
            render_table_with_click(additions_filtered, "addition", allow_move=True)
        else:
            st.success("🎉 لا توجد حالات إضافة أو طلبات مفقودة قيد المتابعة حالياً.")

    with tab_returns:
        st.markdown(f"### 📤 الإرجاعات والزيادات المستندة للفروع")
        # تم تغيير المتغير لـ returns_filtered لإنهاء الـ NameError
        if not returns_filtered.empty:
            render_table_with_click(returns_filtered, "return", allow_move=True)
        else:
            st.success("🎉 ممتاز! لا توجد فواتير زيادة أو قيد إرجاع معلق.")

    with tab_conflicts:
        st.markdown(f"### 📊 فواتير معلقة بسبب التداخل والتكرار بين الفروع ({total_conflicts})")
        # الجدول المدمج والمحدث للنزاعات النشطة والقديمة المتداخلة
        if not conflicts_filtered.empty:
            render_table_with_click(conflicts_filtered, "branch_conflict", allow_move=True)
        else:
            st.success("🎉 ممتاز! لا توجد فواتير معلقة أو متداخلة بين الفروع.")
            
    with tab_post_cutoff:
        st.markdown(f"### ⏰ فواتير ABC تم إنشاؤها بعد توقيت آخر طلب في سلة")
        # تم تغيير المتغير لـ post_cutoff_filtered
        if not post_cutoff_filtered.empty:
            render_table_with_click(post_cutoff_filtered, "post_cutoff", allow_move=True)
        else:
            st.success("🎉 لا توجد فواتير مخرجات مسجلة بعد توقيت قطع الجلسة.")

    with tab_payment:
        st.markdown(f"### 💰 طلبات سلة معلقة بانتظار إتمام الدفع بالفرع")
        # تم تغيير المتغير لـ payment_filtered
        if not payment_filtered.empty:
            render_table_with_click(payment_filtered, "pending_payment", allow_move=True)
        else:
            st.success("🎉 لا توجد طلبات معلقة بانتظار الدفع.")

    with tab_cancelled:
        st.markdown(f"### ⚠️ فواتير ABC التابعة لطلبات ملغية أو مسترجعة في سلة")
        # تم تغيير المتغير لـ cancelled_filtered
        if not cancelled_filtered.empty:
            render_table_with_click(cancelled_filtered, "cancelled_returned", allow_move=True)
        else:
            st.success("🎉 نظيف! لا توجد فواتير مضروبة لطلبات ملغية.")

    with tab_completed:
        st.markdown("### ✅ التسويات والطلبات التي تم الانتهاء منها وإغلاقها")
        if completed_df_admin is not None and not completed_df_admin.empty:
            render_completed_table(completed_df_admin, is_admin=True)
        else:
            st.info("📭 لم يتم اعتماد أو إكمال أي حالات في هذه الجلسة بعد.")

    with tab_old_orders:
        st.markdown("### 📋 أرشيف الطلبات القديمة التاريخية (أكثر من 6 أشهر)")
        if not old_orders_df.empty:
            # عرض الأرشيف التاريخي الكامل للطلبات القديمة
            st.dataframe(old_orders_df, use_container_width=True)
        else:
            st.success("🎉 لا توجد طلبات قديمة مؤرشفة.")

    with tab_old_invoices:
        st.markdown("### 🧾 أرشيف الفواتير القديمة التاريخية (أكثر من 6 أشهر)")
        if not old_invoices_df.empty:
            # عرض الأرشيف التاريخي الكامل للفواتير القديمة
            st.dataframe(old_invoices_df, use_container_width=True)
        else:
            st.success("🎉 لا توجد فواتير قديمة مؤرشفة.")
   
    with tab_old_stats:
        st.markdown("### 📊 إحصائيات العناصر القديمة")
        st.markdown("---")
        
        if selected_branch != "الكل":
            st.info(f"🏥 عرض الإحصائيات للفرع: {selected_branch}")
        
        old_orders_stats_data = get_old_orders(pharmacy_name=selected_branch_name, months=6)
        old_invoices_stats_data = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 📅 الطلبات القديمة")
            st.markdown(f"""
            - **إجمالي الطلبات القديمة:** {len(old_orders_stats_data)}
            - **إضافات قديمة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "addition"])}
            - **إرجاعات قديمة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "return"])}
            - **طلبات بدون فاتورة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "orphan_salla"])}
            """)
            
            if selected_branch == "الكل" and not old_orders_stats_data.empty:
                st.markdown("#### 🏥 التوزيع حسب الفرع")
                for branch, count in old_orders_stats_data.groupby("pharmacy_name").size().items():
                    st.markdown(f"- {branch}: {count} طلب")
        
        with col2:
            st.markdown("#### 🧾 الفواتير القديمة")
            st.markdown(f"""
            - **إجمالي الفواتير القديمة:** {len(old_invoices_stats_data)}
            - **إضافات:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "addition"])}
            - **إرجاعات:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "return"])}
            - **فواتير بدون طلب:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "orphan_abc"])}
            """)
            
            if selected_branch == "الكل" and not old_invoices_stats_data.empty:
                st.markdown("#### 🏥 التوزيع حسب الفرع")
                for branch, count in old_invoices_stats_data.groupby("pharmacy_name").size().items():
                    st.markdown(f"- {branch}: {count} فاتورة")
        
        st.markdown("---")
        total_old = len(old_orders_stats_data) + len(old_invoices_stats_data)
        if total_old > 0:
            st.warning(f"⚠️ إجمالي العناصر القديمة (طلبات + فواتير): {total_old}")
            st.info("💡 هذه العناصر تم استبعادها تلقائياً من التبويبات الأخرى (الإضافات، الإرجاعات، إلخ)")
        else:
            st.success("🎉 لا توجد عناصر قديمة (طلبات أو فواتير)")
    
    st.markdown('<div class="section-title">👥 آخر دخول للصيدليات والفروع</div>', unsafe_allow_html=True)
    last_logins = get_all_last_logins()
    
    if not last_logins.empty:
        # 💡 [تصفية أمنية]: استبعاد حسابات الإدارة وعرض الفروع فقط في جدول المراقبة
        pharmacy_logins = last_logins[~last_logins['username'].isin(['manager'])]
        
        cols = st.columns(4)
        for idx, (_, row) in enumerate(pharmacy_logins.head(8).iterrows()):
            with cols[idx % 4]:
                st.markdown(f"""
                <div class="note-card">
                    <strong>🏥 {row['pharmacy_name']}</strong><br>
                    <span>👤 {row['pharmacist_name'] or 'غير مسجل'}</span><br>
                    <span>📅 {row['last_login'][:16] if row['last_login'] else 'لم يدخل'}</span><br>
                    <span>🌐 IP: {row['last_ip'] or 'غير معروف'}</span>
                </div>
                """, unsafe_allow_html=True)
        
        with st.expander("📋 عرض جميع الصيدليات"):
            st.dataframe(last_logins[['pharmacy_name', 'pharmacist_name', 'last_login', 'last_ip']], use_container_width=True)
    else:
        st.info("لا توجد سجلات دخول للصيدليات بعد")

def show_special_offers_page():
    """واجهة مستقلة تماماً تعرض على كامل مساحة الشاشة لإدارة العروض ورفع ملفات الـ Excel"""
    st.markdown("# 🎁 مركز إدارة العروض الترويجية الخاصة (Live Sync)")
    st.markdown("---")
    
    # جلب توكن الأمان النشط لمنصة سلة من الجلسة
    access_token = st.session_state.get('salla_access_token', 'SAMPLE_TOKEN')
    
    st.markdown("### 📊 رفع وتحديث العروض دفعة واحدة عبر ملف Excel")
    with st.container():
        st.info("📋 يجب أن يحتوي شيت الإكسيل على الأعمدة التالية بشكل قياسي: (name, message, offer_type, min_purchase_amount)")
        offers_file = st.file_uploader("اختر ملف Excel العروض الخاصة التراكمي:", type=["xlsx", "xls"], key="excel_offers_uploader")
        
        if offers_file is not None:
            if st.button("🚀 بدء معالجة الملف وبث العروض حياً إلى متجر سلة", use_container_width=True):
                from utils.api_connectors import process_special_offers_excel_sync
                success, msg = process_special_offers_excel_sync(offers_file, access_token)
                if success:
                    st.success(msg)
                    st.balloons()
                else:
                    st.error(msg)
                    
    st.markdown("---")
    st.markdown("### 📋 العروض الحالية النشطة بالمتجر (مراقبة وتعديل لحظي)")
    
    # جلب العروض الحية وعرضها مع أزرار التحكم الفردية (تفعيل / تعطيل)
    from utils.api_connectors import get_salla_special_offers, change_special_offer_status
    offers = get_salla_special_offers(access_token)
    
    if offers:
        for offer in offers:
            off_id = offer.get('id')
            current_status = offer.get('status', 'inactive')
            status_color = '#27ae60' if current_status == 'active' else '#e74c3c'
            
            st.markdown(f"""
            <div style="background:#f8f9fa; border-radius:12px; padding:1.2rem; margin-bottom:1rem; border-right:5px solid {status_color}; box-shadow:0 2px 4px rgba(0,0,0,0.02);">
                <h4 style="margin:0; color:#1f7a8c;">🎁 {offer.get('name')} ({'🟢 نشط حالياً' if current_status == 'active' else '🔴 معطل'})</h4>
                <p style="margin:0.5rem 0; color:#444;">💬 الرسالة التسويقية: {offer.get('message')}</p>
                <small style="color:#777;">🆔 معرف السيرفر: {off_id} | ⚙️ نوع العرض الحسابي: {offer.get('offer_type')}</small>
            </div>
            """, unsafe_allow_html=True)
            
            c1, c2 = st.columns([1, 5])
            with c1:
                if current_status == "active":
                    if st.button("🔴 إيقاف العرض", key=f"stop_{off_id}", use_container_width=True):
                        if change_special_offer_status(access_token, off_id, "inactive"):
                            st.toast("🔄 تم إيقاف وتعطيل العرض في المتجر!")
                            st.rerun()
                else:
                    if st.button("🟢 تفعيل الآن", key=f"start_{off_id}", use_container_width=True):
                        if change_special_offer_status(access_token, off_id, "active"):
                            st.toast("✅ تم إطلاق وتفعيل العرض حياً للعملاء!")
                            st.rerun()
            st.markdown("---")
    else:
        st.warning("📭 لا توجد عروض خاصة منشأة في متجرك حالياً.")
